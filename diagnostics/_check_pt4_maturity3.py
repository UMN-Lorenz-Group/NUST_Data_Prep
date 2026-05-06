"""Find PT-IV Maturity data for Portageville/Queenstown/Eldorado directly in XLSX."""
import openpyxl, datetime
from pathlib import Path

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "input_1980/1980/Sojabone-1980 (90-164 OR).xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

# Find all rows that mention Portageville or Queenstown in any cell
print("=== Rows containing 'Portageville' or 'Queenstown' or 'Eldorado' ===")
for row in ws.iter_rows():
    for cell in row:
        v = str(cell.value or "").lower()
        if any(x in v for x in ["portageville", "queenstown", "eldorado"]):
            vals = [ws.cell(row=cell.row, column=c).value for c in range(1, 15)]
            print(f"  Row {cell.row}: {vals}")
            break

print()

# Also: find rows where col A is 'tp8' OR contains tp8 in any cell (case insensitive, partial)
print("=== All cells containing 'tp8' (partial match) ===")
for row in ws.iter_rows():
    for cell in row:
        if "tp8" in str(cell.value or "").lower():
            vals = [ws.cell(row=cell.row, column=c).value for c in range(1, 8)]
            print(f"  Row {cell.row} Col {cell.column}: '{cell.value}' -> {vals}")
            break
