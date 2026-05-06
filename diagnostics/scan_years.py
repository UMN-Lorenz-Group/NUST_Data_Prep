import openpyxl
from pathlib import Path
import re
import sys

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

GREEN = Path("R:/NUST_Historical_Data/Green-20260427T181938Z-3-001/Green")

PER_LOC = {"tp6","tp7","tp8","tp9","tp10","tp11a","tp11b","tp12a","tp12b"}

for year_dir in sorted(GREEN.iterdir()):
    year_match = re.search(r'\b(19|20)\d{2}\b', year_dir.name)
    if not year_match:
        continue
    year = year_match.group(0)
    xlsx_files = sorted(f for f in year_dir.rglob("*.xlsx") if not f.name.startswith("~$"))
    if not xlsx_files:
        print(f"{year} | no xlsx")
        sys.stdout.flush()
        continue

    # Use first file (representative for structure check)
    xlsx = xlsx_files[0]
    try:
        wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
        ws = wb.active
        rows_count = ws.max_row
        col_a_tp = set()
        trait_labels = set()
        loc_samples = set()

        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
            v = str(row[0]).strip() if row[0] else ""
            if v.startswith("tp"):
                col_a_tp.add(v)
                # trait label is col B on tp6-tp12b rows
                if v in PER_LOC and len(row) > 1 and row[1]:
                    trait_labels.add(str(row[1]).strip())
            # Collect location-like values from header rows (non-tp, non-empty col C+)
            # Heuristic: rows where col A looks like a state/region abbreviation pattern
            if not v and len(row) > 2:
                for cell in row[2:8]:
                    if cell and isinstance(cell, str) and len(cell) < 25 and "." in str(cell):
                        loc_samples.add(str(cell).strip())

        wb.close()
        tp_sorted = sorted(col_a_tp)
        tl_sorted = sorted(trait_labels)
        loc_sample_list = sorted(loc_samples)[:8]
        print(f"{year} | rows={rows_count} | tp={tp_sorted} | traits={tl_sorted} | loc_samples={loc_sample_list}")
    except Exception as e:
        print(f"{year} | ERROR: {e}")
    sys.stdout.flush()
