"""Read raw XLSX rows around PT-IV maturity blocks at rows 1530 and 1621.
Compute expected DOY and compare to CSV values.
"""
import openpyxl, datetime
from pathlib import Path
import pandas as pd

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "input_1980/1980/Sojabone-1980 (90-164 OR).xlsx"
CSV  = ROOT / "output_1980/validated/combined_1980_phenotypesTable_approved.csv"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

def doy_1980(dt):
    """Convert openpyxl datetime to 1980 DOY (uses month/day only)."""
    if isinstance(dt, (datetime.datetime, datetime.date)):
        return datetime.date(1980, dt.month, dt.day).timetuple().tm_yday
    return None

def read_block(start_row, label):
    print(f"\n=== Block at row {start_row} ({label}) ===")
    # Print marker row(s) just before
    for r in range(max(1, start_row - 3), start_row):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 8)]
        print(f"  Row {r} (pre-header): {vals}")

    # Header
    cols = [ws.cell(row=start_row, column=c).value for c in range(1, 14)]
    print(f"  Header: {cols}")

    # Build col index: location name -> column number
    loc_cols = {}
    for c in range(2, 14):
        h = str(ws.cell(row=start_row, column=c).value or "")
        if h:
            loc_cols[c] = h

    # Reference row (first data row)
    ref_row = start_row + 1
    ref_strain = ws.cell(row=ref_row, column=1).value
    print(f"  Reference strain: {ref_strain}")
    ref_doys = {}
    for c, name in loc_cols.items():
        v = ws.cell(row=ref_row, column=c).value
        doy = doy_1980(v)
        if doy:
            ref_doys[c] = (name, doy, v)
            print(f"    {name}: {v} → DOY {doy}")
        else:
            ref_doys[c] = (name, None, v)
            print(f"    {name}: raw={v}")

    print()
    print("  Strain DOY calculations (ref_DOY + offset):")
    for r in range(ref_row + 1, start_row + 55):
        strain = ws.cell(row=r, column=1).value
        if strain is None:
            break
        if str(strain).startswith("Date") or str(strain).startswith("*Day"):
            break
        row_data = []
        for c, (name, ref_doy, _) in ref_doys.items():
            offset = ws.cell(row=r, column=c).value
            if ref_doy is not None and isinstance(offset, (int, float)):
                computed = ref_doy + int(offset)
                row_data.append(f"{name.split('.')[-1].strip()[:12]}:{computed}(off={offset:+d})")
            elif isinstance(offset, (datetime.datetime, datetime.date)):
                computed = doy_1980(offset)
                row_data.append(f"{name.split('.')[-1].strip()[:12]}:{computed}(date)")
            else:
                row_data.append(f"{name.split('.')[-1].strip()[:12]}:None")
        print(f"    {str(strain):25s}  {' | '.join(row_data)}")

    return ref_doys

read_block(1530, "PT-IV maturity block A")
read_block(1621, "PT-IV maturity block B")

# Now compare computed vs CSV
print("\n\n=== Spot-check computed DOY vs CSV for flagged strains ===")
df = pd.read_csv(CSV)
pt4_mat = df[(df["Test"] == "PT-IV") & (df["Phenotype"] == "Maturity")]

checks = [
    ("Union (IV)",    "Queenstown",        "MD"),
    ("K1062",         "Portageville Loam", "MO"),
    ("K1063",         "Portageville Loam", "MO"),
    ("K1033 Douglas", "Portageville Loam", "MO"),
    ("LS78-229",      "Queenstown",        "MD"),
    ("LS78-344",      "Eldorado",          "IL"),
]
for strain, city, state in checks:
    row = pt4_mat[(pt4_mat["Strain"] == strain) & (pt4_mat["City"] == city)]
    csv_val = row["Value"].values[0] if len(row) else "NOT FOUND"
    print(f"  CSV: {strain:25s} | {city:22s} {state} | Maturity={csv_val}")
