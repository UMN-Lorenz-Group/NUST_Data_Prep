"""Check all PT-IV tp8 blocks for Portageville/Queenstown/Eldorado locations."""
import openpyxl
from pathlib import Path
from datetime import date

ROOT = Path(__file__).parent.parent
XLSX = ROOT / "input_1980/1980/Sojabone-1980 (90-164 OR).xlsx"

wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

# All tp8 marker rows
tp8_rows = []
for row in ws.iter_rows():
    for cell in row:
        if str(cell.value).strip().lower() == "tp8":
            tp8_rows.append(cell.row)

print(f"All tp8 rows: {tp8_rows}")
print()

# Print each tp8 block header to identify which has our locations
for start in tp8_rows:
    header_row = ws.cell(row=start + 1, column=1).value
    cols = [ws.cell(row=start + 1, column=c).value for c in range(1, 12)]
    ref_row = [ws.cell(row=start + 2, column=c).value for c in range(1, 12)]
    print(f"Row {start}: {cols}")
    print(f"  Reference: {ref_row}")
    print()

# Now find the block(s) with Portageville/Queenstown/Eldorado
TARGET_LOCS = ["portageville", "queenstown", "eldorado", "lexington"]

for start in tp8_rows:
    header = [str(ws.cell(row=start + 1, column=c).value or "").lower() for c in range(1, 15)]
    if any(loc in " ".join(header) for loc in TARGET_LOCS):
        print(f"\n=== PT-IV tp8 block at row {start} (has target locations) ===")
        # Print column headers
        cols = [ws.cell(row=start + 1, column=c).value for c in range(1, 12)]
        print(f"Headers: {cols}")
        # Print reference check row
        ref = [ws.cell(row=start + 2, column=c).value for c in range(1, 12)]
        print(f"Ref check: {ref}")
        print()

        # Find column indices for our locations
        loc_map = {}
        for c in range(1, 15):
            h = str(ws.cell(row=start + 1, column=c).value or "").lower()
            for loc in TARGET_LOCS:
                if loc in h:
                    loc_map[loc] = c

        # Print reference DOY
        ref_strain = ws.cell(row=start + 2, column=1).value
        print(f"Reference strain: {ref_strain}")
        for loc, col in sorted(loc_map.items()):
            val = ws.cell(row=start + 2, column=col).value
            if isinstance(val, date):
                doy = val.timetuple().tm_yday
                # Adjust for 1980 leap year (Excel dates stored as 1980 but read as 2026)
                # The day-of-year is what matters; month/day is the same regardless of year
                # For leap year 1980: Jan=31, Feb=29, so add 1 for dates after Feb 28
                import datetime
                d_1980 = datetime.date(1980, val.month, val.day)
                doy_1980 = d_1980.timetuple().tm_yday
                print(f"  {loc} col {col}: {val.month}/{val.day} = DOY {doy_1980} (1980 leap)")
            else:
                print(f"  {loc} col {col}: raw={val}")

        print()
        print("  Data rows (strain, target-location values):")
        for r in range(start + 2, start + 50):
            if ws.cell(row=r, column=1).value is None:
                break
            strain = ws.cell(row=r, column=1).value
            row_vals = {}
            for loc, col in loc_map.items():
                row_vals[loc] = ws.cell(row=r, column=col).value
            print(f"    {str(strain):25s} {row_vals}")
