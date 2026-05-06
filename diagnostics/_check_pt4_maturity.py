"""Cross-check PT-IV Maturity values in approved CSV vs raw XLSX cells."""
import sys
from pathlib import Path
import pandas as pd
import openpyxl

ROOT = Path(__file__).parent.parent

XLSX = ROOT / "input_1980/1980/Sojabone-1980 (90-164 OR).xlsx"
CSV  = ROOT / "output_1980/validated/combined_1980_phenotypesTable_approved.csv"

# Strains/locations to spot-check
CHECK = [
    # (strain, city, state)
    ("Union (IV)",  "Queenstown", "MD"),
    ("LS78-229",    "Queenstown", "MD"),
    ("LS78-335",    "Queenstown", "MD"),
    ("K1062",       "Portageville Loam", "MO"),
    ("K1063",       "Portageville Loam", "MO"),
    ("K1033 Douglas", "Portageville Loam", "MO"),
    ("LS78-229",    "Eldorado",   "IL"),
    ("LS78-344",    "Eldorado",   "IL"),
]

print("=== CSV values (PT-IV Maturity) ===")
df = pd.read_csv(CSV)
pt4_mat = df[(df["Test"] == "PT-IV") & (df["Phenotype"] == "Maturity")]
for strain, city, state in CHECK:
    row = pt4_mat[(pt4_mat["Strain"] == strain) & (pt4_mat["City"] == city)]
    val = row["Value"].values[0] if len(row) else "NOT FOUND"
    print(f"  {strain:20s} | {city:22s} {state} | {val}")

print()
print("=== Raw XLSX scan (searching for Maturity/tp8 block near PT-IV) ===")
wb = openpyxl.load_workbook(XLSX, data_only=True)
ws = wb.active

# Find tp8 markers (Maturity table) — scan all rows for 'tp8'
tp8_rows = []
for row in ws.iter_rows():
    for cell in row:
        if str(cell.value).strip().lower() == "tp8":
            tp8_rows.append(cell.row)

print(f"  tp8 markers at rows: {tp8_rows}")

# For the last tp8 block (PT-IV is the last group in file 2), print header + data rows
if tp8_rows:
    target_row = tp8_rows[-1]
    print(f"\n  Last tp8 block starting at row {target_row}:")
    print(f"  {'Row':>5}  {'Col A (strain)':25}  cols B-Z (values)")
    for r in range(target_row, min(target_row + 45, ws.max_row + 1)):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 20)]
        # Skip fully empty rows
        if all(v is None for v in vals):
            continue
        row_str = "  ".join(str(v) if v is not None else "-" for v in vals[:15])
        print(f"  {r:5d}  {row_str}")
