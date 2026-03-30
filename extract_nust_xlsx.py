#!/usr/bin/env python
"""
extract_nust_xlsx.py
====================
Extract all NUST trial data from historical XLSX reports using Claude API.
Claude receives the raw cell grid and handles all table detection, location
normalization, and schema mapping — no rule-based parsing.

Usage:
    python extract_nust_xlsx.py --file "Sojabone-1980 (1-89 OR).xlsx" --out_dir ./output/
    python extract_nust_xlsx.py --dir ./NUST_Historical_Data/ --out_dir ./output/

Requirements:
    pip install openpyxl anthropic pandas

Environment:
    ANTHROPIC_API_KEY must be set, or pass --api_key
"""

import argparse
import json
import os
import re
import sys
import time
from pathlib import Path

# Force UTF-8 output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import anthropic
import pandas as pd


# ---------------------------------------------------------------------------
# System prompt — describes NUST schema and output format to Claude
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are extracting structured data from a historical NUST (North American Uniform Soybean Trial) spreadsheet section.

The cell grid is formatted as: R{row_number}<TAB>col1<TAB>col2<TAB>...
Empty cells are empty strings. Row numbers are included for reference.

The section you receive covers either a global parentage table or ONE entry group from the trial.
Entry groups contain several sub-tables tagged by tp* codes in column A:

SUB-TABLE TYPES:
- tp1: Global parentage (Strain, Parentage or Source) — 2 columns only
- tp2: Entry group parentage (Strain, Parentage, Previous Testing, Generation Composited)
- tp3a: Descriptive/morphological (Strain, Descriptive Code, Chlorosis Score Ames, Chlorosis Score Lamberton, Hypocotyl Score Ames, Shattering)
- tp3b: Disease resistance (Strain, BSR %, GERM %, SMV score, PSB %, PS score, PR Reaction)
- tp4: Summary performance current year avg (Strain, Yield bu/a, Rank No., Maturity Date, Lodging Score, Height In., Seed Quality Score, Seed Size g/100, Composition Protein %, Composition Oil %)
- tp5: Multi-year means — SKIP ENTIRELY, do not extract
- tp6: Per-location YIELD (bu/a)
- tp7: Per-location Yield Rank
- tp8: Per-location MATURITY (date)
- tp9: Per-location LODGING (score)
- tp10: Per-location HEIGHT (inches)
- tp11a: Per-location SEED QUALITY (score)
- tp11b: Per-location SEED SIZE (g/100)
- tp12a: Per-location % PROTEIN
- tp12b: Per-location % OIL

PER-LOCATION TABLE STRUCTURE (tp6–tp12b):
- Row with tp* tag: section marker (col A = tp code, col B = trait name label)
- Next non-empty row: column headers — col A = "Strain", col B = "N Mean Tests" or similar, cols C+ = location names
- Data rows: col A = strain name, col B = mean value, cols C+ = per-location values
- FOOTER rows (do NOT treat as strain data, capture as metadata):
  "No. of Tests", "C.V. (%)", "L.S.D. (5%)", "Row sp (in.)", "Rows/plot", "Reps"

LOCATION NORMALIZATION — convert abbreviations to City_State format:
- "Ont. Ottawa" → city="Ottawa", state="ONT"
- "Ont. Elora" → city="Elora", state="ONT"
- "Wisc. Ashland" → city="Ashland", state="WI"
- "N.D. Fargo" → city="Fargo", state="ND"
- "Man. Morden" → city="Morden", state="MAN"
- "Man. Brandon" or "Man. Brandon*" → city="Brandon", state="MAN"
- "Minn. Rosemount" → city="Rosemount", state="MN"
- "Minn. Morris" or "Minn. Norris" → city="Morris", state="MN"
- For any unrecognized location: keep as city, state=""

TRAIT NORMALIZATION:
- "YIELD (bu/a)" or "YIELD" → "YIELD (bu/a)"
- "Yield Rank" or "YIELD RANK" → "YIELD RANK"
- "MATURITY (date)" or "MATURITY DATE" → "MATURITY (date)"
- "LODGING (score)" or "Lodging" or "LODGING" → "LODGING (score)"
- "HEIGHT (inches)" or "HEIGHT" → "PLANT HEIGHT (inches)"
- "QUALITY (score)" or "Quality (score)" or "QUALITY" → "SEED QUALITY (score)"
- "SIZE (g/100)" or "SEED SIZE" or "SIZE" → "SEED SIZE (g/100)"
- "% PROTEIN" or "PROTEIN (%)" or "PROTEIN" → "PROTEIN (%)"
- "% OIL" or "OIL (%)" or "Oil %" or "OIL" → "OIL (%)"

Return ONLY valid JSON (no markdown fences) with this exact structure:
{
  "entry_group": "<group label passed to you, e.g. Group_1>",
  "parentage": [
    {"strain": "...", "parentage": "...", "prev_testing": "...", "generation": "..."}
  ],
  "descriptive": [
    {"strain": "...", "descriptive_code": "...", "chlorosis_ames": "...", "chlorosis_lamberton": "...", "hypocotyl_ames": "...", "shattering": "..."}
  ],
  "disease": [
    {"strain": "...", "BSR_pct": "...", "GERM_pct": "...", "SMV_score": "...", "PSB_pct": "...", "PS_score": "...", "PR_reaction": "..."}
  ],
  "summary": [
    {"strain": "...", "yield_bua": "...", "rank": "...", "maturity": "...", "lodging": "...", "height_in": "...", "quality": "...", "seed_size_g100": "...", "protein_pct": "...", "oil_pct": "..."}
  ],
  "phenotypes": [
    {
      "trait": "YIELD (bu/a)",
      "data": [
        {"strain": "...", "mean": "...", "locations": {"Ottawa_ONT": "...", "Elora_ONT": "..."}}
      ],
      "metadata": {"no_of_tests": {}, "cv_pct": {}, "lsd_5pct": {}, "row_spacing_in": {}, "rows_per_plot": {}, "reps": {}}
    }
  ]
}

IMPORTANT NOTES:
- Include ALL strain data rows found (do not truncate)
- Use null for missing, dash, or blank values
- Preserve original value formatting (e.g. "9-5*", "+4.5 *", "3M", "5S")
- If a section is absent from the grid, return an empty list for that key
- tp5 (multi-year means) must always be skipped entirely
- The "summary" key captures tp4 data (current-year averages only, not per-location)
- location keys in the "locations" dict use City_State format: "Ottawa_ONT", "Morris_MN"
"""


# ---------------------------------------------------------------------------
# Cell grid utilities
# ---------------------------------------------------------------------------

def extract_year_from_filename(filename: str) -> str:
    """Extract 4-digit year from filename like 'Sojabone-1980 (1-89 OR).xlsx'."""
    match = re.search(r'\b(19|20)\d{2}\b', filename)
    return match.group(0) if match else "unknown"


def sheet_to_text(ws, min_row: int = 1, max_row: int = None) -> str:
    """
    Serialize worksheet rows as a tab-separated text grid with row numbers.
    Only includes non-empty rows; trims trailing empty columns per row.
    """
    if max_row is None:
        max_row = ws.max_row
    lines = []
    for i, row in enumerate(
        ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row
    ):
        vals = [str(v) if v is not None else "" for v in row]
        while vals and vals[-1] == "":
            vals.pop()
        if any(v.strip() for v in vals):
            lines.append(f"R{i}\t" + "\t".join(vals))
    return "\n".join(lines)


PER_LOC_MARKERS = ("tp6", "tp7", "tp8", "tp9", "tp10", "tp11a", "tp11b", "tp12a", "tp12b")
ALL_TP_MARKERS = ("tp2", "tp3a", "tp3b", "tp4", "tp5") + PER_LOC_MARKERS


def find_group_boundaries(ws) -> list[tuple[str, int, int]]:
    """
    Locate entry group boundaries by finding tp2 markers (start of each group).
    Also captures the global tp1 parentage section before the first tp2.

    Chunking strategy to stay within reliable API output limits (~150 rows/call):
    - Groups <=300 rows: sent as one chunk.
    - Groups >300 rows: split into:
        a) rows tp2..tp6-1 (parentage, descriptive, disease, summary)
        b) per-location phenotypes, further split per-trait if b >150 rows
           (each tp6/tp7/.../tp12b table becomes its own chunk)
    Falls back to treating the whole sheet as one group if no tp2 found.
    Returns list of (group_label, start_row, end_row).
    """
    # Collect all tp marker row positions in one pass
    tp2_rows = []
    tp_marker_rows: dict[int, str] = {}  # row -> marker value
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
    ):
        v = row[0]
        if v:
            sv = str(v).strip()
            if sv == "tp2":
                tp2_rows.append(i)
            if sv in ALL_TP_MARKERS:
                tp_marker_rows[i] = sv

    if not tp2_rows:
        print("  Warning: no tp2 markers found -- sending entire sheet as one group")
        return [("Group_1", 1, ws.max_row)]

    boundaries = []

    # Global tp1 parentage section (rows before first tp2)
    if tp2_rows[0] > 1:
        boundaries.append(("GlobalParentage", 1, tp2_rows[0] - 1))

    # Entry groups: tp2 -> next tp2 (or end of sheet)
    for idx, start in enumerate(tp2_rows):
        end = tp2_rows[idx + 1] - 1 if idx + 1 < len(tp2_rows) else ws.max_row
        group_label = f"Group_{idx + 1}"
        n_rows = end - start + 1

        if n_rows > 300:
            # Find the first tp6 within this group
            tp6_row = None
            for r in sorted(tp_marker_rows):
                if start < r <= end and tp_marker_rows[r] == "tp6":
                    tp6_row = r
                    break

            if tp6_row and tp6_row > start + 10:
                # Sub-chunk A: parentage + descriptive + disease + summary (before tp6)
                boundaries.append((f"{group_label}a", start, tp6_row - 1))

                # Sub-chunk B: per-location phenotypes
                b_rows = end - tp6_row + 1
                if b_rows > 150:
                    # Split into per-trait chunks (one chunk per tp6/tp7/.../tp12b marker)
                    per_loc_in_group = sorted(
                        r for r in tp_marker_rows
                        if tp6_row <= r <= end and tp_marker_rows[r] in PER_LOC_MARKERS
                    )
                    for t_idx, t_start in enumerate(per_loc_in_group):
                        t_end = per_loc_in_group[t_idx + 1] - 1 if t_idx + 1 < len(per_loc_in_group) else end
                        trait_tag = tp_marker_rows[t_start]
                        boundaries.append((f"{group_label}b_{trait_tag}", t_start, t_end))
                else:
                    boundaries.append((f"{group_label}b", tp6_row, end))
                continue

        boundaries.append((group_label, start, end))

    return boundaries


# ---------------------------------------------------------------------------
# Claude API
# ---------------------------------------------------------------------------

def call_claude(client: anthropic.Anthropic, cell_grid: str, group_label: str,
                max_retries: int = 3, retry_delay: int = 20) -> dict:
    """
    Send a cell grid section to Claude claude-sonnet-4-6 and return parsed JSON.
    Uses streaming to support large outputs. Retries on empty or malformed responses.
    """
    user_content = f"Entry group: {group_label}\n\nCell grid:\n{cell_grid}"

    for attempt in range(1, max_retries + 1):
        if attempt > 1:
            print(f"    Retry {attempt}/{max_retries} for {group_label} (waiting {retry_delay}s)...", flush=True)
            time.sleep(retry_delay)

        try:
            raw_parts = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=32000,
                system=SYSTEM_PROMPT,
                messages=[{"role": "user", "content": user_content}],
            ) as stream:
                for text in stream.text_stream:
                    raw_parts.append(text)

            raw = "".join(raw_parts).strip()

            if not raw:
                print(f"    Empty response on attempt {attempt}", flush=True)
                continue

            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)

            try:
                return json.loads(raw)
            except json.JSONDecodeError as e:
                print(f"    JSON parse error on attempt {attempt}: {e}", flush=True)
                if attempt == max_retries:
                    return {"entry_group": group_label, "_parse_error": str(e), "_raw": raw}

        except Exception as e:
            print(f"    API error on attempt {attempt}: {e}", flush=True)
            if attempt == max_retries:
                return {"entry_group": group_label, "_parse_error": str(e), "_raw": ""}

    return {"entry_group": group_label, "_parse_error": "All retries exhausted", "_raw": ""}


# ---------------------------------------------------------------------------
# JSON → flat row lists
# ---------------------------------------------------------------------------

def flatten_to_rows(group_result: dict, year: str) -> dict[str, list[dict]]:
    """
    Convert Claude's JSON result for one group into flat row dicts
    ready for DataFrame assembly.
    """
    # Normalize sub-chunk labels:
    #   Group_4a, Group_4b, Group_4b_tp6 -> Group_4
    group_raw = group_result.get("entry_group", "unknown")
    group = re.sub(r"(Group_\d+)[ab](?:_tp\w+)?$", r"\1", group_raw)
    rows: dict[str, list[dict]] = {
        "phenotypes": [], "strains": [], "parentage": [],
        "descriptive": [], "disease": [], "summary": [],
    }

    # parentage + strains
    for rec in group_result.get("parentage", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["parentage"].append({
            "Strain": strain,
            "Parentage": rec.get("parentage", ""),
            "PrevTesting": rec.get("prev_testing", ""),
            "Generation": rec.get("generation", ""),
            "Test": group,
            "Year": year,
        })
        rows["strains"].append({"Strain": strain, "Test": group, "Year": year})

    # descriptive (tp3a)
    for rec in group_result.get("descriptive", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["descriptive"].append({
            "Strain": strain,
            "DescriptiveCode": rec.get("descriptive_code", ""),
            "ChlorosisAmes": rec.get("chlorosis_ames", ""),
            "ChlorosisLamberton": rec.get("chlorosis_lamberton", ""),
            "HypocotylAmes": rec.get("hypocotyl_ames", ""),
            "Shattering": rec.get("shattering", ""),
            "Test": group,
            "Year": year,
        })

    # disease resistance (tp3b)
    for rec in group_result.get("disease", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["disease"].append({
            "Strain": strain,
            "BSR_pct": rec.get("BSR_pct", ""),
            "GERM_pct": rec.get("GERM_pct", ""),
            "SMV_score": rec.get("SMV_score", ""),
            "PSB_pct": rec.get("PSB_pct", ""),
            "PS_score": rec.get("PS_score", ""),
            "PR_reaction": rec.get("PR_reaction", ""),
            "Test": group,
            "Year": year,
        })

    # summary averages (tp4)
    for rec in group_result.get("summary", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["summary"].append({
            "Strain": strain,
            "YieldBuA": rec.get("yield_bua", ""),
            "Rank": rec.get("rank", ""),
            "Maturity": rec.get("maturity", ""),
            "Lodging": rec.get("lodging", ""),
            "HeightIn": rec.get("height_in", ""),
            "Quality": rec.get("quality", ""),
            "SeedSizeG100": rec.get("seed_size_g100", ""),
            "Protein_pct": rec.get("protein_pct", ""),
            "Oil_pct": rec.get("oil_pct", ""),
            "Test": group,
            "Year": year,
        })

    # per-location phenotypes (tp6–tp12b) — long format
    for pheno_section in group_result.get("phenotypes", []):
        trait = (pheno_section.get("trait") or "unknown").strip()
        for data_row in pheno_section.get("data", []):
            strain = (data_row.get("strain") or "").strip()
            if not strain:
                continue
            mean_val = data_row.get("mean", "")
            # Mean row
            rows["phenotypes"].append({
                "Strain": strain, "Year": year, "Test": group,
                "City": "Mean", "State": "", "Phenotype": trait, "Value": mean_val,
            })
            # Per-location rows
            for loc_key, val in (data_row.get("locations") or {}).items():
                # loc_key: "Ottawa_ONT" → city="Ottawa", state="ONT"
                parts = loc_key.rsplit("_", 1)
                city = parts[0].replace("_", " ") if len(parts) == 2 else loc_key
                state = parts[1] if len(parts) == 2 else ""
                rows["phenotypes"].append({
                    "Strain": strain, "Year": year, "Test": group,
                    "City": city, "State": state, "Phenotype": trait, "Value": val,
                })

    return rows


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_file(xlsx_path: Path, out_dir: Path, client: anthropic.Anthropic) -> None:
    """Full extraction pipeline for one XLSX file."""
    print(f"\nProcessing: {xlsx_path.name}")

    year = extract_year_from_filename(xlsx_path.name)
    print(f"  Year detected: {year}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.max_row} rows x {ws.max_column} cols")

    boundaries = find_group_boundaries(ws)
    print(f"  Sections found: {[b[0] for b in boundaries]}")

    # Accumulate rows across all groups
    all_rows: dict[str, list[dict]] = {
        k: [] for k in ("phenotypes", "strains", "parentage", "descriptive", "disease", "summary")
    }

    for group_label, start_row, end_row in boundaries:
        print(f"  [{group_label}] rows {start_row}-{end_row} -> Claude API...", end=" ", flush=True)
        cell_grid = sheet_to_text(ws, min_row=start_row, max_row=end_row)
        result = call_claude(client, cell_grid, group_label)

        if "_parse_error" in result:
            print("ERROR")
            err_path = out_dir / f"error_{group_label}.json"
            err_path.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"    Saved error details -> {err_path.name}")
            continue

        group_rows = flatten_to_rows(result, year)
        for key in all_rows:
            all_rows[key].extend(group_rows[key])

        counts = {k: len(v) for k, v in group_rows.items() if v}
        print("done. " + ", ".join(f"{k}:{n}" for k, n in counts.items()))

    # Write output CSVs
    safe_stem = re.sub(r'[\\/*?:"<>|()]+', '', xlsx_path.stem).strip().replace(" ", "_")
    print(f"\n  Writing CSVs (prefix: {safe_stem}_):")
    for table_name, row_list in all_rows.items():
        if row_list:
            df = pd.DataFrame(row_list).drop_duplicates()
            out_path = out_dir / f"{safe_stem}_{table_name}.csv"
            df.to_csv(out_path, index=False)
            print(f"    {table_name}: {len(df)} rows -> {out_path.name}")
        else:
            print(f"    {table_name}: (no data)")


def main():
    parser = argparse.ArgumentParser(
        description="Extract NUST historical XLSX data using Claude API (Option C)"
    )
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--file", type=str, help="Path to a single XLSX file")
    src.add_argument("--dir",  type=str, help="Directory of XLSX files to batch-process")
    parser.add_argument("--out_dir",  type=str, required=True, help="Output directory for CSVs")
    parser.add_argument("--api_key",  type=str, default=None,
                        help="Anthropic API key (or set ANTHROPIC_API_KEY env var)")
    args = parser.parse_args()

    api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        print("Error: Anthropic API key required. Set ANTHROPIC_API_KEY or use --api_key.")
        sys.exit(1)

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    client = anthropic.Anthropic(api_key=api_key)

    if args.file:
        xlsx_files = [Path(args.file)]
    else:
        xlsx_files = sorted(
            p for p in Path(args.dir).glob("*.xlsx") if not p.name.startswith("~$")
        )
        print(f"Found {len(xlsx_files)} XLSX file(s) in {args.dir}")

    for xlsx_path in xlsx_files:
        process_file(xlsx_path, out_dir, client)

    print("\nAll done.")


if __name__ == "__main__":
    main()
