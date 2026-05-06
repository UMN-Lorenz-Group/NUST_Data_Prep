#!/usr/bin/env python
"""
extract_nust_xlsx.py
====================
Extract all NUST trial data from historical XLSX reports using Claude API.
Claude receives the raw cell grid and handles all table detection, location
normalization, and schema mapping — no rule-based parsing.

Usage:
    python extract_nust_xlsx.py --file "Sojabone-1980 (1-89 OR).xlsx" --out_dir ./output/
    python extract_nust_xlsx.py --dir ./NUST_Historical_Data/ --out_dir ./output/

Requirements:
    pip install openpyxl anthropic pandas

Environment:
    ANTHROPIC_API_KEY must be set, or pass --api_key
"""

import argparse
import json
import os
import re
import sys
import time
from collections import Counter
from pathlib import Path

# Force UTF-8 output on Windows
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import anthropic
import pandas as pd


# ---------------------------------------------------------------------------
# System prompt — describes NUST schema and output format to Claude
# ---------------------------------------------------------------------------

SYSTEM_PROMPT = """You are extracting structured data from a historical NUST (North American Uniform Soybean Trial) spreadsheet section.

The cell grid is formatted as: R{row_number}<TAB>col1<TAB>col2<TAB>...
Empty cells are empty strings. Row numbers are included for reference.

The section you receive covers either a global parentage table or ONE entry group from the trial.
Entry groups contain several sub-tables tagged by tp* codes in column A:

SUB-TABLE TYPES:
- tp1: Global parentage (Strain, Parentage or Source) — 2 columns only
- tp2: Entry group parentage (Strain, Parentage, Previous Testing, Generation Composited)
- tp3a: Descriptive/morphological (Strain, Descriptive Code, Chlorosis Score Ames, Chlorosis Score Lamberton, Hypocotyl Score Ames, Shattering)
- tp3b: Disease resistance (Strain, BSR %, GERM %, SMV score, PSB %, PS score, PR Reaction)
- tp4: Summary performance current year avg (Strain, Yield bu/a, Rank No., Maturity Date, Lodging Score, Height In., Seed Quality Score, Seed Size g/100, Composition Protein %, Composition Oil %)
- tp5: Multi-year means — SKIP ENTIRELY, do not extract
- tp6: Per-location YIELD (bu/a)
- tp7: Per-location Yield Rank
- tp8: Per-location MATURITY (date)
- tp9: Per-location LODGING (score)
- tp10: Per-location HEIGHT (inches)
- tp11a: Per-location SEED QUALITY (score)
- tp11b: Per-location SEED SIZE (g/100)
- tp12a: Per-location % PROTEIN
- tp12b: Per-location % OIL

PER-LOCATION TABLE STRUCTURE (tp6–tp12b):
- Row with tp* tag: section marker (col A = tp code, col B = trait name label)
- Next non-empty row: column headers — col A = "Strain", col B = "N Mean Tests" or similar, cols C+ = location names
- Data rows: col A = strain name, col B = mean value, cols C+ = per-location values
- FOOTER rows (do NOT treat as strain data, capture as metadata):
  "No. of Tests", "C.V. (%)", "L.S.D. (5%)", "Row sp (in.)", "Rows/plot", "Reps"

LOCATION NORMALIZATION — convert abbreviations to City_State format.
Modern format (State. City — used ~1970+):
- "Ont. Ottawa" → city="Ottawa", state="ONT"
- "Ont. Elora" → city="Elora", state="ONT"
- "Wisc. Ashland" → city="Ashland", state="WI"
- "N.D. Fargo" → city="Fargo", state="ND"
- "Man. Morden" → city="Morden", state="MAN"
- "Man. Brandon" or "Man. Brandon*" → city="Brandon", state="MAN"
- "Minn. Rosemount" → city="Rosemount", state="MN"
- "Minn. Morris" or "Minn. Norris" → city="Morris", state="MN"
- For any unrecognized location: keep as city, state=""

TRAIT NORMALIZATION — map all observed label variants to canonical names:
Yield:
- "YIELD (bu/a)", "YIELD", "Yield in bushels per acre", "Yield in Bu/A", "Yield Bu/A",
  "Yield (bushels per acre)", "Yields in bushels", "YIELD (bu./A)",
  "{year} YIELD (bu/a)", "1970 YIELD (bu/a)" → "YIELD (bu/a)"
Yield Rank:
- "Yield Rank", "YIELD RANK", "Rank of Yield", "Rank Yield", "Summary Yield Rank" → "YIELD RANK"
Maturity:
- "MATURITY (date)", "MATURITY DATE", "MATURITY (relative date)", "Maturity data",
  "Summary Maturity", "Maturity (date)" → "MATURITY (date)"
  (Preserve the original maturity value as-is whether it is a calendar date or a relative offset.)
Plant Height:
- "PLANT HEIGHT (inches)", "HEIGHT (inches)", "HEIGHT", "Plant height",
  "Summary of Height data", "PLANT HEIGHT (in)", "PLANT HEIGHT (in.)" → "PLANT HEIGHT (inches)"
Lodging:
- "LODGING (score)", "Lodging", "LODGING", "Lodging Score",
  "Summary of lodging data", "LODGING (SCORE)" → "LODGING (score)"
Seed Quality:
- "SEED QUALITY (score)", "QUALITY (score)", "Quality (score)", "QUALITY",
  "Seed Quality Score", "Seed quality", "Summary of seed quality" → "SEED QUALITY (score)"
Seed Size:
- "SEED SIZE (g/100)", "SIZE (g/100)", "SEED SIZE", "SIZE",
  "Seed Weight", "Summary of seed weight data in grams per 100" → "SEED SIZE (g/100)"
Protein:
- "PROTEIN (%)", "% PROTEIN", "PROTEIN (%)", "Percentage of Protein",
  "% Protein", "Protein", "PROTEIN  (%)", "PROTIEN (%)" → "PROTEIN (%)"
Oil:
- "OIL (%)", "% OIL", "OIL (%)", "Oil (%)", "Percentage of Oil",
  "% Oil", "Oil", "Percentages of oil" → "OIL (%)"

SKIP THESE ENTIRELY (multi-year summaries and continuation headers — return no data):
- "2-year summary", "3-year summary", "4-year summary", "Five-year summary",
  "Six-year summary", "Three-year summary", "Four-year summary",
  "{n}-year mean Yield", "{year1}-{year2} MEAN", "1968-70 MEAN YIELD"
- Any row where column A or the trait label contains "(Continued)" or "(cont.)" alone —
  treat as a continuation header, not a new section.

Return ONLY valid JSON (no markdown fences) with this exact structure:
{
  "entry_group": "<group label passed to you, e.g. Group_1>",
  "parentage": [
    {"strain": "...", "parentage": "...", "prev_testing": "...", "generation": "..."}
  ],
  "descriptive": [
    {"strain": "...", "descriptive_code": "...", "chlorosis_ames": "...", "chlorosis_lamberton": "...", "hypocotyl_ames": "...", "shattering": "..."}
  ],
  "disease": [
    {"strain": "...", "BSR_pct": "...", "GERM_pct": "...", "SMV_score": "...", "PSB_pct": "...", "PS_score": "...", "PR_reaction": "..."}
  ],
  "summary": [
    {"strain": "...", "yield_bua": "...", "rank": "...", "maturity": "...", "lodging": "...", "height_in": "...", "quality": "...", "seed_size_g100": "...", "protein_pct": "...", "oil_pct": "..."}
  ],
  "phenotypes": [
    {
      "trait": "YIELD (bu/a)",
      "data": [
        {"strain": "...", "mean": "...", "locations": {"Ottawa_ONT": "...", "Elora_ONT": "..."}}
      ],
      "metadata": {"no_of_tests": {}, "cv_pct": {}, "lsd_5pct": {}, "row_spacing_in": {}, "rows_per_plot": {}, "reps": {}}
    }
  ]
}

IMPORTANT NOTES:
- Include ALL strain data rows found (do not truncate)
- Use null for missing, dash, or blank values
- Preserve original value formatting (e.g. "9-5*", "+4.5 *", "3M", "5S")
- If a section is absent from the grid, return an empty list for that key
- tp5 (multi-year means) must always be skipped entirely
- The "summary" key captures tp4 data (current-year averages only, not per-location)
- location keys in the "locations" dict use City_State format: "Ottawa_ONT", "Morris_MN"
"""

# ---------------------------------------------------------------------------
# Era-specific addenda — appended to SYSTEM_PROMPT for pre-1970 years
# ---------------------------------------------------------------------------

ERA_ADDENDUM_EARLY = """
=== EARLY-ERA RULES (years 1941–1956) ===

LOCATION FORMAT: City name comes FIRST, state abbreviation SECOND with trailing period.
Examples:
- "Morris Minn."      → city="Morris",        state="MN"
- "Fargo N.D."        → city="Fargo",         state="ND"
- "Madison Wis."      → city="Madison",       state="WI"
- "Ashland Wis."      → city="Ashland",       state="WI"
- "Guelph Ont."       → city="Guelph",        state="ONT"
- "Ottawa Ont."       → city="Ottawa",        state="ONT"
- "Ridgetown Ont."    → city="Ridgetown",     state="ONT"
- "Harrow Ont."       → city="Harrow",        state="ONT"
- "Morden Man."       → city="Morden",        state="MAN"
- "Brandon Man."      → city="Brandon",       state="MAN"
- "Crookston Minn."   → city="Crookston",     state="MN"
- "Lamberton Minn."   → city="Lamberton",     state="MN"
- "East Lansing Mich."→ city="East Lansing",  state="MI"
- "Mt. Morris Ill."   → city="Morris",        state="IL"
- "Urbana Ill."       → city="Urbana",        state="IL"
- "DeKalb Ill."       → city="DeKalb",        state="IL"
- "Beltsville Md."    → city="Beltsville",    state="MD"
- "Georgetown Del."   → city="Georgetown",    state="DE"
- "State College Pa." → city="State College", state="PA"
- "Lafayette Ind."    → city="Lafayette",     state="IN"
- "Evansville Ind."   → city="Evansville",    state="IN"
- "Vincennes Ind."    → city="Vincennes",     state="IN"
- "Columbia Mo."      → city="Columbia",      state="MO"
- "Manhattan Kans."   → city="Manhattan",     state="KS"
- "Ontario Ore."      → city="Ontario",       state="OR"
- "Prosser Wash."     → city="Prosser",       state="WA"

State abbreviation key: Ill.=IL  Ind.=IN  Mo.=MO  Wis.=WI  Mich.=MI  Minn.=MN
N.D.=ND  S.D.=SD  Kans.=KS  Md.=MD  Va.=VA  Pa.=PA  Del.=DE  N.J.=NJ  Ohio=OH
Ore.=OR  Wash.=WA  Neb.=NE  Ont.=ONT  Man.=MAN  Sask.=SK

FOOTNOTE STRIPPING: Many location headers carry trailing footnote digits or superscript
characters. Strip these before normalizing:
  "Guelph Ont.1" → "Guelph Ont."   "Mt. Morris Ill.¹" → "Mt. Morris Ill."
  "Fall City Wis. 1" → "Fall City Wis."

FUSED TP MARKERS: Some years compress two sub-tables into one block. Use column headers
to identify which columns belong to each trait and emit two separate entries in "phenotypes":
- "tp6+7"   → YIELD (bu/a) columns (tp6) AND YIELD RANK columns (tp7) in same block
- "tp9+10"  → LODGING (score) columns (tp9) AND PLANT HEIGHT (inches) columns (tp10)
- "tp9+tp10"→ same as tp9+10
- "tp8+12a" → MATURITY (date) columns (tp8) AND PROTEIN (%) columns (tp12a)
- "tp8+11a" → MATURITY (date) columns (tp8) AND SEED QUALITY (score) columns (tp11a)

ABSENT SECTIONS (normal for this era — return empty list, do not flag as errors):
- tp1 (global parentage): typically absent before 1957
- tp3a (descriptive/morphological): typically absent before 1963
- tp3b (disease resistance): typically absent before 1957

ADDITIONAL TRAITS present in early years:
- "Iodine number of oil" / "Iodine Number of Oil" → trait="IODINE NUMBER OF OIL", units=""
  Extract as a per-location phenotype (tp-style table). Not present after ~1948.
- "SEED WEIGHT (cg)" / "Seed weight (cg)" / "Seed weight in centigrams" →
  trait="SEED WEIGHT (cg)", units="cg"
  This is centigrams per seed — DO NOT merge with "SEED SIZE (g/100)". Keep as a distinct
  phenotype name.
"""

ERA_ADDENDUM_TRANSITIONAL = """
=== TRANSITIONAL-ERA RULES (years 1957–1969) ===

LOCATION FORMAT: Mixed — some files use modern "State. City" format, others use early
"City State." format (with dotted abbreviations). Both may even appear in the same file.
Apply all location mappings from the base rules PLUS the following dotted-abbreviation set:
  "Morris Minn."       → city="Morris",       state="MN"
  "Fargo N.D."         → city="Fargo",        state="ND"
  "Guelph Ont."        → city="Guelph",       state="ONT"
  "Ottawa Ont."        → city="Ottawa",       state="ONT"
  "Harrow Ont."        → city="Harrow",       state="ONT"
  "Ridgetown Ont."     → city="Ridgetown",    state="ONT"
  "Morden Man."        → city="Morden",       state="MAN"
  "Brandon Man."       → city="Brandon",      state="MAN"
  "East Lansing Mich." → city="East Lansing", state="MI"
  "Madison Wis."       → city="Madison",      state="WI"
  "Urbana Ill."        → city="Urbana",       state="IL"
  "DeKalb Ill."        → city="DeKalb",       state="IL"
  "Lafayette Ind."     → city="Lafayette",    state="IN"
  "Manhattan Kans."    → city="Manhattan",    state="KS"
State abbreviation key: Ill.=IL  Ind.=IN  Mo.=MO  Wis.=WI  Mich.=MI  Minn.=MN
N.D.=ND  S.D.=SD  Kans.=KS  Pa.=PA  Del.=DE  N.J.=NJ  Ohio=OH  Ont.=ONT  Man.=MAN

FOOTNOTE STRIPPING: Strip trailing digits and superscripts from location names:
  "Guelph Ont.1" → "Guelph Ont."   "DeKalb Ill. 1" → "DeKalb Ill."

FUSED TP MARKERS:
- "tp12a & tp12b" or "tp12a&tp12b" → PROTEIN (%) columns (tp12a) AND OIL (%) columns (tp12b)
- "tp10 & tp12b"                    → PLANT HEIGHT (inches) (tp10) AND OIL (%) (tp12b)
- "tp6+7"                           → YIELD (bu/a) (tp6) AND YIELD RANK (tp7)

OCR ARTIFACTS — normalize before processing:
- "tptp11a"      → treat as tp11a (SEED QUALITY)
- "tp 2", "tp 7" → treat as tp2, tp7 (strip the internal space)
- "tp24"         → skip entirely (OCR error)
- "tp3" (bare, no letter suffix) → if columns contain morphological data (Descriptive Code,
  Chlorosis, Shattering), treat as tp3a; otherwise skip
- "tp3c"         → extract into "descriptive" table with whatever columns are present;
  add a note field: "section": "tp3c"
- "tp2 (NB ...)" → treat as tp2 (strip the Afrikaans/parenthetical annotation)

MULTI-YEAR SUMMARIES — skip entirely (more common in transitional years):
"2-year summary", "3-year summary", "4-year summary", "Five-year summary",
"Six-year summary", "Three-year summary", "Four-year summary",
"{n}-year mean Yield", "{year1}-{year2} MEAN", "1968-70 MEAN YIELD",
"1968-70 RANK", "Summary Yields" appearing as a section header (not a trait label)

ABSENT SECTIONS (may or may not be present — return empty list if absent):
- tp1 (global parentage): appears from ~1957 onward; may still be absent in early transitional
- tp3b (disease resistance): appears from ~1957; may be absent before 1960
"""


def get_era(year: int) -> str:
    """Return era key for a given trial year."""
    if year <= 1956:
        return "early"
    if year <= 1969:
        return "transitional"
    return "modern"


def build_system_prompt(year: str) -> str:
    """Return the full system prompt for the given year string (e.g. '1952')."""
    try:
        yr_int = int(year)
    except (ValueError, TypeError):
        return SYSTEM_PROMPT
    era = get_era(yr_int)
    if era == "early":
        return SYSTEM_PROMPT + "\n" + ERA_ADDENDUM_EARLY
    if era == "transitional":
        return SYSTEM_PROMPT + "\n" + ERA_ADDENDUM_TRANSITIONAL
    return SYSTEM_PROMPT


# ---------------------------------------------------------------------------
# Cell grid utilities
# ---------------------------------------------------------------------------

def extract_year_from_filename(filename: str) -> str:
    """Extract 4-digit year from filename like 'Sojabone-1980 (1-89 OR).xlsx'."""
    match = re.search(r'\b(19|20)\d{2}\b', filename)
    return match.group(0) if match else "unknown"


def sheet_to_text(ws, min_row: int = 1, max_row: int = None) -> str:
    """
    Serialize worksheet rows as a tab-separated text grid with row numbers.
    Only includes non-empty rows; trims trailing empty columns per row.
    """
    if max_row is None:
        max_row = ws.max_row
    lines = []
    for i, row in enumerate(
        ws.iter_rows(min_row=min_row, max_row=max_row, values_only=True), start=min_row
    ):
        vals = [str(v) if v is not None else "" for v in row]
        while vals and vals[-1] == "":
            vals.pop()
        if any(v.strip() for v in vals):
            lines.append(f"R{i}\t" + "\t".join(vals))
    return "\n".join(lines)


PER_LOC_MARKERS = ("tp6", "tp7", "tp8", "tp9", "tp10", "tp11a", "tp11b", "tp12a", "tp12b")
ALL_TP_MARKERS = ("tp1", "tp2", "tp3a", "tp3b", "tp4", "tp5") + PER_LOC_MARKERS

# Fused markers: map raw cell value -> canonical marker for boundary/chunking purposes.
# Claude handles the actual split; we just need to know where each block starts.
_FUSED_MARKER_MAP: dict[str, str | None] = {
    "tp6+7":         "tp6",
    "tp9+10":        "tp9",
    "tp9+tp10":      "tp9",
    "tp8+11a":       "tp8",
    "tp8+12a":       "tp8",
    "tp7+8":         "tp7",
    "tp12a & tp12b": "tp12a",
    "tp12a&tp12b":   "tp12a",
    "tp10 & tp12b":  "tp10",
    "tptp11a":       "tp11a",
    "tp":            None,   # bare separator artifact — skip
    "tp24":          None,   # OCR error — skip
}
_NOISY_TP_RE = re.compile(r'^tp\?+$')   # tp??, tp??? → skip


def _normalize_tp(sv: str) -> str | None:
    """
    Map a raw column-A cell value to a canonical tp marker string, or None to skip.
    Handles fused markers (tp6+7), noisy codes (tp??), spaced codes (tp 2),
    continuation annotations (tp6 (Continued)), and Afrikaans notes (tp2 (NB ...)).
    """
    if sv in _FUSED_MARKER_MAP:
        return _FUSED_MARKER_MAP[sv]
    if _NOISY_TP_RE.match(sv):
        return None
    # "tp 2", "tp 7" etc. — strip internal space
    m = re.match(r'^tp\s+(\w+)$', sv)
    if m:
        return f"tp{m.group(1)}"
    # "tp6 (Continued)" / "tp2 (NB ...)" — strip annotation
    m = re.match(r'^(tp\w+)\s*\(', sv)
    if m:
        return m.group(1)
    return sv


def classify_marker(sv: str) -> str:
    """Classify a raw column-A tp value into one of four categories."""
    if sv in ALL_TP_MARKERS:
        return "standard"
    norm = _normalize_tp(sv)
    if norm is None:
        return "known-skip"
    if sv in _FUSED_MARKER_MAP:
        return "known-fused"
    if norm != sv and norm in ALL_TP_MARKERS:
        return "known-noisy"
    return "unknown"


def _row_to_list(row_tuple: tuple) -> list[str]:
    """Convert a row values-tuple to a trimmed list of strings."""
    vals = [str(v) if v is not None else "" for v in row_tuple]
    while vals and not vals[-1].strip():
        vals.pop()
    return vals


def scan_markers(ws) -> dict:
    """
    Scan column A of ws for all tp-marker values and classify each one.

    Returns a dict with keys: standard, known_skip, known_fused, known_noisy, unknown.
    Each entry includes row number, raw value, canonical form, and for fused/unknown
    markers a preview of surrounding rows (column headers or context) so the caller
    can verify column assignment without an API call.
    """
    all_rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))

    findings: dict[str, list[dict]] = {
        "standard": [], "known_skip": [], "known_fused": [],
        "known_noisy": [], "unknown": [],
    }

    for i, row in enumerate(all_rows, start=1):
        v = row[0]
        if not v:
            continue
        sv = str(v).strip()
        if not sv.lower().startswith("tp"):
            continue

        cls = classify_marker(sv)
        norm = _normalize_tp(sv)
        entry: dict = {"row": i, "raw": sv, "canonical": norm}

        if cls == "known-fused":
            # Rows i, i+1, i+2 (1-indexed) = marker + trait-label row + column-header row
            preview = []
            for k in range(i - 1, min(i + 2, len(all_rows))):
                cells = _row_to_list(all_rows[k])
                if cells:
                    preview.append({"sheet_row": k + 1, "cells": cells})
            entry["header_preview"] = preview

        elif cls == "unknown":
            # ±2 rows of context for investigation
            context = []
            for k in range(max(0, i - 3), min(i + 3, len(all_rows))):
                cells = _row_to_list(all_rows[k])
                if cells:
                    context.append({"sheet_row": k + 1, "cells": cells})
            entry["context"] = context

        findings[cls.replace("-", "_")].append(entry)

    return findings


def find_group_boundaries(ws) -> list[tuple[str, int, int]]:
    """
    Locate entry group boundaries by finding tp2 markers (start of each group).
    Also captures the global tp1 parentage section before the first tp2.

    Chunking strategy to stay within reliable API output limits (~150 rows/call):
    - Groups <=300 rows: sent as one chunk.
    - Groups >300 rows: split into:
        a) rows tp2..tp6-1 (parentage, descriptive, disease, summary)
        b) per-location phenotypes, further split per-trait if b >150 rows
           (each tp6/tp7/.../tp12b table becomes its own chunk)
    Falls back to treating the whole sheet as one group if no tp2 found.
    Returns list of (group_label, start_row, end_row).
    """
    # Collect all tp marker row positions in one pass
    tp2_rows = []
    tp_marker_rows: dict[int, str] = {}  # row -> canonical marker value
    for i, row in enumerate(
        ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1
    ):
        v = row[0]
        if v:
            sv = str(v).strip()
            sv_norm = _normalize_tp(sv)
            if sv_norm is None:
                continue
            if sv_norm == "tp2":
                tp2_rows.append(i)
            if sv_norm in ALL_TP_MARKERS:
                tp_marker_rows[i] = sv_norm

    if not tp2_rows:
        print("  Warning: no tp2 markers found -- sending entire sheet as one group")
        return [("Group_1", 1, ws.max_row)]

    boundaries = []

    # Global tp1 parentage section (rows before first tp2)
    if tp2_rows[0] > 1:
        boundaries.append(("GlobalParentage", 1, tp2_rows[0] - 1))

    # Entry groups: tp2 -> next tp2 (or end of sheet)
    for idx, start in enumerate(tp2_rows):
        end = tp2_rows[idx + 1] - 1 if idx + 1 < len(tp2_rows) else ws.max_row
        group_label = f"Group_{idx + 1}"
        n_rows = end - start + 1

        if n_rows > 300:
            # Find the first tp6 within this group
            tp6_row = None
            for r in sorted(tp_marker_rows):
                if start < r <= end and tp_marker_rows[r] == "tp6":
                    tp6_row = r
                    break

            if tp6_row and tp6_row > start + 10:
                # Sub-chunk A: parentage + descriptive + disease + summary (before tp6)
                boundaries.append((f"{group_label}a", start, tp6_row - 1))

                # Sub-chunk B: per-location phenotypes
                b_rows = end - tp6_row + 1
                if b_rows > 150:
                    # Split into per-trait chunks (one chunk per tp6/tp7/.../tp12b marker)
                    per_loc_in_group = sorted(
                        r for r in tp_marker_rows
                        if tp6_row <= r <= end and tp_marker_rows[r] in PER_LOC_MARKERS
                    )
                    for t_idx, t_start in enumerate(per_loc_in_group):
                        t_end = per_loc_in_group[t_idx + 1] - 1 if t_idx + 1 < len(per_loc_in_group) else end
                        trait_tag = tp_marker_rows[t_start]
                        boundaries.append((f"{group_label}b_{trait_tag}", t_start, t_end))
                else:
                    boundaries.append((f"{group_label}b", tp6_row, end))
                continue

        boundaries.append((group_label, start, end))

    return boundaries


# ---------------------------------------------------------------------------
# Claude API
# ---------------------------------------------------------------------------

def call_claude(client: anthropic.Anthropic, cell_grid: str, group_label: str,
                year: str = "unknown",
                max_retries: int = 3, retry_delay: int = 20) -> dict:
    """
    Send a cell grid section to Claude claude-sonnet-4-6 and return parsed JSON.
    Uses streaming to support large outputs. Retries on empty or malformed responses.
    The system prompt is era-aware: early (≤1956) and transitional (1957–1969) years
    receive additional structural rules appended to the base prompt.
    """
    user_content = f"Entry group: {group_label}\n\nCell grid:\n{cell_grid}"
    system_prompt = build_system_prompt(year)

    for attempt in range(1, max_retries + 1):
        if attempt > 1:
            print(f"    Retry {attempt}/{max_retries} for {group_label} (waiting {retry_delay}s)...", flush=True)
            time.sleep(retry_delay)

        try:
            raw_parts = []
            with client.messages.stream(
                model="claude-sonnet-4-6",
                max_tokens=32000,
                system=system_prompt,
                messages=[{"role": "user", "content": user_content}],
            ) as stream:
                for text in stream.text_stream:
                    raw_parts.append(text)

            raw = "".join(raw_parts).strip()

            if not raw:
                print(f"    Empty response on attempt {attempt}", flush=True)
                continue

            raw = re.sub(r"^```(?:json)?\s*", "", raw)
            raw = re.sub(r"\s*```$", "", raw)

            try:
                return json.loads(raw)
            except json.JSONDecodeError as e:
                print(f"    JSON parse error on attempt {attempt}: {e}", flush=True)
                if attempt == max_retries:
                    return {"entry_group": group_label, "_parse_error": str(e), "_raw": raw}

        except anthropic.RateLimitError as e:
            wait = 60
            print(f"    Rate limit hit on attempt {attempt} — sleeping {wait}s...", flush=True)
            time.sleep(wait)
        except anthropic.APIStatusError as e:
            if e.status_code == 529:  # API overloaded
                wait = 30
                print(f"    API overloaded (529) on attempt {attempt} — sleeping {wait}s...", flush=True)
                time.sleep(wait)
            else:
                print(f"    API error on attempt {attempt}: {e}", flush=True)
                if attempt == max_retries:
                    return {"entry_group": group_label, "_parse_error": str(e), "_raw": ""}
        except Exception as e:
            print(f"    API error on attempt {attempt}: {e}", flush=True)
            if attempt == max_retries:
                return {"entry_group": group_label, "_parse_error": str(e), "_raw": ""}

    return {"entry_group": group_label, "_parse_error": "All retries exhausted", "_raw": ""}


# ---------------------------------------------------------------------------
# JSON → flat row lists
# ---------------------------------------------------------------------------

def split_trait_units(trait: str) -> tuple[str, str]:
    """Split 'YIELD (bu/a)' -> ('YIELD', 'bu/a'). Returns (trait, '') if no parenthetical."""
    m = re.match(r'^(.+?)\s*\(([^)]+)\)\s*$', trait.strip())
    if m:
        return m.group(1).strip(), m.group(2).strip()
    return trait.strip(), ""


def flatten_to_rows(group_result: dict, year: str) -> dict[str, list[dict]]:
    """
    Convert Claude's JSON result for one group into flat row dicts
    ready for DataFrame assembly.
    """
    # Normalize sub-chunk labels:
    #   Group_4a, Group_4b, Group_4b_tp6 -> Group_4
    group_raw = group_result.get("entry_group", "unknown")
    group = re.sub(r"(Group_\d+)[ab](?:_tp\w+)?$", r"\1", group_raw)
    rows: dict[str, list[dict]] = {
        "phenotypes": [], "strains": [], "parentage": [],
        "descriptive": [], "disease": [], "summary": [],
    }

    # parentage + strains
    for rec in group_result.get("parentage", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["parentage"].append({
            "Strain": strain,
            "Parentage": rec.get("parentage", ""),
            "PrevTesting": rec.get("prev_testing", ""),
            "Generation": rec.get("generation", ""),
            "Test": group,
            "Year": year,
        })
        rows["strains"].append({"Strain": strain, "Test": group, "Year": year})

    # descriptive (tp3a)
    for rec in group_result.get("descriptive", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["descriptive"].append({
            "Strain": strain,
            "DescriptiveCode": rec.get("descriptive_code", ""),
            "ChlorosisAmes": rec.get("chlorosis_ames", ""),
            "ChlorosisLamberton": rec.get("chlorosis_lamberton", ""),
            "HypocotylAmes": rec.get("hypocotyl_ames", ""),
            "Shattering": rec.get("shattering", ""),
            "Test": group,
            "Year": year,
        })

    # disease resistance (tp3b)
    for rec in group_result.get("disease", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["disease"].append({
            "Strain": strain,
            "BSR_pct": rec.get("BSR_pct", ""),
            "GERM_pct": rec.get("GERM_pct", ""),
            "SMV_score": rec.get("SMV_score", ""),
            "PSB_pct": rec.get("PSB_pct", ""),
            "PS_score": rec.get("PS_score", ""),
            "PR_reaction": rec.get("PR_reaction", ""),
            "Test": group,
            "Year": year,
        })

    # summary averages (tp4)
    for rec in group_result.get("summary", []):
        strain = (rec.get("strain") or "").strip()
        if not strain:
            continue
        rows["summary"].append({
            "Strain": strain,
            "YieldBuA": rec.get("yield_bua", ""),
            "Rank": rec.get("rank", ""),
            "Maturity": rec.get("maturity", ""),
            "Lodging": rec.get("lodging", ""),
            "HeightIn": rec.get("height_in", ""),
            "Quality": rec.get("quality", ""),
            "SeedSizeG100": rec.get("seed_size_g100", ""),
            "Protein_pct": rec.get("protein_pct", ""),
            "Oil_pct": rec.get("oil_pct", ""),
            "Test": group,
            "Year": year,
        })

    # per-location phenotypes (tp6–tp12b) — long format
    for pheno_section in group_result.get("phenotypes", []):
        trait_raw = (pheno_section.get("trait") or "unknown").strip()
        trait_name, units = split_trait_units(trait_raw)
        for data_row in pheno_section.get("data", []):
            strain = (data_row.get("strain") or "").strip()
            if not strain:
                continue
            mean_val = data_row.get("mean", "")
            # Mean row
            rows["phenotypes"].append({
                "Strain": strain, "Year": year, "Test": group,
                "City": "Mean", "State": "", "Phenotype": trait_name, "Value": mean_val, "Units": units,
            })
            # Per-location rows
            for loc_key, val in (data_row.get("locations") or {}).items():
                # loc_key: "Ottawa_ONT" → city="Ottawa", state="ONT"
                parts = loc_key.rsplit("_", 1)
                city = parts[0].replace("_", " ") if len(parts) == 2 else loc_key
                state = parts[1] if len(parts) == 2 else ""
                rows["phenotypes"].append({
                    "Strain": strain, "Year": year, "Test": group,
                    "City": city, "State": state, "Phenotype": trait_name, "Value": val, "Units": units,
                })

    return rows


# ---------------------------------------------------------------------------
# Dry-run: marker scan (no API calls)
# ---------------------------------------------------------------------------

def scan_file(xlsx_path: Path, out_dir: Path) -> bool:
    """
    Scan one XLSX file for tp markers and print a classified inventory.
    Writes a JSON report to out_dir. Returns True if safe to extract (no unknowns).

    Use this before paying for API calls on an unfamiliar year:
      python scripts/extract_nust_xlsx.py --file X.xlsx --out_dir Y/ --dry-run
    """
    print(f"\nScanning: {xlsx_path.name}")
    year = extract_year_from_filename(xlsx_path.name)
    era  = get_era(int(year)) if year.isdigit() else "unknown"
    print(f"  Year: {year}  Era: {era}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.max_row} rows × {ws.max_column} cols")

    findings = scan_markers(ws)
    std     = findings["standard"]
    skips   = findings["known_skip"]
    fused   = findings["known_fused"]
    noisy   = findings["known_noisy"]
    unknown = findings["unknown"]

    # Standard — summarize as counts per marker type
    std_counts = Counter(e["raw"] for e in std)
    print(f"\n  STANDARD markers ({len(std)}):")
    if std_counts:
        print("    " + "  ".join(f"{k}×{v}" for k, v in sorted(std_counts.items())))
    else:
        print("    (none)")

    # Known-skip
    print(f"\n  KNOWN-SKIP markers ({len(skips)}):")
    for e in skips:
        print(f"    {e['raw']!r:25s} row {e['row']}")
    if not skips:
        print("    (none)")

    # Known-noisy
    print(f"\n  KNOWN-NOISY markers ({len(noisy)}):")
    for e in noisy:
        print(f"    {e['raw']!r:25s} row {e['row']}  → {e['canonical']}")
    if not noisy:
        print("    (none)")

    # Fused — show column-header preview so assignment can be verified without API
    print(f"\n  KNOWN-FUSED markers ({len(fused)}):")
    for e in fused:
        print(f"    {e['raw']!r:25s} row {e['row']}  → canonical: {e['canonical']}")
        for pr in e.get("header_preview", []):
            cells_str = " | ".join(pr["cells"][:10])  # cap at 10 cols for readability
            print(f"      R{pr['sheet_row']:>4}: {cells_str}")
    if not fused:
        print("    (none)")

    # Unknown — show context rows for investigation
    print(f"\n  UNKNOWN markers ({len(unknown)}):")
    for e in unknown:
        print(f"    {e['raw']!r:25s} row {e['row']}  ← INVESTIGATE")
        for ctx in e.get("context", []):
            cells_str = " | ".join(ctx["cells"][:10])
            print(f"      R{ctx['sheet_row']:>4}: {cells_str}")
    if not unknown:
        print("    (none)")

    # JSON report
    safe_stem = re.sub(r'[\\/*?:"<>|()]+', '', xlsx_path.stem).strip().replace(" ", "_")
    report = {
        "file": xlsx_path.name,
        "year": year,
        "era": era,
        "sheet_rows": ws.max_row,
        "summary": {k: len(v) for k, v in findings.items()},
        "findings": findings,
    }
    out_dir.mkdir(parents=True, exist_ok=True)
    report_path = out_dir / f"{safe_stem}_marker_scan.json"
    report_path.write_text(json.dumps(report, indent=2, ensure_ascii=False), encoding="utf-8")

    if unknown:
        print(f"\n  ⚠  {len(unknown)} unknown marker(s) found — inspect before extracting")
        print(f"  Report: {report_path}")
        return False
    else:
        print(f"\n  ✓  All markers recognized — safe to proceed with extraction")
        print(f"  Report: {report_path}")
        return True


# ---------------------------------------------------------------------------
# Main pipeline
# ---------------------------------------------------------------------------

def process_file(xlsx_path: Path, out_dir: Path, client: anthropic.Anthropic) -> None:
    """Full extraction pipeline for one XLSX file."""
    print(f"\nProcessing: {xlsx_path.name}")

    year = extract_year_from_filename(xlsx_path.name)
    print(f"  Year detected: {year}")

    wb = openpyxl.load_workbook(str(xlsx_path), data_only=True)
    ws = wb.active
    print(f"  Sheet: {ws.max_row} rows x {ws.max_column} cols")

    boundaries = find_group_boundaries(ws)
    print(f"  Sections found: {[b[0] for b in boundaries]}")

    # Accumulate rows across all groups
    all_rows: dict[str, list[dict]] = {
        k: [] for k in ("phenotypes", "strains", "parentage", "descriptive", "disease", "summary")
    }

    era = get_era(int(year)) if year.isdigit() else "modern"
    print(f"  Era: {era} (system prompt: {len(build_system_prompt(year))} chars)")

    for group_label, start_row, end_row in boundaries:
        print(f"  [{group_label}] rows {start_row}-{end_row} -> Claude API...", end=" ", flush=True)
        cell_grid = sheet_to_text(ws, min_row=start_row, max_row=end_row)
        result = call_claude(client, cell_grid, group_label, year=year)

        if "_parse_error" in result:
            print("ERROR")
            err_path = out_dir / f"error_{group_label}.json"
            err_path.write_text(json.dumps(result, indent=2, ensure_ascii=False), encoding="utf-8")
            print(f"    Saved error details -> {err_path.name}")
            continue

        group_rows = flatten_to_rows(result, year)
        for key in all_rows:
            all_rows[key].extend(group_rows[key])

        counts = {k: len(v) for k, v in group_rows.items() if v}
        print("done. " + ", ".join(f"{k}:{n}" for k, n in counts.items()))

    # Write output CSVs
    safe_stem = re.sub(r'[\\/*?:"<>|()]+', '', xlsx_path.stem).strip().replace(" ", "_")
    print(f"\n  Writing CSVs (prefix: {safe_stem}_):")
    for table_name, row_list in all_rows.items():
        if row_list:
            df = pd.DataFrame(row_list).drop_duplicates()
            out_path = out_dir / f"{safe_stem}_{table_name}.csv"
            df.to_csv(out_path, index=False)
            print(f"    {table_name}: {len(df)} rows -> {out_path.name}")
        else:
            print(f"    {table_name}: (no data)")


def load_env_file() -> str | None:
    """Read API key from .Env file in the same directory as this script."""
    env_path = Path(__file__).parent / ".Env"
    if not env_path.exists():
        return None
    for line in env_path.read_text(encoding="utf-8").splitlines():
        line = line.strip()
        if line.startswith("ANTHROPIC_API_KEY="):
            return line.split("=", 1)[1].strip()
        if line.startswith("sk-ant-"):
            return line
    return None


def main():
    parser = argparse.ArgumentParser(
        description="Extract NUST historical XLSX data using Claude API (Option C)"
    )
    src = parser.add_mutually_exclusive_group(required=True)
    src.add_argument("--file", type=str, help="Path to a single XLSX file")
    src.add_argument("--dir",  type=str, help="Directory of XLSX files to batch-process")
    parser.add_argument("--out_dir",  type=str, required=True, help="Output directory for CSVs")
    parser.add_argument("--api_key",  type=str, default=None,
                        help="Anthropic API key (overrides .Env / ANTHROPIC_API_KEY env var)")
    parser.add_argument("--dry-run", action="store_true",
                        help="Scan tp markers only — classify and preview without calling the API. "
                             "Writes a *_marker_scan.json report per file. Use this before "
                             "extracting an unfamiliar year to verify all markers are recognized.")
    args = parser.parse_args()

    out_dir = Path(args.out_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    if args.file:
        xlsx_files = [Path(args.file)]
    else:
        xlsx_files = sorted(
            p for p in Path(args.dir).glob("*.xlsx") if not p.name.startswith("~$")
        )
        print(f"Found {len(xlsx_files)} XLSX file(s) in {args.dir}")

    if args.dry_run:
        any_unknown = False
        for xlsx_path in xlsx_files:
            safe = scan_file(xlsx_path, out_dir)
            if not safe:
                any_unknown = True
        print("\nDry-run complete.")
        if any_unknown:
            print("⚠  Unknown markers detected — resolve before running extraction.")
            sys.exit(1)
        else:
            print("✓  All files clean — run without --dry-run to extract.")
        return

    api_key = args.api_key or os.environ.get("ANTHROPIC_API_KEY") or load_env_file()
    if not api_key:
        print("Error: Anthropic API key required. Add it to .Env, set ANTHROPIC_API_KEY, or use --api_key.")
        sys.exit(1)

    client = anthropic.Anthropic(api_key=api_key)

    for xlsx_path in xlsx_files:
        process_file(xlsx_path, out_dir, client)

    print("\nAll done.")


if __name__ == "__main__":
    main()
