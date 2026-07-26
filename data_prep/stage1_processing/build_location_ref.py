#!/usr/bin/env python
"""
build_location_ref.py
=====================
Build a master NUST location reference table (nust_locations_ref.csv) by:
  1. Collecting unique City+State pairs from all existing extracted CSVs
  2. Scanning all Green XLSX files for location strings in per-location table headers
  3. Geocoding each location via Nominatim (1 req/sec, free)
  4. Flagging known NUST research stations for manual coordinate verification

Output: nust_locations_ref.csv
  City, State, Country, lat, lon, StationName, Geocoded_City, NeedsVerification, Source

Usage:
    python build_location_ref.py
"""

import re
import sys
import time
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import pandas as pd
from geopy.geocoders import Nominatim
from geopy.exc import GeocoderTimedOut, GeocoderServiceError

# ---------------------------------------------------------------------------
# Paths
# ---------------------------------------------------------------------------
GREEN = Path("R:/NUST_Historical_Data_1941_1988/Green-20260427T181938Z-3-001/Green")
OUT_DIR = Path("reference")
CSV_DIRS = list(Path(".").glob("output_*/"))

# ---------------------------------------------------------------------------
# Known NUST research stations
# These locations recur frequently and city-center coordinates are wrong.
# Flagged for manual verification — StationName gives the user enough context
# to look up the exact GPS coordinates.
# ---------------------------------------------------------------------------
KNOWN_STATIONS = {
    # (city_norm, state_norm): station name
    ("ames",           "ia"):   "ISU Agronomy and Agricultural Engineering Research Center",
    ("rosemount",      "mn"):   "UMN Rosemount Research and Outreach Center",
    ("morris",         "mn"):   "UMN West Central Research and Outreach Center",
    ("crookston",      "mn"):   "UMN Northwest Research and Outreach Center",
    ("lamberton",      "mn"):   "UMN Southwest Research and Outreach Center",
    ("waseca",         "mn"):   "UMN Southern Research and Outreach Center",
    ("st paul",        "mn"):   "UMN St Paul Campus Farm",
    ("fargo",          "nd"):   "NDSU Main Station / USDA ARS Fargo",
    ("brookings",      "sd"):   "SDSU Agricultural Experiment Station",
    ("west lafayette", "in"):   "Purdue Agronomy Center for Research and Education",
    ("lafayette",      "in"):   "Purdue Agronomy Center for Research and Education",
    ("wooster",        "oh"):   "OARDC (Ohio Ag Research and Development Center)",
    ("columbia",       "mo"):   "MU Bradford Research Center",
    ("mt vernon",      "mo"):   "MU Southwest Research Center",
    ("portageville",   "mo"):   "MU Portageville Research Center",
    ("portageville loam","mo"): "MU Portageville Research Center",
    ("manhattan",      "kan"):  "Kansas State University Agronomy Farm",
    ("manhattan",      "ks"):   "Kansas State University Agronomy Farm",
    ("urbana",         "il"):   "UIUC Crop Sciences Research and Education Center",
    ("dekalb",         "il"):   "NIU / DeKalb County research station",
    ("east lansing",   "mi"):   "MSU Agronomy Farm",
    ("beltsville",     "md"):   "USDA ARS Beltsville Agricultural Research Center",
    ("state college",  "pa"):   "Penn State Agronomy Research Farm",
    ("ithaca",         "ny"):   "Cornell University Experiment Station",
    ("geneva",         "ny"):   "Cornell AgriTech (NYS Agricultural Experiment Station)",
    ("ottawa",         "ont"):  "Agriculture and Agri-Food Canada Ottawa Research Centre",
    ("ottawa",         "on"):   "Agriculture and Agri-Food Canada Ottawa Research Centre",
    ("harrow",         "ont"):  "Agriculture and Agri-Food Canada Harrow Research Centre",
    ("elora",          "ont"):  "University of Guelph Elora Research Station",
    ("guelph",         "ont"):  "University of Guelph",
    ("ridgetown",      "ont"):  "University of Guelph Ridgetown Campus",
    ("morden",         "man"):  "Agriculture and Agri-Food Canada Morden Research Centre",
    ("brandon",        "man"):  "Agriculture and Agri-Food Canada Brandon Research Centre",
    ("winnipeg",       "man"):  "Agriculture and Agri-Food Canada Winnipeg",
    ("swift current",  "sk"):   "Agriculture and Agri-Food Canada Swift Current",
    ("ashland",        "wi"):   "Northern Soils and Crops Research Station (USDA / UW)",
    ("arlington",      "wi"):   "UW-Madison Arlington Agricultural Research Station",
    ("madison",        "wi"):   "UW-Madison Agronomy Farm",
    ("belleville",     "il"):   "USDA ARS Dale Bumpers Small Farms Research Center (check)",
    ("stoneville",     "ms"):   "USDA ARS Stoneville Mississippi Delta",
    ("baton rouge",    "la"):   "LSU AgCenter Dean Lee Research Station",
    ("lubbock",        "tx"):   "Texas A&M AgriLife Research Center Lubbock",
    ("queenstown",     "md"):   "University of Maryland Upper Chesapeake Agricultural Center",
    ("stutts",         "al"):   "Alabama A&M University research farm (verify)",
}

# ---------------------------------------------------------------------------
# State abbreviation normalisation
# Maps historical variants → standard 2-letter codes
# ---------------------------------------------------------------------------
STATE_NORM = {
    # US
    "ill": "IL", "ill.": "IL",
    "ind": "IN", "ind.": "IN",
    "mo": "MO", "mo.": "MO",
    "wis": "WI", "wis.": "WI", "wisc": "WI", "wisc.": "WI",
    "mich": "MI", "mich.": "MI",
    "minn": "MN", "minn.": "MN",
    "n.d": "ND", "n.d.": "ND", "nd": "ND",
    "s.d": "SD", "s.d.": "SD", "sd": "SD",
    "kans": "KS", "kans.": "KS", "kan": "KS", "kan.": "KS", "ks": "KS",
    "md": "MD", "md.": "MD",
    "va": "VA", "va.": "VA",
    "pa": "PA", "pa.": "PA",
    "del": "DE", "del.": "DE",
    "n.j": "NJ", "n.j.": "NJ", "nj": "NJ",
    "ohio": "OH", "oh": "OH",
    "ore": "OR", "ore.": "OR",
    "wash": "WA", "wash.": "WA",
    "neb": "NE", "neb.": "NE", "nebr": "NE",
    "ia": "IA", "iowa": "IA",
    "ky": "KY", "ky.": "KY",
    "ny": "NY", "n.y": "NY",
    "la": "LA", "la.": "LA",
    "ms": "MS",
    "al": "AL", "ala": "AL",
    "ga": "GA",
    "nc": "NC", "n.c": "NC",
    "sc": "SC",
    "tx": "TX", "tex": "TX",
    "ok": "OK", "okla": "OK",
    "ar": "AR", "ark": "AR",
    "tn": "TN", "tenn": "TN",
    "in": "IN",  # also Indiana
    "il": "IL",
    "wi": "WI",
    "mi": "MI",
    "mn": "MN",
    "mo2": "MO",  # avoid collision with "mo" already there
    # Canadian
    "ont": "ONT", "ont.": "ONT", "on": "ONT",
    "man": "MAN", "man.": "MAN", "mb": "MAN",
    "sask": "SK", "sask.": "SK",
    "alta": "AB", "ab": "AB",
    "bc": "BC",
    "que": "QC", "que.": "QC", "qc": "QC",
    "ns": "NS",
    "nb": "NB",
}

COUNTRY_BY_STATE = {
    "ONT": "Canada", "MAN": "Canada", "SK": "Canada",
    "AB": "Canada", "BC": "Canada", "QC": "Canada", "NS": "Canada", "NB": "Canada",
}


def norm_state(s: str) -> str:
    s = s.strip().rstrip(".")
    return STATE_NORM.get(s.lower(), s.upper())


def parse_location_string(raw: str) -> tuple[str, str] | None:
    """
    Parse a raw location string into (city, state).
    Handles both era formats and strips footnote markers.
    Returns None if unparseable.
    """
    if not raw or str(raw).strip() in ("", "nan", "None"):
        return None

    s = str(raw).strip()
    # Strip trailing footnote markers: digits, *, ¹, ², superscript chars
    s = re.sub(r'[\s\*\d¹²³⁴⁵]+$', '', s).strip()
    s = re.sub(r'\s+\d+$', '', s).strip()

    # Skip pure-numeric, single chars, or meta labels
    if re.match(r'^[\d\.\s\+\-]+$', s) or len(s) < 3:
        return None
    if re.search(r'(mean|lsd|c\.v|rep|row|plot|test|strain|yield|no\.|n mean)', s, re.I):
        return None

    # Modern format: "State. City" or "State City" e.g. "Minn. Morris", "Ont. Ottawa"
    m = re.match(r'^([A-Za-z]{1,5}\.?)\s+(.+)$', s)
    if m:
        possible_state = m.group(1).rstrip(".")
        if possible_state.lower() in STATE_NORM or possible_state.upper() in STATE_NORM.values():
            city  = m.group(2).strip().title()
            state = norm_state(possible_state)
            return city, state

    # Early format: "City State" e.g. "Fargo N.D.", "Eau Claire Wis.", "Mt. Morris Ill."
    # State abbreviation at end (with or without dots)
    m = re.match(r'^(.+?)\s+([A-Za-z]{1,6}\.?)$', s)
    if m:
        possible_state = m.group(2).rstrip(".")
        if possible_state.lower() in STATE_NORM:
            city  = m.group(1).strip().title()
            state = norm_state(possible_state)
            return city, state

    # Two-word state abbrev at end: "N.D.", "S.D.", "N.J.", "N.C."
    m = re.match(r'^(.+?)\s+([A-Z]\.[A-Z]\.)$', s)
    if m:
        state = norm_state(m.group(2))
        city  = m.group(1).strip().title()
        return city, state

    return None


# ---------------------------------------------------------------------------
# Step 1: Collect locations from extracted CSVs
# ---------------------------------------------------------------------------

def collect_from_csvs() -> list[dict]:
    rows = []
    for csv_dir in sorted(CSV_DIRS):
        for loc_csv in csv_dir.glob("combined_*_locationsTable.csv"):
            df = pd.read_csv(loc_csv)
            for _, r in df.iterrows():
                rows.append({
                    "City": str(r["City"]).strip(),
                    "State": str(r["State"]).strip() if not pd.isna(r["State"]) else "",
                    "Source": loc_csv.name,
                })
    print(f"  From extracted CSVs: {len(rows)} location rows")
    return rows


# ---------------------------------------------------------------------------
# Step 2: Scan XLSX headers for location strings
# ---------------------------------------------------------------------------

def scan_xlsx_locations() -> list[dict]:
    rows = []
    per_loc_markers = {"tp6","tp7","tp8","tp9","tp10","tp11a","tp11b","tp12a","tp12b"}

    for year_dir in sorted(GREEN.iterdir()):
        year_m = re.search(r'\b(19|20)\d{2}\b', year_dir.name)
        if not year_m:
            continue
        year = year_m.group(0)
        xlsx_files = sorted(f for f in year_dir.rglob("*.xlsx") if not f.name.startswith("~$"))
        if not xlsx_files:
            continue
        xlsx = xlsx_files[0]  # first file per year is representative

        try:
            wb = openpyxl.load_workbook(str(xlsx), data_only=True, read_only=True)
            ws = wb.active
            prev_tp = None
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                col_a = str(row[0]).strip() if row[0] else ""
                # Track tp marker
                if col_a in per_loc_markers or re.match(r'tp6\+', col_a):
                    prev_tp = col_a
                    continue
                # Row after a per-loc marker = header row with location names
                if prev_tp and col_a == "" and row[1] is not None:
                    # cols C+ are location names
                    for cell in row[2:]:
                        if cell:
                            parsed = parse_location_string(str(cell))
                            if parsed:
                                rows.append({
                                    "City": parsed[0],
                                    "State": parsed[1],
                                    "Source": f"xlsx_scan_{year}",
                                })
                    prev_tp = None
                elif col_a:
                    prev_tp = None
            wb.close()
        except Exception as e:
            print(f"  XLSX scan error {year}: {e}")

    print(f"  From XLSX scans: {len(rows)} location strings parsed")
    return rows


# ---------------------------------------------------------------------------
# Step 3: Geocode with Nominatim
# ---------------------------------------------------------------------------

def geocode_locations(unique_locs: pd.DataFrame) -> pd.DataFrame:
    geolocator = Nominatim(user_agent="nust_location_ref_v1", timeout=10)
    results = []

    for _, row in unique_locs.iterrows():
        city   = row["City"]
        state  = row["State"]
        country = COUNTRY_BY_STATE.get(state, "USA")

        query = f"{city}, {state}, {country}"
        lat, lon, geocoded_city = None, None, None

        try:
            loc = geolocator.geocode(query)
            if loc:
                lat, lon = round(loc.latitude, 6), round(loc.longitude, 6)
                geocoded_city = loc.address.split(",")[0].strip()
            else:
                # Try without state
                loc2 = geolocator.geocode(f"{city}, {country}")
                if loc2:
                    lat, lon = round(loc2.latitude, 6), round(loc2.longitude, 6)
                    geocoded_city = loc2.address.split(",")[0].strip()
        except (GeocoderTimedOut, GeocoderServiceError) as e:
            print(f"    Geocode error for {query}: {e}")

        # Flag known research stations
        city_norm  = city.strip().lower()
        state_norm = state.strip().lower()
        station = KNOWN_STATIONS.get((city_norm, state_norm), "")

        results.append({
            "City":              city,
            "State":             state,
            "Country":           country,
            "lat":               lat,
            "lon":               lon,
            "StationName":       station,
            "Geocoded_City":     geocoded_city,
            "NeedsVerification": 1 if station else 0,
        })

        status = f"{'[STATION] ' if station else ''}{query} -> {lat},{lon}"
        print(f"  {status}", flush=True)
        time.sleep(1.1)  # Nominatim rate limit: 1 req/sec

    return pd.DataFrame(results)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    print("=== Building NUST master location reference ===\n")

    # Collect from all sources
    print("Step 1: Collecting from extracted CSVs...")
    csv_locs = collect_from_csvs()

    print("Step 2: Scanning XLSX files for location headers...")
    xlsx_locs = scan_xlsx_locations()

    all_locs = csv_locs + xlsx_locs
    df = pd.DataFrame(all_locs)

    # Normalise: title-case city, strip trailing commas, upper + normalise state
    df["City"]  = df["City"].str.strip().str.rstrip(",").str.strip().str.title()
    df["State"] = df["State"].str.strip().str.upper().apply(
        lambda s: STATE_NORM.get(s.lower(), s)
    )

    # Drop empties, Mean rows, non-state labels (soil types, etc.)
    real_states = set(STATE_NORM.values()) | {"IL","IN","IA","MN","WI","MI","OH","MO",
                                              "ND","SD","KS","NE","KY","MD","VA","PA",
                                              "NJ","NY","DE","GA","NC","SC","TX","OK",
                                              "AR","TN","AL","MS","LA","WA","OR","CA",
                                              "ONT","MAN","SK","AB","BC","QC","NS","NB"}
    df = df[df["City"].notna() & (df["City"] != "") & (df["City"] != "Mean")]
    df = df[df["City"].str.len() > 1]
    df = df[df["State"].isin(real_states)]  # drop soil-type "states" like CLAY, LOAM

    # Deduplicate (state already normalised, so ILL==IL, IND==IN, etc. collapse)
    unique = (
        df[["City", "State"]]
        .drop_duplicates()
        .sort_values(["State", "City"])
        .reset_index(drop=True)
    )
    print(f"\nTotal unique City+State pairs: {len(unique)}")

    # Save unique list before geocoding (fast checkpoint)
    unique.to_csv("reference/nust_locations_unique.csv", index=False)
    print(f"  Saved to reference/nust_locations_unique.csv")

    # Geocode
    print(f"\nStep 3: Geocoding {len(unique)} locations via Nominatim (~{len(unique)} seconds)...")
    ref = geocode_locations(unique)

    # Save reference table
    out_path = OUT_DIR / "nust_locations_ref.csv"
    ref.to_csv(out_path, index=False)

    n_geocoded  = ref["lat"].notna().sum()
    n_stations  = (ref["NeedsVerification"] == 1).sum()
    n_failed    = ref["lat"].isna().sum()

    print(f"\n=== Summary ===")
    print(f"  Total locations : {len(ref)}")
    print(f"  Geocoded        : {n_geocoded}")
    print(f"  Research stations flagged for verification: {n_stations}")
    print(f"  Failed (lat=NULL): {n_failed}")
    print(f"\nSaved -> {out_path}")
    print("\nNext step: manually verify/correct coordinates for NeedsVerification=1 rows.")


if __name__ == "__main__":
    main()
