#!/usr/bin/env python
"""
standardize_yellow_year.py
==========================
Walk a year's Yellow doc-AI folder once, classify every xlsx by table type,
extract Maturity anchors via a two-tier fallback ladder, and emit a
canonical `Sojabone-{year}_yellow_standardized.xlsx` plus QC summary.

Tier A — per-location Maturity tables: reference-strain row has absolute
M-D calendar dates per city column. High confidence, one anchor per
(Test, City).

Tier B — Regional Summary tables: reference-strain row's Maturity cell
shows the test-level mean calendar date (e.g. "9-19.0", "10-1.5").
Medium confidence; broadcast to every city in the test that lacks a Tier
A anchor.

Output (per year): `analysis/data/yellow_standardized/Sojabone-{year}_yellow_standardized.xlsx`
with sheets:
- `anchors`: one row per (Test, City) — RefStrain, AnchorDate_MD, AnchorDOY,
  Tier (A or B), Source.
- `qc`: header stats + per-Test breakdown + inventory + anomalies.

Plus a cross-year `yellow_standardization_summary.csv` for batch view.

Usage:
    uv run python fixes/standardize_yellow_year.py --year 1972
    uv run python fixes/standardize_yellow_year.py --years 1968,1969,1972
    uv run python fixes/standardize_yellow_year.py --all-yellow
"""
import argparse
import re
import shutil
import sys
import tempfile
import zipfile
from collections import Counter, defaultdict
from pathlib import Path

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
import pandas as pd

# Reuse from sibling scripts
_FIXES_DIR = Path(__file__).parent
_REPO_ROOT = _FIXES_DIR.parent
sys.path.insert(0, str(_FIXES_DIR))
sys.path.insert(0, str(_REPO_ROOT / "scripts"))

from apply_patches_corpus_maturity_doy_via_ocr import (  # type: ignore
    parse_test_from_title,
    _norm,
    _looks_like_calendar,
    _looks_like_offset,
    extract_city_columns,
    detect_reference_row,
    has_maturity_label,
    parse_maturity_table,
    find_source,
    get_working_folder,
    YELLOW_LOCAL,
)
from apply_patches_corpus_maturity_doy import (  # type: ignore
    parse_calendar,
    canonicalize_city,
    DOY_LO, DOY_HI,
)

# Regional summary parser shared with the OCR comparison script
from compare_1975_ocr_vs_claude import (  # type: ignore
    REGIONAL_COL_MAP,
    _norm_header as _norm_regional_header,
)


# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

OUT_DIR_DEFAULT = _REPO_ROOT / "analysis" / "data" / "yellow_standardized"

# Known-good baselines (anchor counts from the API patch). Used for the
# console summary when piloting/comparing.
API_ANCHOR_BASELINES = {
    1972: 173,
    1965: 153,
}


# ---------------------------------------------------------------------------
# Classification
# ---------------------------------------------------------------------------

REGIONAL_RE = re.compile(r"regional\s+summar", re.IGNORECASE)
DESCRIPTIVE_RE = re.compile(r"descriptive|shattering|disease", re.IGNORECASE)
PARENTAGE_RE = re.compile(r"parentage", re.IGNORECASE)
PROTEIN_OIL_RE = re.compile(r"protein|oil", re.IGNORECASE)
SEED_RE = re.compile(r"seed\s+(size|quality|weight)", re.IGNORECASE)
LODGING_RE = re.compile(r"lodging", re.IGNORECASE)
HEIGHT_RE = re.compile(r"height", re.IGNORECASE)
YIELD_RE = re.compile(r"yield", re.IGNORECASE)


def looks_like_city_name(s) -> bool:
    """Heuristic: real city names start with a letter and contain at least
    one alphabetic character. Filters out offsets ('+5', '-2'), numbers
    ('14.1', '157'), and empty/punctuation-only cells that sometimes leak
    through extract_city_columns when header rows are mis-aligned."""
    if s is None:
        return False
    s = str(s).strip()
    if not s:
        return False
    # First non-space char must be a letter (excludes +1, 0, 1.0, etc.)
    if s[0] in "+-0123456789.":
        return False
    # Must have at least one alphabetic character
    if not any(c.isalpha() for c in s):
        return False
    # Common header fragments that aren't cities
    if re.match(r"^(no\.|mean|tests?|c\.\s*v\.|l\.\s*s\.\s*d\.)", s,
                re.IGNORECASE):
        return False
    return True


def classify_table(title: str, rows: list[list]) -> str:
    """Return one of: regional_summary, per_loc_mat, per_loc_yield,
    per_loc_lodging, per_loc_height, per_loc_seed, per_loc_protein_oil,
    descriptive, parentage, multi_year_mean, no_title, other.
    """
    if not title or title.strip().lower() == "no_title":
        return "no_title"
    if re.search(r"\d-year\s+mean|\d{4}\s*-\s*\d{2,4}\s*mean", title, re.IGNORECASE):
        return "multi_year_mean"
    if REGIONAL_RE.search(title):
        return "regional_summary"
    if PARENTAGE_RE.search(title):
        return "parentage"
    if DESCRIPTIVE_RE.search(title):
        return "descriptive"
    # Trait-specific titles (1968-style sentence + 1984+-style suffix)
    if has_maturity_label(rows, title=title):
        return "per_loc_mat"
    if YIELD_RE.search(title):
        return "per_loc_yield"
    if LODGING_RE.search(title):
        return "per_loc_lodging"
    if HEIGHT_RE.search(title):
        return "per_loc_height"
    if SEED_RE.search(title):
        return "per_loc_seed"
    if PROTEIN_OIL_RE.search(title):
        return "per_loc_protein_oil"
    return "other"


# ---------------------------------------------------------------------------
# Tier B — Regional Summary Maturity anchor extraction
# ---------------------------------------------------------------------------

# Map a NUST state-abbreviation hint to a canonical 2-letter code.
# Best-effort — only used to populate the `State` column.
_STATE_ABBR = {
    "il": "IL", "ill": "IL", "illinois": "IL",
    "in": "IN", "ind": "IN", "indiana": "IN",
    "ia": "IA", "iowa": "IA",
    "mn": "MN", "minn": "MN", "minnesota": "MN",
    "mo": "MO", "missouri": "MO",
    "oh": "OH", "ohio": "OH",
    "ks": "KS", "kan": "KS", "kansas": "KS",
    "ne": "NE", "nebraska": "NE",
    "sd": "SD", "s.d": "SD", "s. dakota": "SD", "south dakota": "SD",
    "nd": "ND", "n.d": "ND", "n. dakota": "ND", "north dakota": "ND",
    "wi": "WI", "wisc": "WI", "wisconsin": "WI",
    "mi": "MI", "mich": "MI", "michigan": "MI",
    "ky": "KY", "kentucky": "KY",
    "va": "VA", "virginia": "VA",
    "md": "MD", "maryland": "MD",
    "de": "DE", "delaware": "DE",
    "pa": "PA", "penn": "PA",
    "nj": "NJ", "new jersey": "NJ",
    "ny": "NY", "new york": "NY",
    "ont": "ONT", "ontario": "ONT",
    "man": "MAN", "manitoba": "MAN",
    "sk": "SK", "sask": "SK", "saskatchewan": "SK",
    "qc": "QUE", "que": "QUE", "quebec": "QUE",
}


def find_regional_maturity_column(rows: list[list]) -> int | None:
    """Return col index of the Maturity column in a Regional Summary table."""
    # Build composite header from rows 0+1
    r0 = [_norm(c) for c in rows[0]] if rows else []
    r1 = [_norm(c) for c in rows[1]] if len(rows) > 1 else []
    n_cols = max(len(r0), len(r1))
    while len(r0) < n_cols: r0.append("")
    while len(r1) < n_cols: r1.append("")
    for ci in range(1, n_cols):
        sub, main = r1[ci], r0[ci]
        label = sub if sub else main
        if not label:
            continue
        normed = _norm_regional_header(label)
        # Look up via the regional col map
        for key, canon in REGIONAL_COL_MAP.items():
            if canon == "Maturity" and key in normed:
                return ci
    return None


def parse_regional_maturity_anchor(xlsx_path: Path, title: str,
                                   year: int | None = None) -> dict | None:
    """For a Regional Summary xlsx, return Tier B anchor info:
    {test, ref_strain, anchor_date_md, anchor_doy, source}.
    Returns None if no Maturity column or no reference row found.
    """
    test_code, is_multi = parse_test_from_title(title)
    if not test_code or is_multi:
        return None
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return None

    mat_col = find_regional_maturity_column(rows)
    if mat_col is None:
        return None

    # Find the first data row whose Maturity cell parses as a calendar
    # date (e.g. "9-19.0") — that's the reference strain.
    for r in rows[2:20]:
        if not r or mat_col >= len(r):
            continue
        strain_cell = _norm(r[0])
        mat_cell = _norm(r[mat_col]) if r[mat_col] is not None else ""
        if not strain_cell or not mat_cell:
            continue
        # Skip header-like col-A values
        if re.match(r"^(no\.|c\.\s*v\.|l\.\s*s\.\s*d\.|mean|tests?)\b",
                    strain_cell, re.IGNORECASE):
            continue
        doy = parse_calendar(mat_cell, year)
        if doy is None or not (DOY_LO <= doy <= DOY_HI):
            continue
        # Strain name might carry a "+" ref marker; strip for cleanliness
        ref_clean = strain_cell.rstrip("+").strip()
        return {
            "test": test_code,
            "ref_strain": ref_clean,
            "anchor_date_md": mat_cell.rstrip("+"),
            "anchor_doy": int(doy),
            "source": xlsx_path.name,
        }
    return None


# ---------------------------------------------------------------------------
# Tier A — Per-location Maturity tables (RefStrain + State extraction)
# ---------------------------------------------------------------------------

def parse_maturity_table_full(xlsx_path: Path, title: str,
                              year: int | None = None) -> list[dict]:
    """Like parse_maturity_table() but emits richer per-(Test, City) rows
    with RefStrain, AnchorDate_MD, AnchorDOY, State."""
    test_code, is_multi = parse_test_from_title(title)
    if not test_code or is_multi:
        return []
    try:
        wb = openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
        wb.close()
    except Exception:
        return []
    if not has_maturity_label(rows, title=title):
        return []

    cities = extract_city_columns(rows)
    ref_row_idx = detect_reference_row(rows)
    if ref_row_idx is None or not cities:
        return []

    # Reference strain name (col A of the reference row)
    ref_row = rows[ref_row_idx]
    ref_strain = _norm(ref_row[0]).rstrip("+").strip() if ref_row else ""

    # Build State mapping from row 0 (state spans) — best-effort
    r0 = [_norm(c) for c in rows[0]] if rows else []
    state_for_col: dict[int, str] = {}
    current_state = ""
    for ci, label in enumerate(r0):
        if label:
            # Try to parse as a state abbreviation
            norm = label.lower().rstrip(".").strip()
            if norm in _STATE_ABBR:
                current_state = _STATE_ABBR[norm]
            else:
                # Multi-word fragments — try matching common patterns
                for key in sorted(_STATE_ABBR, key=len, reverse=True):
                    if key in norm:
                        current_state = _STATE_ABBR[key]
                        break
        if ci in cities:
            state_for_col[ci] = current_state

    out = []
    for col_idx, city in cities.items():
        if col_idx >= len(ref_row):
            continue
        if not looks_like_city_name(city):
            continue  # filter out offsets/numbers mis-detected as cities
        cell = _norm(ref_row[col_idx])
        if not cell:
            continue
        doy = parse_calendar(cell, year)
        if doy is None or not (DOY_LO <= doy <= DOY_HI):
            continue
        out.append({
            "Test": test_code,
            "City": city,
            "City_canon": canonicalize_city(city),
            "State": state_for_col.get(col_idx, ""),
            "RefStrain": ref_strain,
            "AnchorDate_MD": cell.rstrip("+"),
            "AnchorDOY": int(doy),
            "Tier": "A",
            "Source": xlsx_path.name,
        })
    return out


# ---------------------------------------------------------------------------
# Year-level orchestration
# ---------------------------------------------------------------------------

def standardize_year(year: int, source: Path, out_dir: Path) -> dict:
    """Walk a year's Yellow folder; emit standardized xlsx; return stats."""
    folder = get_working_folder(source, year)
    print(f"\n=== {year} ===")
    print(f"  Working folder: {folder}")

    xlsx_files = sorted(folder.glob("*.xlsx"))
    title_dir = folder
    n_total = len(xlsx_files)
    print(f"  XLSX files: {n_total}")

    # Pass 1: classify every xlsx
    classifications: dict[str, list[Path]] = defaultdict(list)
    title_for: dict[str, str] = {}
    for xp in xlsx_files:
        tp = title_dir / f"{xp.stem}_title.txt"
        title = tp.read_text(encoding="utf-8").strip() if tp.exists() else ""
        title_for[xp.name] = title
        try:
            wb = openpyxl.load_workbook(xp, data_only=True, read_only=True)
            ws = wb.active
            rows = list(ws.iter_rows(values_only=True))
            wb.close()
        except Exception:
            classifications["error"].append(xp)
            continue
        cls = classify_table(title, rows)
        classifications[cls].append(xp)

    inv = {k: len(v) for k, v in classifications.items()}
    print(f"  Inventory: {dict(sorted(inv.items()))}")

    # Pass 2 — Tier A from per_loc_mat tables
    tier_a_rows: list[dict] = []
    for xp in classifications.get("per_loc_mat", []):
        rows_out = parse_maturity_table_full(xp, title_for.get(xp.name, ""), year)
        tier_a_rows.extend(rows_out)

    # Index Tier A by (test, city_canon) — for Tier B gap-fill
    tier_a_keys = {(r["Test"], r["City_canon"]) for r in tier_a_rows}
    tier_a_cities_per_test: dict[str, set[str]] = defaultdict(set)
    for r in tier_a_rows:
        tier_a_cities_per_test[r["Test"]].add(r["City_canon"])

    # Pass 3 — Tier B from regional summary tables
    tier_b_per_test: dict[str, dict] = {}  # test → anchor dict
    for xp in classifications.get("regional_summary", []):
        a = parse_regional_maturity_anchor(xp, title_for.get(xp.name, ""), year)
        if a:
            # Keep first found per test (Regional Summaries may repeat across pages)
            if a["test"] not in tier_b_per_test:
                tier_b_per_test[a["test"]] = a

    # Pass 4 — Build per-test city universe (from per-loc tables, all
    # traits, not just maturity) so Tier B can broadcast to cities that
    # appear in OTHER trait tables but not in maturity tables.
    test_cities: dict[str, set[tuple[str, str]]] = defaultdict(set)
    # (test, (city, city_canon))
    for cls_key in ("per_loc_mat", "per_loc_yield", "per_loc_lodging",
                      "per_loc_height", "per_loc_seed", "per_loc_protein_oil"):
        for xp in classifications.get(cls_key, []):
            title = title_for.get(xp.name, "")
            test_code, is_multi = parse_test_from_title(title)
            if not test_code or is_multi:
                continue
            try:
                wb = openpyxl.load_workbook(xp, data_only=True, read_only=True)
                ws = wb.active
                rows = list(ws.iter_rows(values_only=True))
                wb.close()
            except Exception:
                continue
            cities = extract_city_columns(rows)
            for ci, city in cities.items():
                if looks_like_city_name(city):
                    test_cities[test_code].add((city, canonicalize_city(city)))

    # Tier B rows — one row per test with City="_test_mean_". The patcher
    # uses this as a fallback when a specific (Test, City) Tier A anchor
    # isn't available. This avoids the "broadcast-to-wrong-cities" bug
    # that arose when extract_city_columns picked up garbage from
    # mis-aligned headers in other trait tables.
    tier_b_rows: list[dict] = []
    for test, anchor in tier_b_per_test.items():
        tier_b_rows.append({
            "Test": test,
            "City": "_test_mean_",
            "City_canon": "_test_mean_",
            "State": "",
            "RefStrain": anchor["ref_strain"],
            "AnchorDate_MD": anchor["anchor_date_md"],
            "AnchorDOY": anchor["anchor_doy"],
            "Tier": "B",
            "Source": anchor["source"],
        })

    # Combined anchors
    anchors_df = pd.DataFrame(tier_a_rows + tier_b_rows)
    if not anchors_df.empty:
        anchors_df = anchors_df.sort_values(["Test", "City_canon", "Tier"]).reset_index(drop=True)

    # QC table
    qc_rows = []
    qc_rows.append({"Section": "header", "Key": "Year", "Value": year})
    qc_rows.append({"Section": "header", "Key": "anchors_total",
                     "Value": len(anchors_df)})
    qc_rows.append({"Section": "header", "Key": "anchors_tier_A",
                     "Value": int((anchors_df["Tier"] == "A").sum()) if not anchors_df.empty else 0})
    qc_rows.append({"Section": "header", "Key": "anchors_tier_B",
                     "Value": int((anchors_df["Tier"] == "B").sum()) if not anchors_df.empty else 0})
    qc_rows.append({"Section": "header", "Key": "distinct_tests",
                     "Value": int(anchors_df["Test"].nunique()) if not anchors_df.empty else 0})
    qc_rows.append({"Section": "header", "Key": "distinct_test_city",
                     "Value": int(anchors_df.groupby(["Test", "City_canon"]).ngroups)
                              if not anchors_df.empty else 0})
    qc_rows.append({"Section": "header", "Key": "n_xlsx_total", "Value": n_total})

    # Inventory by classification
    for k, v in sorted(inv.items()):
        qc_rows.append({"Section": "inventory", "Key": k, "Value": v})

    # Per-test breakdown
    if not anchors_df.empty:
        for test, grp in anchors_df.groupby("Test"):
            qc_rows.append({"Section": "per_test", "Key": f"{test}_anchors_A",
                             "Value": int((grp["Tier"] == "A").sum())})
            qc_rows.append({"Section": "per_test", "Key": f"{test}_anchors_B",
                             "Value": int((grp["Tier"] == "B").sum())})
            qc_rows.append({"Section": "per_test", "Key": f"{test}_distinct_city",
                             "Value": int(grp["City_canon"].nunique())})

    # Anomalies
    anomalies = []
    if not anchors_df.empty:
        out_of_range = anchors_df[~anchors_df["AnchorDOY"].between(DOY_LO, DOY_HI)]
        if not out_of_range.empty:
            anomalies.append(f"{len(out_of_range)} anchors with DOY out of {DOY_LO}-{DOY_HI}")
        suspicious_city = anchors_df[anchors_df["City"].str.len() < 3]
        if not suspicious_city.empty:
            anomalies.append(f"{len(suspicious_city)} anchors with very short city name")
    for a in anomalies:
        qc_rows.append({"Section": "anomalies", "Key": "anomaly", "Value": a})

    qc_df = pd.DataFrame(qc_rows)

    # Write xlsx
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"Sojabone-{year}_yellow_standardized.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        if anchors_df.empty:
            pd.DataFrame(columns=["Test", "City", "City_canon", "State",
                                    "RefStrain", "AnchorDate_MD", "AnchorDOY",
                                    "Tier", "Source"]).to_excel(
                writer, sheet_name="anchors", index=False)
        else:
            anchors_df.to_excel(writer, sheet_name="anchors", index=False)
        qc_df.to_excel(writer, sheet_name="qc", index=False)

    n_a = int((anchors_df["Tier"] == "A").sum()) if not anchors_df.empty else 0
    n_b = int((anchors_df["Tier"] == "B").sum()) if not anchors_df.empty else 0
    print(f"  Anchors: Tier A = {n_a}, Tier B = {n_b}, total = {n_a+n_b}")
    if year in API_ANCHOR_BASELINES:
        baseline = API_ANCHOR_BASELINES[year]
        delta = (n_a + n_b) - baseline
        print(f"  vs API baseline ({baseline}): {'+' if delta >= 0 else ''}{delta}")
    print(f"  Wrote {out_path.name}")

    return {
        "Year": year,
        "n_xlsx_total": n_total,
        "n_per_loc_mat": inv.get("per_loc_mat", 0),
        "n_regional_summary": inv.get("regional_summary", 0),
        "n_anchors_tier_A": n_a,
        "n_anchors_tier_B": n_b,
        "n_distinct_test_city": int(anchors_df.groupby(["Test", "City_canon"]).ngroups)
                                if not anchors_df.empty else 0,
        "n_anomalies": len(anomalies),
    }


def discover_yellow_years() -> list[tuple[int, Path]]:
    """Find every {year}_done folder under YELLOW_LOCAL; return [(year, path)]."""
    found = []
    if not YELLOW_LOCAL.exists():
        return found
    for sub in YELLOW_LOCAL.iterdir():
        if not sub.is_dir():
            continue
        for inner in sub.iterdir():
            m = re.match(r"^(\d{4})_done$", inner.name)
            if m and inner.is_dir():
                found.append((int(m.group(1)), inner))
    return sorted(found)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    ap = argparse.ArgumentParser()
    grp = ap.add_mutually_exclusive_group(required=True)
    grp.add_argument("--year",  type=int, help="Single year to standardize")
    grp.add_argument("--years", type=str, help="Comma-separated list of years")
    grp.add_argument("--all-yellow", action="store_true",
                     help="Standardize every year found in Yellow extract")
    ap.add_argument("--out_dir", default=str(OUT_DIR_DEFAULT),
                    help="Output directory (default: analysis/data/yellow_standardized/)")
    ap.add_argument("--source", default=None,
                    help="Explicit source folder/zip (single-year mode only)")
    args = ap.parse_args()

    out_dir = Path(args.out_dir)

    jobs: list[tuple[int, Path]] = []
    if args.year:
        src = Path(args.source) if args.source else find_source(args.year)
        if src is None:
            sys.exit(f"ERROR: no source for {args.year} found.")
        jobs = [(args.year, src)]
    elif args.years:
        for y in [int(y) for y in args.years.split(",")]:
            src = find_source(y)
            if src:
                jobs.append((y, src))
            else:
                print(f"WARN: no source for {y}; skipping")
    else:  # --all-yellow
        jobs = discover_yellow_years()

    if not jobs:
        sys.exit("ERROR: no jobs to run.")

    print(f"\nStandardizing {len(jobs)} year(s) -> {out_dir}")

    all_stats = []
    for year, src in jobs:
        try:
            stats = standardize_year(year, src, out_dir)
            all_stats.append(stats)
        except Exception as e:
            import traceback
            traceback.print_exc()
            all_stats.append({"Year": year, "error": str(e)})

    summary_df = pd.DataFrame(all_stats)
    summary_csv = out_dir / "yellow_standardization_summary.csv"
    summary_df.to_csv(summary_csv, index=False)
    print(f"\nSummary -> {summary_csv.name}")
    if len(summary_df) > 1:
        print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()
