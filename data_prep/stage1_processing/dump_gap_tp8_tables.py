#!/usr/bin/env python
"""
dump_gap_tp8_tables.py
======================
For every year with remaining Maturity gaps, extract the full tp8
MATURITY data tables from the Sojabone xlsx (or input/output combined
CSV for 1987-88) into a single multi-sheet workbook for visual review.

For each year, the sheet contains:
  - One block per tp8 MATURITY section found in the Sojabone
  - Each block has: section marker, strain header (with city columns),
    all strain rows (offset values), and footer rows (Date planted,
    X matured, Days to mature)
  - Blank rows between blocks

Output: logs/NUST_maturity_gap_tp8_dump.xlsx (one sheet per year)

Usage:
    uv run python fixes/dump_gap_tp8_tables.py
"""
import sys, re
from pathlib import Path
from datetime import datetime, date

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import pandas as pd

_FIXES = Path(__file__).parent
sys.path.insert(0, str(_FIXES))

from extract_sojabone_anchors import (  # type: ignore
    find_sojabone_xlsx_all, REPO_INPUT_ROOT,
    parse_compound_city, _composite_header, _norm, looks_like_city,
    load_test_map,
)
from apply_patches_corpus_maturity_doy import canonicalize_city  # type: ignore

REPO_ROOT = _FIXES.parent
LOGS = REPO_ROOT / "logs"
GAPS_CSV = LOGS / "NUST_maturity_gaps_consolidated.csv"
# Allow override at module load time via env var (for v2 / temp runs)
import os as _os
_alt = _os.environ.get("NUST_GAPS_CSV")
if _alt:
    GAPS_CSV = Path(_alt)

# Years still containing gap cells (from the latest gaps report)
GAP_YEARS = [
    1942, 1943, 1944, 1945, 1946, 1947, 1948, 1949, 1950,
    1951, 1952, 1954, 1955, 1957, 1958, 1959, 1960, 1961, 1962,
    1963, 1964, 1965, 1966, 1967, 1968, 1969, 1970,
]

MAT_RE = re.compile(r"^(.+?)\s+mat[a-z]*\.?$", re.IGNORECASE)
DAYS_RE = re.compile(r"^(days?|da\.)\s+to\s+ma(?:t|tur|ture)?\.?$", re.IGNORECASE)
PL_RE = re.compile(r"^date\s+pl(?:anted|td\.?|t\.?)?$", re.IGNORECASE)


# ---------------------------------------------------------------------------
# Gap-list filtering helpers
# ---------------------------------------------------------------------------

def load_gap_cities_by_year() -> dict[int, dict[str, set[str]]]:
    """Read NUST_maturity_gaps_consolidated.csv and return a nested dict:
        {year: {test_code: set(city_canon)}}
    Empty dict if the CSV is missing."""
    if not GAPS_CSV.exists():
        return {}
    df = pd.read_csv(GAPS_CSV, dtype=str)
    out: dict[int, dict[str, set[str]]] = {}
    for _, row in df.iterrows():
        try:
            y = int(row["Year"])
        except (TypeError, ValueError):
            continue
        t = (row.get("Test") or "").strip()
        c = (row.get("City_canon") or "").strip().lower()
        if not (t and c):
            continue
        out.setdefault(y, {}).setdefault(t, set()).add(c)
    return out


def load_year_gap_cities_flat(gaps_map: dict, year: int) -> set[str]:
    """All canonical cities with remaining Maturity gaps for the year,
    pooled across Tests (used for permissive filtering when block→Test
    attribution is unreliable in the 1942-50 era)."""
    yr = gaps_map.get(year, {})
    out: set[str] = set()
    for cities in yr.values():
        out |= cities
    return out


def block_cities(rows: list, block_start: int) -> set[str]:
    """Return the set of canonical city names in the strain header
    immediately above (or at) `block_start`. Searches upward for the
    closest 'Strain' row within 200 rows."""
    hdr_idx = -1
    lo = max(block_start - 200, 0)
    for j in range(block_start, lo - 1, -1):
        r = rows[j]
        if r and r[0] is not None and _norm(str(r[0])).lower() == "strain":
            n_non_empty = sum(1 for c in r[1:] if c is not None and _norm(c))
            if n_non_empty >= 2:
                hdr_idx = j
                break
    if hdr_idx < 0:
        return set()
    cities: set[str] = set()
    for h in _composite_header(rows, hdr_idx)[1:]:
        if not h or not looks_like_city(h):
            continue
        c = canonicalize_city(parse_compound_city(h))
        if c:
            cities.add(c.lower())
    return cities


def block_test_code(group_idx: int, groups: list[dict]) -> str:
    """Map a Sojabone tp2-group_idx to its test_code via test_map.json.
    Returns '' if not found."""
    if group_idx < 0 or group_idx >= len(groups):
        return ""
    return (groups[group_idx].get("test_code") or "").strip()


def cell_display(v):
    """Render a cell for output: datetime -> 'M-D', else str."""
    if v is None:
        return ""
    if isinstance(v, (datetime, date)):
        return f"{v.month}-{v.day}"
    return str(v)


def find_tp8_blocks(rows: list) -> list[tuple[int, int]]:
    """Return list of (start_row, end_row) for ACTUAL MATURITY blocks.

    1942-1950 era tp markers are not phenotype-aligned — `tp8` can mean
    "Maturity" OR "Yield bushels" OR another trait depending on year/page
    (user-flagged 1948 R153: `tp8 3-year summary Yield bushels`). So we
    can't just include every tp8 marker. Instead, include a block if EITHER:
      (a) The tp marker's label literally contains "matur" / "matured", OR
      (b) The block content contains a "<strain> matured" / "X mat." row
          (Pattern 1, the X-matured-footer detector), OR
      (c) The block content has at least one cell with a M-D / DD-Mon /
          datetime calendar-date in the September-November range AND at
          least 3 strain rows with integer offsets (in-table date pattern).
    """
    import re as _re
    from datetime import datetime as _dt, date as _date
    blocks = []
    n = len(rows)
    # First pass: collect all tp markers + their ranges
    tp_markers = []
    for i, r in enumerate(rows):
        if r and r[0] is not None and _re.match(r"^tp\d+[a-z]?$", str(r[0]).strip()):
            tp_markers.append(i)
    tp_markers.append(n)  # sentinel for last block end

    MAT_FOOTER_RE = _re.compile(r"^(?:Date\s+)?(.+?)\s+[mnh]at[a-z]*\.?$", _re.IGNORECASE)
    DAYS_RE = _re.compile(r"^(days?|da\.)\s+to\s+(ma|na|ha)", _re.IGNORECASE)
    PL_RE = _re.compile(r"^date\s+pl", _re.IGNORECASE)

    for idx in range(len(tp_markers) - 1):
        start = tp_markers[idx]
        end = tp_markers[idx + 1]
        marker_row = rows[start]
        marker_a = str(marker_row[0]).strip()
        marker_label = (str(marker_row[1]).strip().lower()
                        if len(marker_row) > 1 and marker_row[1] is not None
                        else "")

        # (a) Label says Maturity
        if "matur" in marker_label or "matured" in marker_label:
            blocks.append((start, end))
            continue

        # (b)/(c) Inspect block content
        has_mat_footer = False
        sep_oct_dates = 0
        for j in range(start + 1, end):
            r2 = rows[j]
            if not r2:
                continue
            a2 = str(r2[0]).strip() if r2[0] is not None else ""
            # Skip rows that aren't matured-like
            if a2 and MAT_FOOTER_RE.match(a2) and not DAYS_RE.match(a2) and not PL_RE.match(a2):
                # Could be a matured row — confirm content has dates
                n_dates = 0
                for c in r2[1:]:
                    if isinstance(c, (_dt, _date)) and 7 <= c.month <= 12:
                        n_dates += 1
                    elif c is not None:
                        cs = str(c).strip()
                        if _re.match(r"^\d{1,2}[-/]\d{1,2}", cs):
                            n_dates += 1
                if n_dates >= 3:
                    has_mat_footer = True
                    break
            # Count Sept-Oct dates anywhere in body (for in-table-date heuristic)
            for c in r2[1:]:
                if isinstance(c, (_dt, _date)) and 8 <= c.month <= 11:
                    sep_oct_dates += 1
        if has_mat_footer or sep_oct_dates >= 5:
            blocks.append((start, end))
    return blocks


def dump_year_to_sheet(wb, year: int, gaps_map: dict | None = None,
                        only_gaps: bool = False) -> int:
    """Append a sheet for this year containing tp8 MATURITY blocks
    from its Sojabone xlsx(s). Returns # blocks dumped.

    If `only_gaps` is True, blocks are filtered to those whose strain
    header contains at least one city still in the (Year, Test) gap list
    in NUST_maturity_gaps_consolidated.csv. The block label is augmented
    with the matching gap cities for quick triage.
    """
    xps = find_sojabone_xlsx_all(year)
    if not xps:
        return 0

    # Pre-compute the year's gap lookup (per-test + flat) once.
    test_to_gaps: dict[str, set[str]] = (gaps_map or {}).get(year, {}) if gaps_map else {}
    flat_gaps: set[str] = set()
    for _v in test_to_gaps.values():
        flat_gaps |= _v

    # Load test_map for group_idx -> test_code mapping (best-effort)
    try:
        groups = load_test_map(year)
    except Exception:
        groups = []

    # Collect all blocks (across multiple Sojabone files)
    all_blocks = []
    for xp in xps:
        try:
            sw = openpyxl.load_workbook(xp, data_only=True, read_only=True)
            ws = sw["Sheet1"]
            file_rows = list(ws.iter_rows(values_only=True))
            sw.close()
        except Exception as e:
            continue
        # Track group_idx for context
        group_idx = -1
        gi_at = []
        for r in file_rows:
            if r and r[0] is not None and str(r[0]).strip().startswith("tp2"):
                group_idx += 1
            gi_at.append(group_idx)
        for (s, e) in find_tp8_blocks(file_rows):
            all_blocks.append((xp.name, s, e, gi_at[s], file_rows))

    if not all_blocks:
        return 0

    # Optional filter: only blocks whose strain header overlaps the
    # remaining gap-list cities for this year.
    #
    # Tightening strategy:
    #   - If we can attribute the block to a single test_code via group_idx
    #     -> test_map, require (test_code, city) pair membership in the
    #     per-test gap set. Adjacent test codes (UT/PT siblings for the
    #     same MG) are also considered because Sojabone tp8 reference
    #     rows broadcast across MG-sibling tests.
    #   - If group_idx attribution is unavailable (no test_map, year in
    #     1942-50 era where tp markers aren't phenotype-aligned), fall
    #     back to the permissive flat-gap-cities intersection.
    # Hard short-circuit: only_gaps mode with no gap data for this year
    # means there's nothing to surface — skip the year entirely.
    if only_gaps and not test_to_gaps:
        return 0

    annotated_blocks = []  # (fname, s, e, gi, file_rows, matched_pairs)
    if only_gaps and test_to_gaps:
        for fname, s, e, gi, file_rows in all_blocks:
            cities = block_cities(file_rows, s)
            if not cities:
                continue
            tc_block = block_test_code(gi, groups)
            matched_pairs: set[str] = set()
            if tc_block:
                # Build the set of tests whose gap-list we should match
                # against — the block's own test plus MG-sibling tests
                # (e.g. UT-IV + PT-IV + UPT-IV all share the MG IV ref
                # row). MG-siblings = tests sharing the trailing MG token.
                import re as _re2
                m_mg = _re2.search(r"-([0O]+|I+V?|VI*)\s*$", tc_block.upper())
                mg = m_mg.group(1) if m_mg else None
                tests_to_check = set()
                for t in test_to_gaps:
                    if t == tc_block:
                        tests_to_check.add(t)
                    elif mg:
                        m2 = _re2.search(r"-([0O]+|I+V?|VI*)\s*$", t.upper())
                        if m2 and m2.group(1) == mg:
                            tests_to_check.add(t)
                for t in tests_to_check:
                    for c in cities & test_to_gaps.get(t, set()):
                        matched_pairs.add(f"{t}:{c}")
                if not matched_pairs:
                    continue
            else:
                # No test attribution — fall back to flat-city overlap.
                m_flat = cities & flat_gaps
                if not m_flat:
                    continue
                matched_pairs = {f"?:{c}" for c in m_flat}
            annotated_blocks.append((fname, s, e, gi, file_rows, matched_pairs))
        all_blocks_view = annotated_blocks
    elif only_gaps and flat_gaps:
        # No per-test gap info (year not in gaps CSV table_to_gaps) —
        # fall back to permissive flat-city filter
        for fname, s, e, gi, file_rows in all_blocks:
            cities = block_cities(file_rows, s)
            matched = cities & flat_gaps
            if not matched:
                continue
            annotated_blocks.append((fname, s, e, gi, file_rows,
                                     {f"?:{c}" for c in matched}))
        all_blocks_view = annotated_blocks
    else:
        all_blocks_view = [(fname, s, e, gi, file_rows,
                            {f"?:{c}" for c in block_cities(file_rows, s) & flat_gaps})
                           for (fname, s, e, gi, file_rows) in all_blocks]

    if not all_blocks_view:
        return 0

    # Create sheet
    sheet = wb.create_sheet(str(year))
    row_idx = 1

    # Header
    title_suffix = " (gap-overlap only)" if only_gaps else ""
    sheet.cell(row=row_idx, column=1, value=f"Year {year} - tp8 MATURITY blocks{title_suffix}")
    sheet.cell(row=row_idx, column=1).font = Font(bold=True, size=14)
    row_idx += 1
    sheet.cell(row=row_idx, column=1, value=f"Total blocks: {len(all_blocks_view)}, "
              f"source files: {len(set(b[0] for b in all_blocks_view))}, "
              f"year-gap-cities: {len(flat_gaps)}")
    row_idx += 2

    bold = Font(bold=True)
    yellow = PatternFill(start_color="FFFFC0", end_color="FFFFC0", fill_type="solid")
    pink = PatternFill(start_color="FFD0D0", end_color="FFD0D0", fill_type="solid")
    grey = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

    for blk_i, (fname, s, e, gi, rows, matched_pairs) in enumerate(all_blocks_view):
        # Block header
        tc = block_test_code(gi, groups)
        gap_blurb = ""
        if matched_pairs:
            gap_blurb = "  gap_hits=" + ",".join(sorted(matched_pairs))
        block_label = (f"[Block {blk_i+1}] file={fname}  start_row={s}  "
                       f"group_idx={gi}  test={tc or '?'}{gap_blurb}")
        sheet.cell(row=row_idx, column=1, value=block_label).font = bold
        sheet.cell(row=row_idx, column=1).fill = grey
        row_idx += 1

        for j in range(s, e):
            r = rows[j]
            if not r:
                row_idx += 1
                continue
            label = str(r[0]).strip() if r[0] is not None else ""
            # Color-code rows
            is_mat = bool(MAT_RE.match(label)) and not DAYS_RE.match(label) and not PL_RE.match(label)
            is_days = bool(DAYS_RE.match(label))
            is_pl = bool(PL_RE.match(label))
            is_strain_hdr = label.lower() == "strain"
            # Write soruce row number for traceability
            sheet.cell(row=row_idx, column=1, value=f"R{j}")
            for c_idx, c in enumerate(r):
                if c is None:
                    continue
                cell = sheet.cell(row=row_idx, column=c_idx + 2, value=cell_display(c))
                if is_strain_hdr:
                    cell.font = bold
                    cell.fill = yellow
                elif is_mat:
                    cell.fill = pink
                    if c_idx == 0:
                        cell.font = bold
                elif is_days or is_pl:
                    cell.fill = grey
            row_idx += 1
        row_idx += 2  # blank between blocks

    # Autosize column A
    sheet.column_dimensions["A"].width = 7
    sheet.column_dimensions["B"].width = 22
    for col_letter in "CDEFGHIJKL":
        sheet.column_dimensions[col_letter].width = 16
    return len(all_blocks_view)


def dump_combined_csv_year(wb, year: int, gaps_map: dict | None = None,
                            only_gaps: bool = False) -> int:
    """For 1987/1988 (no Sojabone xlsx), pull all Maturity rows from
    output_files/output_{year}/combined_{year}_phenotypesTable.csv organized by Test.

    When `only_gaps` is True, rows are filtered to (Test, canonicalized City)
    pairs that appear in the gap list for this year.
    """
    csv_path = REPO_INPUT_ROOT / f"output_files/output_{year}" / f"combined_{year}_phenotypesTable.csv"
    if not csv_path.exists():
        return 0
    df = pd.read_csv(csv_path, dtype=str)
    mat = df[df["Phenotype"] == "Maturity"].copy()
    if mat.empty:
        return 0

    # Optional gap-only filter
    year_gaps = (gaps_map or {}).get(year, {}) if gaps_map else {}
    flat_gaps: set[str] = set()
    for _v in year_gaps.values():
        flat_gaps |= _v
    # Skip the year entirely if only_gaps is on but the year has no gaps.
    if only_gaps and not year_gaps:
        return 0
    if only_gaps and year_gaps:
        def _is_gap(row):
            t = (row.get("Test") or "").strip()
            c = canonicalize_city(str(row.get("City") or "")).lower()
            # Strict (Test, City) pair match — flat_gaps fallback was too
            # permissive and produced ~3700 rows when only ~30 cells in
            # the year actually need anchoring.
            return t in year_gaps and c in year_gaps[t]
        mat = mat[mat.apply(_is_gap, axis=1)]
        if mat.empty:
            return 0

    sheet = wb.create_sheet(str(year))
    title_suffix = " (gap-overlap only)" if only_gaps else ""
    sheet.cell(row=1, column=1, value=f"Year {year} - Maturity rows{title_suffix} "
               f"(from combined_{year}_phenotypesTable.csv)").font = Font(bold=True, size=14)
    sheet.cell(row=2, column=1, value=f"Total rows: {len(mat)}").font = Font(italic=True)

    # Write header
    cols = ["Test", "Strain", "City", "State", "Value"]
    for ci, c in enumerate(cols):
        cell = sheet.cell(row=4, column=ci + 1, value=c)
        cell.font = Font(bold=True)
    # Write data, sorted by Test then Strain
    mat_sorted = mat.sort_values(["Test", "Strain", "City"])
    for ri, (_, row) in enumerate(mat_sorted.iterrows(), start=5):
        for ci, c in enumerate(cols):
            sheet.cell(row=ri, column=ci + 1, value=str(row.get(c, "")))
    for col_letter in "ABCDE":
        sheet.column_dimensions[col_letter].width = 22
    return len(mat)


def main():
    import argparse
    ap = argparse.ArgumentParser()
    ap.add_argument("--start_year", type=int, default=1942,
                    help="First year to include (default: 1942)")
    ap.add_argument("--end_year", type=int, default=1988,
                    help="Last year to include (default: 1988)")
    ap.add_argument("--out", default=None,
                    help="Output xlsx filename (defaults to "
                         "NUST_maturity_gap_tp8_dump_{start}_{end}.xlsx)")
    ap.add_argument("--only_gaps", action="store_true", default=True,
                    help="(default) Only include blocks whose strain header "
                         "contains a city still in NUST_maturity_gaps_consolidated.csv "
                         "for that year. Filters out the 95%% of blocks that are "
                         "already fully patched.")
    ap.add_argument("--all_blocks", action="store_true", default=False,
                    help="Disable the only_gaps filter and dump every "
                         "Maturity block (legacy behavior).")
    args = ap.parse_args()

    only_gaps = args.only_gaps and not args.all_blocks

    years_in_range = [y for y in range(args.start_year, args.end_year + 1)
                      if y != 1975]
    # When the user explicitly picks the 1942-1988 default + filters by
    # gap, walk EVERY year in that range — the filter will drop years
    # whose blocks no longer overlap any gaps. When all_blocks mode is on
    # and the range is the legacy default, keep the GAP_YEARS shortlist
    # for back-compat with the original dump.
    if args.start_year == 1942 and args.end_year == 1988 and not only_gaps:
        years = GAP_YEARS
    else:
        years = years_in_range

    gaps_map = load_gap_cities_by_year() if only_gaps else {}
    if only_gaps:
        n_gap_years = sum(1 for y in years if y in gaps_map and gaps_map[y])
        n_gap_cells = 0
        if GAPS_CSV.exists():
            df = pd.read_csv(GAPS_CSV, dtype=str)
            try:
                df["Year"] = df["Year"].astype(int)
                df_yr = df[df["Year"].between(args.start_year, args.end_year)]
                n_gap_cells = df_yr["Cells"].astype(int).sum()
            except Exception:
                pass
        print(f"only_gaps=True: {n_gap_years} years with gaps in "
              f"{args.start_year}-{args.end_year}, totaling {n_gap_cells} cells")

    wb = openpyxl.Workbook()
    default = wb.active
    wb.remove(default)

    summary = []
    for year in years:
        if year in (1987, 1988):
            n = dump_combined_csv_year(wb, year, gaps_map=gaps_map, only_gaps=only_gaps)
            summary.append((year, n, "combined_csv"))
        else:
            n = dump_year_to_sheet(wb, year, gaps_map=gaps_map, only_gaps=only_gaps)
            summary.append((year, n, "sojabone_tp8"))

    # Drop years where no blocks survived the filter
    summary_kept = [s for s in summary if s[1] > 0]

    # Default output name reflects year range + filter mode
    suffix = "_only_gaps" if only_gaps else ""
    if args.out:
        out = LOGS / args.out
    elif args.start_year == 1942 and args.end_year == 1988 and not only_gaps:
        out = LOGS / "NUST_maturity_gap_tp8_dump.xlsx"
    else:
        out = LOGS / (f"NUST_maturity_gap_tp8_dump_"
                       f"{args.start_year}_{args.end_year}{suffix}.xlsx")
    wb.save(out)
    print(f"Wrote {out} (years kept: {len(summary_kept)} / {len(summary)})")
    for y, n, src in summary:
        marker = "  " if n > 0 else "--"
        print(f"  {marker} {y}: {n} {src} block(s)/rows")


if __name__ == "__main__":
    main()
