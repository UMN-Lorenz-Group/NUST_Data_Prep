"""tp2 REPAIR SPEC for the test-map deficit years.

A deficit year has FEWER tp2 markers in the Green than it has real tests (from the PDF captions), so
one `Group_N` spans >=2 tests and `apply_test_map` cannot label them (one Group -> one code). The
upstream fix (precedent: archive/scripts/preprocess_1974_f2_add_tp2.py) is to INSERT the missing tp2
marker into an empty column-A row just before the dropped test's block, then RE-EXTRACT.

This tool DERIVES and VERIFIES the insertion point(s) per year, WITHOUT touching the API:
  * reuse derive_test_map to learn which Green section (tp6 row) each true test starts at, and which
    tp2 Group currently owns it;
  * for every Group that owns >1 section, the 2nd..k-th sections are the dropped tests -> the missing
    tp2 goes at the PARENTAGE BOUNDARY of that dropped test, NOT at its yield. `tp2` IS the per-test
    parentage-table marker (extract_nust_xlsx: "tp2: Entry group parentage"), so the boundary must fall
    at/above the dropped test's parentage table -- otherwise the test's own parentage + disease +
    descriptive tables get folded into the PREVIOUS group. `parentage_boundary()` walks up from the
    dropped test's tp6 to the block start (stopping at the previous test's first per-loc marker) and
    anchors on the parentage header when present (modern `Parentage...` OR early `Originating Agency |
    Origin | Generation` dialect), else on the block's topmost pre-yield marker when the parentage
    table is physically absent (OCR-lost page, e.g. 1962 "7_OR Excel ontbreek"). A leftover mislabeled
    marker (tp1/tp3a) directly above a present parentage header is inert -- find_group_boundaries uses
    tp2 only as the group delimiter and the API re-reads the cell grid, so the blank tp2 above it is
    correct and non-destructive.
  * SIMULATE find_group_boundaries with the proposed inserts and assert the new group count == #tests.

`--apply` writes the tp2 markers into the Green (into EMPTY cells only; backs up to
`<file>.bak_pre_tp2`). It does NOT re-extract or rebuild -- those remain the user's (API-spending) step.
The Green edit is inert until re-extraction, so applying is safe and reversible.
"""
import argparse
import importlib
import re
import sys
from pathlib import Path

import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(REPO / "data_prep" / "stage0_extraction"))
A = importlib.import_module("audit_test_map_green")
DTM = importlib.import_module("derive_test_map")
from extract_nust_xlsx import _normalize_tp  # noqa: E402

OUT_SPEC = REPO / "reference" / "tp2_repair_spec.csv"


def tp2_rows(rows):
    return [i + 1 for i, r in enumerate(rows)
            if r and r[0] and _normalize_tp(str(r[0]).strip()) == "tp2"]


def blank_row_above(rows, xlsx_row):
    """Last empty col-A row at or just above `xlsx_row` (1-based) -- the insertion slot."""
    i = xlsx_row
    while i >= 1:
        v = rows[i - 1][0] if i - 1 < len(rows) and rows[i - 1] else None
        if v is None or str(v).strip() == "":
            return i
        i -= 1
    return None


# A test's pre-yield tables (parentage/summaries/descriptive/disease) all carry these markers;
# everything from the yield onward (tp6..tp12b) is a PER-LOCATION table. Walking UP from a test's
# tp6, the first per-loc marker we meet belongs to the PREVIOUS test -> it delimits the block start.
PRE_YIELD_MARKERS = {"tp1", "tp2", "tp3a", "tp3b", "tp4", "tp5", "tp5a", "tp5b", "tp5d"}
PER_LOC_MARKERS = {"tp6", "tp7", "tp8", "tp9", "tp10", "tp11a", "tp11b", "tp12a", "tp12b"}

# Parentage-table HEADER dialects (era-aware). A header row matches when it carries a NAME token AND a
# QUALIFIER token. Modern (1970s-80s): "Strain | Parentage | Previous Testing* | Generation Composited".
# Early (1950s-60s): "Strain | Originating Agency | Origin | Generation Composited" (no word "Parentage").
_PARENTAGE_NAME = re.compile(r"parentage|originating\s+agency", re.I)
_PARENTAGE_QUAL = re.compile(r"previous|generation|\bline\b|\bsource\b|\borigin\b", re.I)


def _norm_a(row):
    """Canonical tp-marker in column A of `row`, or None."""
    v = row[0] if row else None
    if v is None:
        return None
    s = str(v).strip()
    return _normalize_tp(s) if s else None


def _marker_class(row):
    """Classify column A as 'perloc' (tp6..tp12b), 'preyield' (parentage/summary/desc/disease), or None.

    Because per-location tables (tp6..tp12b) only ever appear at/after a test's yield, ANY `tp`-prefixed
    token found ABOVE the yield is a pre-yield table marker -- so a garbled/placeholder token the
    canonicaliser can't resolve (e.g. the literal "tp??" the OCR writes for a smudged marker, 1953) is
    still treated as a pre-yield block marker rather than skipped.
    """
    v = row[0] if row else None
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    n = _normalize_tp(s)
    if n in PER_LOC_MARKERS:
        return "perloc"
    if n in PRE_YIELD_MARKERS:
        return "preyield"
    if s.lower().startswith("tp"):     # garbled/unknown tp-token above the yield -> pre-yield table
        return "preyield"
    return None


def _row_text(row):
    return " | ".join(str(v) for v in row if v is not None) if row else ""


def parentage_boundary(rows, green_row):
    """Insertion slot for a dropped test whose deriver `green_row` is at/near its tp6 yield.

    tp2 IS the parentage marker, so the boundary must land at/above the dropped test's parentage
    table. Walk UP from the test's tp6 to the block start (stop at the previous test's first per-loc
    marker) and anchor on the parentage header if present (either dialect), else on the block's topmost
    pre-yield marker (parentage physically absent). Return (slot, kind) where kind records the anchor.
    """
    # locate this section's tp6 near green_row (deriver row can be off by ~1)
    tp6_row = None
    for j in range(max(1, green_row - 2), min(len(rows), green_row + 4) + 1):
        if _norm_a(rows[j - 1] if j - 1 < len(rows) else None) == "tp6":
            tp6_row = j
            break
    if tp6_row is None:
        tp6_row = green_row

    topmost = None   # block's topmost pre-yield marker
    hdr = None       # parentage header row, if present
    dialect = None
    i = tp6_row - 1
    while i >= 1:
        row = rows[i - 1] if i - 1 < len(rows) else None
        cls = _marker_class(row)
        if cls == "perloc":             # previous test's per-loc table -> block start passed
            break
        if cls == "preyield":
            topmost = i
        txt = _row_text(row)
        if _PARENTAGE_NAME.search(txt) and _PARENTAGE_QUAL.search(txt):
            hdr = i
            dialect = "modern" if re.search(r"parentage", txt, re.I) else "early"
        i -= 1

    anchor = hdr if hdr else topmost
    if anchor is None:
        return None, "no-anchor"
    slot = blank_row_above(rows, anchor - 1)
    kind = f"parentage-header({dialect})" if hdr else "block-start-marker"
    if slot is None or anchor - slot > 8:   # sanity: blank should sit just above the anchor
        return slot, f"WARN-far-blank:{kind}"
    return slot, kind


def spec_for_year(year):
    df, pdf_secs, green_secs, _un_g, _un_p, testmaps, n_groups = DTM.derive(year)
    codes = [p["code"] for p in pdf_secs]
    if not green_secs or n_groups == len(codes):
        return None, f"{year}: no tp2 deficit (Groups {n_groups} == tests {len(codes)})", []
    if n_groups > len(codes):
        return None, f"{year}: tp2 SURPLUS ({n_groups} > {len(codes)}) -- not an insertion case", []

    # cache each Green file's rows once
    filerows = {}
    for f in A.green_files(year):
        wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        filerows[Path(f).name] = list(wb["Sheet1"].iter_rows(values_only=True))
        wb.close()

    # which Group owns each section, and the section's tp6 row (0-based list idx -> xlsx row +1)
    by_group = {}
    for r in df.itertuples():
        if pd.isna(r.group_n):
            continue
        by_group.setdefault(int(r.group_n), []).append(r)
    inserts = []
    for g, secs in sorted(by_group.items()):
        secs = sorted(secs, key=lambda s: s.green_row)
        for extra in secs[1:]:                     # 2nd.. section in this Group == a dropped test
            rows = filerows[extra.green_file]
            slot, boundary = parentage_boundary(rows, extra.green_row)  # parentage, NOT yield
            inserts.append(dict(Year=year, green_file=extra.green_file, insert_row=slot,
                                before_code=extra.true_code, before_tp6_row=extra.green_row,
                                boundary=boundary, shared_group=g))

    # Dropped tests with NO group_n: the section landed in a Green file's PRE-tp2 "GlobalParentage"
    # region because its parentage opener is a mislabeled tp1 at the head of a split file (e.g. 1985
    # UT-III sits at the top of file 2 with tp1@3, so combine folds it into GlobalParentage and the 12
    # tp2-groups are the OTHER 12 tests). Insert a tp2 at its parentage boundary so it becomes its own
    # first group; the section's data is already fully in the Green (no PDF re-extraction needed).
    for r in df.itertuples():
        if pd.notna(r.group_n):
            continue
        if pd.isna(r.green_row) or r.green_file not in filerows:
            continue
        ov = getattr(r, "roster_overlap", None)
        if pd.notna(ov) and ov < 0.5:              # skip junk / low-confidence unmapped rows
            continue
        rows = filerows[r.green_file]
        slot, boundary = parentage_boundary(rows, int(r.green_row))
        inserts.append(dict(Year=year, green_file=r.green_file, insert_row=slot,
                            before_code=r.true_code, before_tp6_row=int(r.green_row),
                            boundary=boundary, shared_group="pre-tp2"))
    # SIMULATE: add the inserts, recount tp2 per file, sum across files
    sim = 0
    for fn, rows in filerows.items():
        base = set(tp2_rows(rows))
        base |= {d["insert_row"] for d in inserts if d["green_file"] == fn and d["insert_row"]}
        sim += len(base)
    ok = (sim == len(codes))
    msg = (f"{year}: {n_groups} Groups + {len(inserts)} inserts = {sim} vs {len(codes)} tests "
           f"-> {'OK' if ok else 'MISMATCH'}")
    return inserts, msg, filerows


def apply_inserts(inserts, filerows):
    for fn in {d["green_file"] for d in inserts}:
        path = next(p for p in _all_green() if Path(p).name == fn)
        wb = openpyxl.load_workbook(path)
        ws = wb["Sheet1"]
        bak = Path(path).with_suffix(".xlsx.bak_pre_tp2")
        if not bak.exists():
            import shutil
            shutil.copy2(path, bak)
        for d in [x for x in inserts if x["green_file"] == fn]:
            cell = ws.cell(row=d["insert_row"], column=1)
            if cell.value not in (None, "") and str(cell.value).strip():
                print(f"    !! {fn} row {d['insert_row']} NOT empty ({cell.value!r}) -- skipped")
                continue
            cell.value = "tp2"
            print(f"    {fn} row {d['insert_row']}: inserted tp2 (before {d['before_code']})")
        wb.save(path)


_GREEN_CACHE = {}


def _all_green():
    if "x" not in _GREEN_CACHE:
        seen = []
        for y in range(1941, 1989):
            for f in A.green_files(y):
                if f not in seen:
                    seen.append(f)
        _GREEN_CACHE["x"] = seen
    return _GREEN_CACHE["x"]


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("years", nargs="+", type=int)
    ap.add_argument("--apply", action="store_true")
    args = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    all_specs = []
    for y in args.years:
        inserts, msg, filerows = spec_for_year(y)
        print(msg)
        if inserts:
            for d in inserts:
                print(f"    -> insert tp2 in {d['green_file']} at row {d['insert_row']} "
                      f"(before {d['before_code']} @tp6 row {d['before_tp6_row']}, "
                      f"boundary={d['boundary']}, splitting Group_{d['shared_group']})")
            all_specs += inserts
            if args.apply:
                apply_inserts(inserts, filerows)
    if all_specs:
        pd.DataFrame(all_specs).to_csv(OUT_SPEC, index=False)
        print(f"\nwrote {OUT_SPEC.relative_to(REPO)} ({len(all_specs)} inserts)")
        if not args.apply:
            print("(dry run — pass --apply to write the tp2 markers into the Green; "
                  "re-extraction + rebuild remain the user's API step)")


if __name__ == "__main__":
    main()
