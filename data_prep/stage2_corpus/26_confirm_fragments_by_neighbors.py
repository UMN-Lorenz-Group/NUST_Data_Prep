"""
26_confirm_fragments_by_neighbors.py
====================================
Confirm the orphaned-fragment corrections (script 24 HIGH safe_ocr, + any chosen tier)
by the SOURCE-TABLE NEIGHBOUR method: in the original tp2 XLSX the tables preserve row
order, and a true OCR garble leaves the strain at the SAME row position — so the garbled
code's neighbours (the strain IDs immediately above/below it) must match the proposed
parent's neighbours in the other tables. This is the value/position method that resolved
250->C160, generalised to a batch gate.

For each (fragment, year, parent):
  * collect the first-cell neighbours (+/-2 rows) of every occurrence of the fragment and of
    the parent across all sheets of the year's source XLSX,
  * CONFIRMED if they share >= MIN_SHARED neighbour IDs (excluding the two codes themselves),
    else REVIEW (neighbours don't line up — eyeball before applying).

Read-only. Input: orphaned_trait_fragments.csv (HIGH + safe_ocr). Output:
analysis/data/analysis_results/Corpus_QC/fragment_neighbor_confirmation.{csv,md}
"""
import sys
import re
import glob
from pathlib import Path
import pandas as pd
import openpyxl


def norm(s):
    return re.sub(r"[^a-z0-9]", "", str(s).lower())

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO = Path("C:/Users/vramasub/Desktop/UMN_GIT/NUST_Data_Prep")
INPUT = REPO / "input_files"
QC = REPO / "analysis" / "data" / "analysis_results" / "Corpus_QC"
MIN_SHARED = 2
WINDOW = 2


def year_xlsx(year):
    return sorted(set(glob.glob(str(INPUT / f"input_{year}" / "*.xlsx"))
                      + glob.glob(str(INPUT / f"input_{year}" / str(year) / "*.xlsx"))))


def neighbours(year, code, cache):
    """First-cell IDs within +/-WINDOW rows of every occurrence of `code` (col 0), any sheet."""
    if year not in cache:
        firsts_by_sheet = []
        for p in year_xlsx(year):
            try:
                wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
            except Exception:
                continue
            for sh in wb.sheetnames:
                firsts = [(str(r[0]).strip() if r and r[0] is not None else "")
                          for r in wb[sh].iter_rows(values_only=True)]
                firsts_by_sheet.append(firsts)
        cache[year] = firsts_by_sheet
    out, nc = [], norm(code)
    for firsts in cache[year]:
        for i, fc in enumerate(firsts):
            if norm(fc) == nc:                          # normalized: tolerate spacing/punctuation
                out.append([x for x in firsts[max(0, i - WINDOW):i + WINDOW + 1] if norm(x) != nc])
    return out


def main():
    src = QC / "orphaned_trait_fragments.csv"
    d = pd.read_csv(src, keep_default_na=False)
    # run on the whole HIGH tier: safe_ocr (glyph) AND check_digit (a digit differs -> the
    # neighbour match is decisive: same row position => same line (OCR digit error); different
    # position => a genuine adjacent SISTER selection to leave alone).
    sub = d[d.confidence == "HIGH"].copy()

    cache, rows = {}, []
    for r in sub.itertuples():
        frag, parent, year = str(r.fragment), str(r.proposed_parent), int(r.Year)
        fn = neighbours(year, frag, cache)
        pn = neighbours(year, parent, cache)
        # compare neighbour IDs by normalized form (neighbours may be spelt slightly differently)
        fmap = {norm(x): x for w in fn for x in w}
        pset = {norm(x) for w in pn for x in w}
        shared = sorted(fmap[k] for k in (set(fmap) & pset) if k)
        status = ("CONFIRMED" if len(shared) >= MIN_SHARED
                  else "REVIEW_no_neighbors" if not fn or not pn
                  else "REVIEW_mismatch")
        rows.append({"Year": year, "MG": r.MG, "fragment": frag, "trait": r.trait,
                     "proposed_parent": parent, "match_class": r.match_class,
                     "n_shared": len(shared), "shared_neighbors": ", ".join(shared[:6]),
                     "status": status})

    out = pd.DataFrame(rows).sort_values(["match_class", "status", "Year", "fragment"])
    QC.mkdir(parents=True, exist_ok=True)
    out.to_csv(QC / "fragment_neighbor_confirmation.csv", index=False)
    L = ["# Fragment corrections — source-table neighbour confirmation (safe_ocr tier)\n",
         f"{len(out)} safe_ocr fragments. {out.status.value_counts().to_dict()}\n",
         "CONFIRMED = the garbled code and its proposed parent share >=2 source-table neighbour "
         "IDs (same row position). REVIEW = neighbours absent or don't line up — eyeball.\n",
         "| Year | MG | fragment | -> parent | trait | shared | neighbours | status |",
         "|------|----|----------|-----------|-------|-------:|-----------|--------|"]
    for _, r in out.iterrows():
        L.append(f"| {r.Year} | {r.MG} | {r.fragment} | {r.proposed_parent} | {r.trait} "
                 f"| {r.n_shared} | {r.shared_neighbors} | {r.status} |")
    (QC / "fragment_neighbor_confirmation.md").write_text("\n".join(L), encoding="utf-8")

    print(f"{len(out)} HIGH fragments — neighbour confirmation by class:")
    print(out.groupby(["match_class", "status"]).size().to_string())
    rev = out[out.status != "CONFIRMED"]
    if len(rev):
        print(f"\nNOT auto-confirmed ({len(rev)}) — likely sister selections / eyeball:")
        print(rev[["Year", "MG", "fragment", "proposed_parent", "trait", "match_class", "n_shared"]]
              .to_string(index=False))
    print(f"\nWrote fragment_neighbor_confirmation.csv + .md to {QC.name}/")


if __name__ == "__main__":
    main()
