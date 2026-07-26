"""Green-direct (openpyxl, NO API) re-extraction of four DROPPED early-era Preliminary Test (PT)
sections whose tp2 openers were missing and have now been inserted, so the Green OCR XLSX delimits
them correctly:

    1953 PT-IV   input_files/input_1953/1953-Sojabone (0-67 OR).xlsx   rows 1586-1656
    1961 PT-III  input_files/input_1961/1961-Sojabone (0-106 OR).xlsx  rows 1858-1964
    1962 PT-00   input_files/input_1962/1962-Sojabone (0-95 OR).xlsx   rows 130-189
    1962 PT-0    input_files/input_1962/1962-Sojabone (0-95 OR).xlsx   rows 391-455

Part of the NUST test-map source-repair (C1, no-API path).  These are small early-1950s/60s
preliminary tests: CLEANER than the 1970s Green (single location-group per trait, clean "City State"
headers, full/period state names), but with three early-era wrinkles that this parser handles:

  1. COMBINED trait markers.  1953 PT-IV prints Maturity and Oil under ONE marker `tp8+12b`, in a
     single table with TWO "Mean of N Tests" columns -- one segment per sub-trait.  A marker name is
     split on '+' into sub-codes (tp8, tp12b); the header is split into SEGMENTS at each "Mean ...
     Tests" column; segment i feeds sub-trait i.

  2. EMBEDDED Yield-Rank block.  1953 PT-IV `tp6` (unlike 1961/62 which break rank out as `tp7`)
     appends the per-location Yield RANK columns to the SAME table, re-using the location headers.
     Within a segment the location columns are de-duplicated (keep first), which drops the rank block.

  3. "Mean of N Tests" with N < number of location columns.  The printed mean routinely excludes the
     western irrigated nurseries (e.g. "Ontario Ore.").  We EXTRACT every location (they are real
     observations) but the reconcile uses a small drop-set search over subsets of size N to identify
     the excluded columns and validate the row-mean.

EXTRACTED per-loc trait markers (only those present per section):
    tp6=YieldBuA  tp8=Maturity  tp9=Lodging  tp10=Height
    tp11a=SeedQuality  tp11b=SeedSize  tp12a=Protein  tp12b=Oil
SKIPPED: tp7 (Yield Rank), tp3b (disease), tp2 (parentage), tp4/tp5 "Summary" (per-strain trait means,
no locations -- used ONLY to cross-report), and 1953's Group-wise protein/oil REGIONAL tables (means
by location for a whole maturity group, not per strain; bounded out by section_end).

Maturity (tp8) is stored AS THE PRINTED SIGNED DAY OFFSET (relative to a check, e.g. Shelby/Grant=0);
NO DOY reconstruction here (that is a separate downstream step).  "--"/blank cells are skipped.

Output: reextract_early_pt_green.csv  (long schema shared with the other stage2 green re-extractions)
        Year,TestType,TestMG,Test,Strain,City,State,Phenotype,Value_num,Units,Source
"""
import sys
import re
from pathlib import Path
from itertools import combinations
from math import comb
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")

REPO = Path("C:/Users/vramasub/Desktop/UMN_GIT/NUST_Data_Prep")
OUT = REPO / "data_prep" / "stage2_corpus" / "reextract_early_pt_green.csv"

F1953 = REPO / "input_files" / "input_1953" / "1953-Sojabone (0-67 OR).xlsx"
F1961 = REPO / "input_files" / "input_1961" / "1961-Sojabone (0-106 OR).xlsx"
F1962 = REPO / "input_files" / "input_1962" / "1962-Sojabone (0-95 OR).xlsx"

# Each section: file, (start_row, end_row) 1-based inclusive bound, Year, TestMG, Test-code.
# end_row is the LAST row that may hold this section's trait data (bounds the final trait block so the
# next test's parentage / the 1953 Group regional tables do NOT leak in).
SECTIONS = [
    dict(name="1953 PT-IV",  file=F1953, a=1586, b=1656, year=1953, mg="IV",  test="PT-IV"),
    dict(name="1961 PT-III", file=F1961, a=1858, b=1964, year=1961, mg="III", test="PT-III"),
    dict(name="1962 PT-00",  file=F1962, a=130,  b=189,  year=1962, mg="00",  test="PT-00"),
    dict(name="1962 PT-0",   file=F1962, a=391,  b=455,  year=1962, mg="0",   test="PT-0"),
]

TP2TRAIT = {"tp6": "YieldBuA", "tp8": "Maturity", "tp9": "Lodging", "tp10": "Height",
            "tp11a": "SeedQuality", "tp11b": "SeedSize", "tp12a": "Protein", "tp12b": "Oil"}
UNITS = {"YieldBuA": "bu/a", "Maturity": "offset_days", "Lodging": "score", "Height": "in",
         "SeedQuality": "score", "SeedSize": "g/100", "Protein": "%", "Oil": "%"}
# Maturity offsets are unbounded -> no range gate (handled by omission from RANGE).
RANGE = {"YieldBuA": (2, 120), "Height": (5, 80), "Lodging": (1, 5), "SeedQuality": (1, 5),
         "SeedSize": (5, 40), "Protein": (25, 55), "Oil": (5, 30)}

FOOTER = re.compile(
    r"^\s*(C\.?\s*V|L\.?\s*S\.?\s*D|Mean\b|Row\s*Sp|Row\s*Spacing|No\.?\s*of\s*Tests|Bu\.?\s*Nec|"
    r"Coef\.?\s*of\s*Var|Date\s*planted|Days?\s*to\s*mat|.*\bmat(?:ured|\.)|Strain\b|tp[\d?]|"
    r"\d+_OR|Group\b|Location\b|Identity\b|N\.?S\.?$)", re.I)
MEANCOL = re.compile(r"Mean\b.*Test|Mean\s+of", re.I)
NTESTS = re.compile(r"Mean\s+of\s+([\d]+)\s+Test", re.I)
MARKER = re.compile(r"^tp[\d?]+[ab]?(?:\+\S+)?$", re.I)
FOOTNOTE = re.compile(r"[*¹²³†‡]")

# state token (as printed, periods kept) -> USPS / Canadian 2-letter code
STATE_MAP = {
    "Ohio": "OH", "Ind.": "IN", "Ill.": "IL", "Iowa": "IA", "Ia.": "IA", "Mo.": "MO",
    "Nebr.": "NE", "Neb.": "NE", "Kans.": "KS", "Kan.": "KS", "Mich.": "MI", "Wis.": "WI",
    "Wash.": "WA", "Ore.": "OR", "Minn.": "MN", "Pa.": "PA", "Del.": "DE", "Md.": "MD",
    "Man.": "MB", "Ont.": "ON", "S.D.": "SD", "N.D.": "ND", "S.C.": "SC", "N.C.": "NC",
    "N.J.": "NJ", "Ky.": "KY", "Tenn.": "TN", "Ark.": "AR", "Va.": "VA", "Ga.": "GA",
}


def load(path):
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    sn = "Sheet1" if "Sheet1" in wb.sheetnames else wb.sheetnames[0]
    rows = list(wb[sn].iter_rows(values_only=True))
    wb.close()
    return rows


def strip_rank(s):
    return re.sub(r"^\s*\d+\s*[.)]\s*", "", str(s)).strip()


def clean_strain(s):
    return FOOTNOTE.sub("", strip_rank(s)).strip()


def fnum(x):
    """Parse a leading signed number from a cell (int/float or OCR string). None if not numeric."""
    if isinstance(x, bool):
        return None
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", ".")
    if s in ("--", "-", "", "N.S.", "NS"):
        return None
    m = re.match(r"^\s*([+-]?\d+(?:\.\d+)?)", s)
    return float(m.group(1)) if m else None


def parse_loc(h):
    """'East Lansing Mich.' -> ('East Lansing','MI'); 'Portage la Prairie Man.' -> (...,'MB');
    'Worthing-ton Ind.' -> ('Worthington','IN').  State = last whitespace token (periods kept)."""
    s = FOOTNOTE.sub("", str(h)).strip()
    s = re.sub(r"\s+", " ", s)
    toks = s.split(" ")
    if len(toks) < 2:
        return None, None
    st_tok = toks[-1]
    city = " ".join(toks[:-1]).strip()
    city = re.sub(r"(?<=[A-Za-z])-(?=[A-Za-z])", "", city)      # de-hyphenate OCR word-splits
    state = STATE_MAP.get(st_tok)
    if state is None:                                          # fall back: strip periods, upper
        bare = st_tok.replace(".", "").upper()
        state = STATE_MAP.get(st_tok + ".", bare)
    return city, state


def markers(rows, a, b):
    """Return [(idx0, raw_marker)] for every tp marker cell in [a,b) (0-based idx)."""
    out = []
    for i in range(a, b):
        c = rows[i]
        first = str(c[0]).strip() if c and c[0] is not None else ""
        if MARKER.match(first):
            out.append((i, first))
    return out


def marker_subtraits(raw):
    """'tp8+12b' -> ['Maturity','Oil'];  'tp6' -> ['YieldBuA'];  unmapped/garbled -> []."""
    parts = [p if p.lower().startswith("tp") else "tp" + p for p in raw.lower().split("+")]
    return [TP2TRAIT[p] for p in parts if p in TP2TRAIT]


def parse_header_segments(hdr):
    """Split a header row into per-trait SEGMENTS at each 'Mean of N Tests' column.
    Returns [ {mean_col, n_tests, locs:[(col, city, state)]} ] with per-segment location de-dup."""
    mean_cols = [ci for ci in range(1, len(hdr))
                 if hdr[ci] is not None and MEANCOL.search(str(hdr[ci]))]
    if not mean_cols:
        return []
    segs = []
    for si, mc in enumerate(mean_cols):
        end = mean_cols[si + 1] if si + 1 < len(mean_cols) else len(hdr)
        m = NTESTS.search(str(hdr[mc]))
        n_tests = int(m.group(1)) if m else None
        locs, seen = [], set()
        for ci in range(mc + 1, end):
            h = hdr[ci]
            if h is None or MEANCOL.search(str(h)):
                continue
            city, state = parse_loc(h)
            if not city:
                continue
            key = (city, state)
            if key in seen:                    # embedded rank block re-uses loc headers -> drop
                continue
            seen.add(key)
            locs.append((ci, city, state))
        segs.append(dict(mean_col=mc, n_tests=n_tests, locs=locs))
    return segs


def data_rows(rows, hdr_idx, end):
    """Yield (strain_name, row_tuple) for real strain rows below a header, footers filtered."""
    for ri in range(hdr_idx + 1, end):
        row = rows[ri]
        if not row or all(v is None for v in row):
            continue
        head = str(row[0]).strip() if row[0] is not None else ""
        if not head or FOOTER.match(head):
            continue
        yield clean_strain(head), row


def reconcile(strain_means, strain_vals, n_tests, n_locs, one_dec_hint):
    """Best drop-set reconcile: search subsets of size n_tests (drop k = n_locs-n_tests columns,
    same drop-set for all strains) maximizing strains whose kept-loc row-mean == printed mean.
    Returns (best_ok, total, dropped_col_positions)."""
    k = (n_locs - n_tests) if (n_tests and 0 < n_tests <= n_locs) else 0
    drops = list(combinations(range(n_locs), k)) if (0 < k and comb(n_locs, k) <= 2000) else [()]
    best = (-1, 0, ())
    for drop in drops:
        keep = [c for c in range(n_locs) if c not in drop]
        ok = tot = 0
        for s, mean in strain_means.items():
            if mean is None:
                continue
            vals = [v for (pos, v) in strain_vals.get(s, []) if pos in keep]
            if not vals:
                continue
            tot += 1
            avg = sum(vals) / len(vals)
            tol = 0.15 if one_dec_hint else 0.56
            if abs(avg - mean) <= tol:
                ok += 1
        if ok > best[0]:
            best = (ok, tot, drop)
    return best


def process(sec, rows):
    a0, b0 = sec["a"] - 1, sec["b"]          # 0-based half-open bound of the section
    mk = markers(rows, a0, b0)
    # parentage roster (first tp2 with an Originating-Agency header) for reporting only
    roster = []
    for (i, raw) in mk:
        if raw.lower() == "tp2" and i + 1 < len(rows):
            h = rows[i + 1]
            if h and any(x and "Origin" in str(x) for x in h):
                for ri in range(i + 2, b0):
                    r = rows[ri]
                    if not r or r[0] is None or MARKER.match(str(r[0]).strip()) or str(r[0]).strip() == "":
                        break
                    if FOOTER.match(str(r[0]).strip()):
                        continue
                    roster.append(clean_strain(r[0]))
                break

    recs, report = [], []
    for gi, (i, raw) in enumerate(mk):
        subtraits = marker_subtraits(raw)
        if not subtraits:
            continue
        end = mk[gi + 1][0] if gi + 1 < len(mk) else b0
        segs = parse_header_segments(rows[i + 1])
        if not segs:
            report.append((raw, subtraits, "NO header segments (skipped)"))
            continue
        n = min(len(subtraits), len(segs))
        drows = list(data_rows(rows, i + 1, end))
        for si in range(n):
            trait = subtraits[si]
            seg = segs[si]
            locs = seg["locs"]
            mean_col = seg["mean_col"]
            n_tests = seg["n_tests"]
            strain_means, strain_vals, kept = {}, {}, 0
            for sname, row in drows:
                pm = fnum(row[mean_col]) if mean_col < len(row) else None
                strain_means[sname] = pm
                vlist = []
                for pos, (ci, city, state) in enumerate(locs):
                    if ci >= len(row):
                        continue
                    v = fnum(row[ci])
                    if v is None:
                        continue
                    if trait in RANGE:
                        lo, hi = RANGE[trait]
                        if not (lo <= v <= hi):
                            print(f"    ! {sec['test']} {trait} {sname} @ {city},{state} "
                                  f"out of range {v} -> dropped")
                            continue
                    vlist.append((pos, v))
                    recs.append((sec["year"], "PT", sec["mg"], sec["test"], sname, city, state,
                                 trait, round(v, 1), UNITS[trait],
                                 f"Green{sec['year']}_{sec['test']}_direct"))
                    kept += 1
                strain_vals[sname] = vlist
            one_dec = trait != "Maturity"          # maturity offsets can be integer-meaned
            ok, tot, drop = reconcile(strain_means, strain_vals, n_tests, len(locs),
                                      one_dec_hint=one_dec)
            dropped = [locs[d][1] + " " + (locs[d][2] or "?") for d in drop]
            report.append((raw, trait, dict(n_locs=len(locs), n_tests=n_tests, n_rows=len(drows),
                                            n_vals=kept, recon=(ok, tot), dropped=dropped,
                                            locs=[c for (_, c, _) in locs])))
    return roster, recs, report


def main():
    all_recs = []
    print("=" * 78)
    for sec in SECTIONS:
        rows = load(sec["file"])
        roster, recs, report = process(sec, rows)
        all_recs.extend(recs)
        print(f"\n### {sec['name']}   (rows {sec['a']}-{sec['b']})")
        print(f"  parentage roster ({len(roster)}): {', '.join(roster) if roster else '(from data rows)'}")
        per_loc = [r for r in report if isinstance(r[2], dict) and r[2].get("n_vals")]
        for raw, trait, info in report:
            if not isinstance(info, dict):
                print(f"  {raw:9} {trait:12} -> {info}")
                continue
            ok, tot = info["recon"]
            drp = f"  [excl {', '.join(info['dropped'])}]" if info["dropped"] else ""
            print(f"  {raw:9} {trait:12} locs={info['n_locs']:2} (Mean of {info['n_tests']} Tests) "
                  f"rows={info['n_rows']:2} vals={info['n_vals']:3}  reconcile {ok}/{tot}{drp}")
            print(f"              locations: {', '.join(info['locs'])}")
        traits_here = sorted({t for (_, t, inf) in report if isinstance(inf, dict) and inf.get('n_vals')})
        print(f"  PER-LOCATION traits recovered: {', '.join(traits_here) if traits_here else 'NONE'}")

    df = pd.DataFrame(all_recs, columns=["Year", "TestType", "TestMG", "Test", "Strain", "City",
                                         "State", "Phenotype", "Value_num", "Units", "Source"])
    df.to_csv(OUT, index=False)
    print("\n" + "=" * 78)
    print(f"wrote {OUT.name}  ({len(df):,} rows)\n")
    print("rows per (Year,Test,Phenotype):")
    print(df.groupby(["Year", "Test", "Phenotype"]).size().to_string())
    return df


if __name__ == "__main__":
    main()
