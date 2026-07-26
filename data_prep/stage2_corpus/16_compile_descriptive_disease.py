"""
16_compile_descriptive_disease.py
=================================
Compile ONE separate unified long-format file holding the NUST *descriptive*
traits across the whole record (1970-2025) — the data the main yield-centric
extraction either captured only post-1989 (Chlorosis, Shattering, Emergence,
GreenStem, DescriptiveCode) or dropped entirely pre-1989 (the "Descriptive and
Other Data" block). This is the descriptive analogue of the corpus, kept
SEPARATE from `phenotypesTable1.csv` / the wide corpus.

Scope = PASS 1 (descriptive scores). The disease-reaction matrix
(BB/BP/BS/frogeye/PM/BSR/Phytophthora/SMV…) is a later PASS 2.

THREE source feeds, each robust on its own (no dependence on the corpus carrying
the recovered/modern descriptive rows, so a future corpus rebuild can't change
this file):
  • HISTORICAL 1970-1988  — Chlorosis, from the PRODUCTION Red-PDF parser
    `99_extract_chlorosis_1970_1988.py --dump-csv` (validated; supersedes the
    provisional `15_extract_descriptive_pre1989.py` / `descriptive_pre1989_long.csv`).
  • DIGITAL 1989-2020     — Chlorosis, Shattering, GreenStem, Emergence, Hypocotyl,
    DescriptiveCode, pulled from `nust_1941_2025_combined.csv` (Master/Queryportal).
    Capped at ≤2020 so it never overlaps the XLSX feed and is insensitive to the
    2020/2022-2025 IDC appends.
  • MODERN XLSX 2022-2025 — the full "DESCRIPTIVE AND DISEASE DATA" block of the
    `<year> UTPT Report - Conventional` XLSX: Chlorosis (IDC), Shattering, GreenStem
    (numeric) + LeafShape (categorical). Generalized from `98_extract_idc_2022_2025.py`.
    (2021 has no raw report; 1941-1969 the block does not exist; 1990 PDF-pilot skipped it.)

Output (analysis/data/_shared/):
  nust_descriptive_disease_1941_2025.csv  — long, one row per
      (Year, Test, City, Strain, Trait) descriptive measurement.

Run:  uv run python data_prep/stage2_corpus/16_compile_descriptive_disease.py
"""
import re
import subprocess
import sys
from pathlib import Path

import openpyxl
import pandas as pd

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO = Path("C:/Users/vramasub/Desktop/UMN_GIT/NUST_Data_Prep")
NUST_DATA = Path("C:/Users/vramasub/Desktop/UMN_Projects/NUST_Projects/NUST_Data")
SHARED = REPO / "analysis" / "data" / "_shared"
CORPUS = SHARED / "nust_1941_2025_combined.csv"
SCRIPT_99 = REPO / "data_prep" / "stage2_corpus" / "99_extract_chlorosis_1970_1988.py"
DESCBLOCK_HIST = SHARED / "descriptive_block_pre1989_long.csv"  # script 19: code+hypocotyl+shattering
OUT = SHARED / "nust_descriptive_disease_1941_2025.csv"

# Descriptive-score phenotypes as they appear in the digital corpus → canonical trait.
DESC_TRAIT_MAP = {
    "Chlorosis": "Chlorosis", "IDC": "Chlorosis",
    "Shattering": "Shattering",
    "Shattering10Oct": "Shattering", "Shattering30Sept": "Shattering",
    "Shattering3Oct": "Shattering", "Shattering9Oct": "Shattering",
    "Hypocotyl": "Hypocotyl", "Emergence": "Emergence",
    "DescriptiveCode": "DescriptiveCode", "GreenStem": "GreenStem",
}
DESCRIPTIVE_NUMERIC = {"Chlorosis", "Shattering", "GreenStem", "Emergence"}

OUT_COLS = ["Year", "Era", "Source", "Test", "TestType", "TestMG", "Variant",
            "City", "State", "Strain", "Phenotype", "Trait", "Value", "Value_num", "Units"]

# ---------------------------------------------------------------------------
# Modern XLSX descriptive-block parser (generalized from script 98's IDC parser)
# ---------------------------------------------------------------------------
MODERN_FILES = {
    2022: NUST_DATA / "2022" / "2022_NUST_Processing"
          / "2022 UTPT Report - Conventional - 01252023 Final (no formulas).xlsx",
    2023: NUST_DATA / "2023" / "2023 UTPT Report - Conventional - 20240206 Final no formulas.xlsx",
    2024: NUST_DATA / "2024" / "2024 UTPT Report - Conventional - 20250212 Final no formulas.xlsx",
    2025: NUST_DATA / "2025" / "2025 UTPT Report - Conventional - 20260120 Final no Formulas.xlsx",
}
MODERN_CSV = {
    2022: NUST_DATA / "2022" / "2022_NUST_Data" / "phenotypesTable1.csv",
    2023: NUST_DATA / "2023" / "2023_NUST_Processing" / "phenotypesTable1.csv",
    2024: NUST_DATA / "2024" / "2024_NUST_Processing" / "Files4Upload" / "phenotypesTable1.csv",
    2025: NUST_DATA / "2025" / "2025_NUST_Processing" / "Files4Upload" / "phenotypesTable1.csv",
}
MODERN_HEADER_TO_TRAIT = {"IDC": "Chlorosis", "Shattering": "Shattering",
                          "Green Stem": "GreenStem", "Leaf Shape": "LeafShape"}
SCORE_LO, SCORE_HI = 0.5, 9.0
SUMMARY_ROWS = {"mean", "c.v. (%)", "l.s.d. (5%)", "average", "c.v.", "lsd"}

LOCN = {
    "DanversMN": ("Danvers", "MN"), "Danvers": ("Danvers", "MN"),
    "Crookston": ("Crookston", "MN"), "Climax": ("Climax", "MN"),
    "Lamberton": ("Lamberton", "MN"), "Humboldt": ("Humboldt", "IA"),
    "Ames": ("Ames", "IA"), "Rosemount": ("Rosemount", "MN"), "Waseca": ("Waseca", "MN"),
    "Manhattan": ("Manhattan", "KS"),
    "St Mathieude Beloeil": ("St Mathieu de Beloeil", "QC"),
    "Elora": ("Elora", "ON"), "EloraONT": ("Elora", "ON"),
    "Woodstock": ("Woodstock", "ON"), "WoodstockONT": ("Woodstock", "ON"),
}
_PROV = {"ONT": "ON", "ON": "ON", "QC": "QC", "MN": "MN", "IA": "IA",
         "KS": "KS", "IL": "IL", "WI": "WI", "IN": "IN", "MO": "MO"}
STATE_NORM = {"ONT": "ON", "QUE": "QC", "MAN": "MB", "MINNESOTA": "MN", "IOWA": "IA"}


def frag_join(fr):
    s = "".join(f.replace("-", "") for f in fr)
    return re.sub(r"Crooks?t+on", "Crookston", s)


def strain_key(s):
    s = re.sub(r"\s*\([^)]*\)", "", str(s)).replace("*", "")
    return re.sub(r"\s+", "", s).upper()


def resolve_locn(locstr):
    if locstr in LOCN:
        return LOCN[locstr]
    for tok, st in _PROV.items():
        if locstr.endswith(tok) and len(locstr) > len(tok):
            base = locstr[: -len(tok)]
            return LOCN.get(base, (base, st))
    return (locstr, "")


def test_to_mg(test):
    t = str(test).replace("-", "")
    m = re.match(r"^(?:UT|PT)(00|0|I{1,3}V?|IV|V)", t)
    return m.group(1) if m else ""


def norm_test(sheet):
    m = re.match(r"^(UT|PT)(00|0|I{1,3}V?|IV|V)([AB]?)$", sheet)
    return f"{m.group(1)}-{m.group(2)}{m.group(3)}" if m else sheet


def find_descriptive_strain_col(ws):
    best = None
    for row in ws.iter_rows(min_row=1, max_row=60):
        for c in row:
            if isinstance(c.value, str) and c.value.strip() == "Strain" and c.column >= 12:
                best = (c.column, c.row)
    return best if best else (None, None)


def trait_locations(ws, span, strain_hdr_row, header_row):
    locs = {}
    for cc in range(span[0], span[1] + 1):
        frags = []
        for rr in range(header_row + 1, strain_hdr_row + 1):
            v = ws.cell(rr, cc).value
            if isinstance(v, str) and v.strip() and v.strip().lower() not in ("score", "strain"):
                frags.append(v.strip())
        locs[cc] = frag_join(frags)
    return locs


def extract_modern_sheet(ws, sheet, key2spelling):
    strain_col, strain_hdr_row = find_descriptive_strain_col(ws)
    if strain_col is None:
        return
    test = norm_test(sheet)
    trait_headers = []
    for row in ws.iter_rows(min_row=1, max_row=strain_hdr_row):
        for c in row:
            if isinstance(c.value, str) and c.value.strip() in MODERN_HEADER_TO_TRAIT and c.column >= 13:
                r, col = c.row, c.column
                span = (col, col)
                for mr in ws.merged_cells.ranges:
                    if mr.min_row <= r <= mr.max_row and mr.min_col <= col <= mr.max_col:
                        span = (mr.min_col, mr.max_col)
                trait_headers.append((MODERN_HEADER_TO_TRAIT[c.value.strip()], r, span))
    if not trait_headers:
        return
    trait_cols = []
    for trait, hrow, span in trait_headers:
        for cc, locstr in trait_locations(ws, span, strain_hdr_row, hrow).items():
            city, state = resolve_locn(locstr)
            kind = "numeric" if trait in DESCRIPTIVE_NUMERIC else "categorical"
            trait_cols.append((trait, cc, city, state, kind))
    rr, blanks = strain_hdr_row + 1, 0
    while rr < ws.max_row:
        sv = ws.cell(rr, strain_col).value
        s = str(sv).strip() if sv is not None else ""
        if s == "":
            blanks += 1
            if blanks >= 3:
                break
            rr += 1
            continue
        blanks = 0
        if s.lower() in SUMMARY_ROWS:
            break
        strain = key2spelling.get(strain_key(s),
                                  re.sub(r"\s*\([^)]*\)", "", s).replace("*", "").strip())
        for trait, cc, city, state, kind in trait_cols:
            v = ws.cell(rr, cc).value
            if kind == "numeric":
                if isinstance(v, (int, float)) and SCORE_LO <= float(v) <= SCORE_HI:
                    yield (test, strain, city, state, trait, round(float(v), 4), float(v))
            elif isinstance(v, str) and v.strip():
                yield (test, strain, city, state, trait, v.strip(), None)
        rr += 1


def from_modern_xlsx():
    rows = []
    for year, xlsx in MODERN_FILES.items():
        if not xlsx.exists():
            print(f"  {year}: MISSING {xlsx.name}")
            continue
        src = pd.read_csv(MODERN_CSV[year], low_memory=False)
        key2spelling = {strain_key(s): str(s).strip() for s in src["Strain"].astype(str).unique()}
        wb = openpyxl.load_workbook(xlsx, data_only=True)
        n0 = len(rows)
        for sheet in wb.sheetnames:
            for test, strain, city, state, trait, value, vnum in extract_modern_sheet(
                    wb[sheet], sheet, key2spelling):
                rows.append({
                    "Year": year, "Era": "modern_xlsx", "Source": "UTPT_XLSX",
                    "Test": test, "TestType": test[:2], "TestMG": test_to_mg(test),
                    "Variant": "Conventional", "City": city, "State": state, "Strain": strain,
                    "Phenotype": trait, "Trait": trait, "Value": str(value), "Value_num": vnum,
                    "Units": "score" if vnum is not None else "code",
                })
        byt = pd.Series([r["Trait"] for r in rows[n0:]]).value_counts().to_dict()
        print(f"  {year}: {len(rows)-n0} rows  {byt}")
    return pd.DataFrame(rows, columns=OUT_COLS)


# ---------------------------------------------------------------------------
# Historical 1970-1988 — production script 99 dump
# ---------------------------------------------------------------------------
def from_historical_99():
    dump = SHARED / "_tmp_chl_99.csv"
    res = subprocess.run(["uv", "run", "python", str(SCRIPT_99), "--dump-csv", str(dump)],
                         cwd=str(REPO), capture_output=True, text=True)
    if not dump.exists():
        sys.exit(f"  ERROR: 99 dump failed:\n{res.stdout[-800:]}\n{res.stderr[-800:]}")
    d = pd.read_csv(dump)
    dump.unlink()
    out = pd.DataFrame({
        "Year": d["Year"], "Era": "historical_pdf", "Source": "RedPDF_script99",
        "Test": d["Test"], "TestType": d["Test"].str[:2], "TestMG": d["Test"].map(test_to_mg),
        "Variant": "Conventional", "City": d["City"], "State": d["State"], "Strain": d["Strain"],
        "Phenotype": "Chlorosis", "Trait": "Chlorosis",
        "Value": d["Value"].round(4).astype(str), "Value_num": d["Value"], "Units": "score",
    })
    print(f"  historical (99) rows: {len(out):,}  ({out.Year.min()}-{out.Year.max()})")
    return out[OUT_COLS]


# ---------------------------------------------------------------------------
# Historical 1970-1988 — descriptive block from script 19:
#   DescriptiveCode + Hypocotyl (categorical codes) + Shattering (numeric score)
# ---------------------------------------------------------------------------
def from_historical_descblock():
    if not DESCBLOCK_HIST.exists():
        print(f"  (skip: {DESCBLOCK_HIST.name} not present — run script 19 first)")
        return pd.DataFrame(columns=OUT_COLS)
    d = pd.read_csv(DESCBLOCK_HIST, low_memory=False)
    out = pd.DataFrame({
        "Year": d["Year"], "Era": "historical_pdf", "Source": "RedPDF_script19",
        "Test": d["Test"], "TestType": d["Test"].str[:2], "TestMG": d["Test"].map(test_to_mg),
        "Variant": "Conventional", "City": "", "State": "", "Strain": d["Strain"],
        "Phenotype": d["Trait"], "Trait": d["Trait"],
        "Value": d["Value"].astype(str), "Value_num": d["Value_num"], "Units": d["Units"],
    })
    byt = out.groupby("Trait").size().to_dict()
    print(f"  historical descriptive-block (19) rows: {len(out):,}  ({out.Year.min()}-{out.Year.max()})  {byt}")
    return out[OUT_COLS]


# ---------------------------------------------------------------------------
# Digital 1989-2020 — from the corpus (capped so it can't overlap the XLSX feed)
# ---------------------------------------------------------------------------
def from_corpus():
    df = pd.read_csv(CORPUS, low_memory=False)
    df = df[(df["Year"] >= 1989) & (df["Year"] <= 2020)].copy()
    sub = df[df["Phenotype"].isin(DESC_TRAIT_MAP)].copy()
    sub["Trait"] = sub["Phenotype"].map(DESC_TRAIT_MAP)
    sub["Era"] = "modern_corpus"
    sub["Variant"] = sub.get("Variant", "")
    sub["Value"] = sub["Value_num"].astype("string")  # text Value lost upstream; mirror Value_num
    if "Units" not in sub.columns:
        sub["Units"] = ""
    for c in OUT_COLS:
        if c not in sub.columns:
            sub[c] = pd.NA
    print(f"  digital (corpus) rows: {len(sub):,}  (1989-2020)")
    return sub[OUT_COLS]


def main():
    if not CORPUS.exists():
        sys.exit(f"ERROR: missing {CORPUS}")
    print("=" * 70)
    print("Compiling unified descriptive file (PASS 1 = descriptive scores)")
    print("=" * 70)
    print("\n[1/4] Historical 1970-1988 Chlorosis (Red PDFs, script 99) ...")
    hist = from_historical_99()
    print("\n[2/4] Historical 1970-1988 descriptive block: Code+Hypocotyl+Shattering (script 19) ...")
    hist_code = from_historical_descblock()
    print("\n[3/4] Digital 1989-2020 (corpus) ...")
    modern = from_corpus()
    print("\n[4/4] Modern 2022-2025 (UTPT XLSX full descriptive block) ...")
    xlsx = from_modern_xlsx()

    out = pd.concat([hist, hist_code, modern, xlsx], ignore_index=True)
    out["State"] = (out["State"].fillna("").astype(str).str.strip().str.upper().replace(STATE_NORM))
    out = out.sort_values(["Trait", "Year", "Test", "City", "Strain"]).reset_index(drop=True)
    out.to_csv(OUT, index=False)

    print(f"\nWrote {OUT.name}: {len(out):,} rows × {len(out.columns)} cols")
    print("\n  Trait × year coverage:")
    print(out.groupby("Trait")["Year"].agg(["min", "max", "nunique", "count"]).to_string())
    chl = sorted(out.loc[out.Trait == "Chlorosis", "Year"].unique())
    missing = sorted(set(range(1970, 2026)) - set(chl))
    print(f"\n  Chlorosis span {chl[0]}-{chl[-1]} ({len(chl)} yrs); missing in 1970-2025: {missing}")
    print(f"  Era breakdown: {out.groupby('Era').size().to_dict()}")


if __name__ == "__main__":
    main()
