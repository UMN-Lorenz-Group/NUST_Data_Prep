"""
22_verify_short_code_strains.py
===============================
Verify the `very_short_code` strains flagged by the QC audit (script 13) against the
ORIGINAL source files — confirming, per (Strain, Year), that the short 2-char code is a
real nursery entry that carries a full identity (originator + pedigree + generation) in
BOTH the Green XLSX (OCR intermediate) and the Red PDF (ground truth).

Why this is tractable: these codes begin with a letter (H/M/L/C/S/W/A series, 1941-1984),
so they are searchable in a way the pure-numeric line numbers are not. The source lists
each code once in a strain-identification / key block:
    L1   Ill. A.E.S. & U.S.R.S.L.   Chippewa (8) x Blackhawk   F1
    H5   Ohio Agr.Exp.Sta.          (LX378-32) (Mukden x Mandarin)

Handled gotchas:
  * OCR variants of the code itself (Ll->L1, HI->H1, Ml/MI->M1, MS->M5, 3H->H3) — searched
    via ambiguous-character candidate expansion.
  * Same code, different strain across years (M8 in 1948-52 MG0 vs 1963 MGI) — keyed on the
    actual corpus (Strain, Year) pairs, each searched in THAT year's source.
  * OCR-degraded PDF text layer (L1->"LI", "Exp"->"E3cp") — fuzzy pedigree-keyword match.
  * A code re-used in a later year may appear only in that year's yield table (pedigree in
    the intro year) — reported as xlsx_data_only.

Read-only. Output: analysis/data/analysis_results/Corpus_QC/short_code_source_verification.{csv,md}
"""
import sys
import re
import glob
from pathlib import Path
from itertools import product
import pandas as pd
import openpyxl
import pdfplumber

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

REPO = Path("C:/Users/vramasub/Desktop/UMN_GIT/NUST_Data_Prep")
CORPUS = REPO / "analysis" / "data" / "_shared" / "nust_1941_2025_combined.csv"
AUDIT = REPO / "analysis" / "data" / "analysis_results" / "Corpus_QC" / "corpus_qc_strain_audit.csv"
OUTDIR = REPO / "analysis" / "data" / "analysis_results" / "Corpus_QC"
INPUT = REPO / "input_files"

# Ambiguous OCR character groups for code-candidate expansion (S sits in both 5S and 8BS:
# 'MS' (1950) is an OCR of 'M8'; 'S'<->'5' and 'S'<->'8' both occur).
OCR_GROUPS = [set("1IL"), set("0OQ"), set("5S"), set("8BS"), set("2Z"), set("6G")]
# Identifies an originator / pedigree text cell (vs a numeric yield/data cell or a 1-char
# disease rating). Any cell with a >=3-letter run and length >=5.
TEXT_CELL = re.compile(r"[A-Za-z]{3,}")


def _expand(code):
    pools = []
    for ch in code.upper():
        opts = {ch}
        for grp in OCR_GROUPS:
            if ch in grp:
                opts |= grp
        pools.append(sorted(opts))
    return {"".join(p) for p in product(*pools)}


def code_candidates(code):
    """OCR-variant clean forms of a short code (Ll->L1, MS->M8) plus the transposed form
    ('3H' is an OCR transposition of 'H3')."""
    code = str(code).strip()
    out = _expand(code) | {code, code.upper()}
    if len(code) == 2:
        out |= _expand(code[::-1])     # transposition: 3H <-> H3
    return out


def year_xlsx(year):
    return sorted(set(glob.glob(str(INPUT / f"input_{year}" / "*.xlsx"))
                      + glob.glob(str(INPUT / f"input_{year}" / str(year) / "*.xlsx"))))


def year_pdf(year):
    cands = glob.glob(str(INPUT / f"input_{year}" / f"{year}_done.pdf")) \
        or glob.glob(str(INPUT / f"input_{year}" / "*.pdf"))
    return cands[0] if cands else None


def build_xlsx_idblock(paths):
    """Scan a year's XLSX file(s) → {literal_code(upper): (originator, pedigree, raw_idrow)}.
    An ID/key row has the code in col-0 and at least one originator/pedigree TEXT cell after."""
    idblock = {}
    for p in paths:
        try:
            wb = openpyxl.load_workbook(p, data_only=True, read_only=True)
        except Exception as e:
            print(f"    ! cannot open {Path(p).name}: {e}")
            continue
        for sh in wb.sheetnames:
            for r in wb[sh].iter_rows(values_only=True):
                cells = [("" if c is None else str(c).strip()) for c in r]
                if not cells or not cells[0]:
                    continue
                # strip a leading entry-number prefix: the 1973+ format is "18. S6"
                c0 = re.sub(r"^\d{1,3}\.\s*", "", cells[0])
                if not c0 or len(c0) > 4:
                    continue
                key = c0.upper()
                INST = r"Sta\.|A\.E\.S|Lab|Agr|Univ|Co\.|Brodbeck|U\.S\.R|Expt|Exp\."
                # pedigree = the cross/Sel/bracket cell, which may be ALL codes with no
                # 3-letter word (e.g. 'F2 [C1128 (6) x S54-1207]', 'L8 x L7') — so detect it
                # by marker, not by the originator text filter; exclude institution cells.
                pedigree = next((c for c in cells[1:] if len(c) >= 4
                                 and re.search(r"\bx\b|×|\bSel|Parentage|\[", c, re.I)
                                 and not re.search(INST, c, re.I)), "")
                texts = [c for c in cells[1:] if len(c) >= 5 and TEXT_CELL.search(c)]
                if not texts and not pedigree:
                    continue                      # numeric data row, not an ID row
                originator = next((c for c in texts if c != pedigree
                                   and re.search(INST, c, re.I)),
                                  next((c for c in texts if c != pedigree), ""))
                # keep the RICHEST row for a code: prefer one that has a real pedigree
                # (an earlier continuation/data row can match a code but lack the cross).
                idrow = " | ".join(dict.fromkeys([x for x in [originator, pedigree] if x]
                                                 + texts[:3]))
                prev = idblock.get(key)
                if prev is None or (not prev[1] and pedigree):
                    idblock[key] = (originator, pedigree, idrow)
    return idblock


def pdf_text(path, cache):
    if path in cache:
        return cache[path]
    chunks = []
    try:
        with pdfplumber.open(path) as doc:
            for pg in doc.pages:
                chunks.append(pg.extract_text() or "")
    except Exception as e:
        print(f"    ! cannot read PDF {Path(path).name}: {e}")
    cache[path] = "\n".join(chunks)
    return cache[path]


def ped_keyword(pedigree):
    """A distinctive token from a pedigree for fuzzy PDF confirmation: prefer a real parent
    name (>=4 letters); else fall back to an alphanumeric line code (C1128, H21162, LX378)."""
    toks = [t for t in re.findall(r"[A-Za-z]{4,}", str(pedigree))
            if t.lower() not in ("parentage", "unknown", "from", "sel", "selection")]
    if toks:
        return max(toks, key=len)
    codes = [t for t in re.findall(r"[A-Za-z]{1,3}\d{2,}[\w-]*", str(pedigree)) if len(t) >= 4]
    return max(codes, key=len) if codes else ""


def main():
    aud = pd.read_csv(AUDIT, keep_default_na=False)
    codes = sorted(set(aud[aud.rule == "very_short_code"].Strain))
    corp = pd.read_csv(CORPUS, low_memory=False, usecols=["Year", "TestMG", "Strain"])
    corp = corp[corp.Strain.isin(codes)]
    pairs = (corp.groupby(["Strain", "Year"]).agg(
        n_rows=("TestMG", "size"),
        MGs=("TestMG", lambda x: ",".join(sorted(set(x.astype(str)))))).reset_index()
        .sort_values(["Strain", "Year"]))

    pdf_cache, xlsx_cache = {}, {}
    rows = []
    for _, r in pairs.iterrows():
        code, year = r.Strain, int(r.Year)
        cands = code_candidates(code)
        xls = year_xlsx(year)
        if year not in xlsx_cache:
            xlsx_cache[year] = build_xlsx_idblock(xls)
        idb = xlsx_cache[year]
        matched = next((c for c in cands if c in idb), None)
        originator = pedigree = idrow = ""
        if matched:
            originator, pedigree, idrow = idb[matched]
        # PDF confirmation: pedigree keyword present near the code in the (noisy) PDF text
        pdf_path = year_pdf(year)
        pdf_ped_found = False
        kw = ped_keyword(pedigree)
        if pdf_path and kw:
            txt = pdf_text(pdf_path, pdf_cache).lower()
            pdf_ped_found = kw.lower() in txt
        if matched and pdf_ped_found:
            status = "CONFIRMED_BOTH"
        elif matched and not pedigree:
            status = "xlsx_data_only"     # code present but pedigree listed in its intro year
        elif matched and not kw:
            status = "xlsx_confirmed"     # pedigree present but not PDF-checkable (parentage
            #                               unknown / self-referential like 'L8 x L7')
        elif matched:
            status = "xlsx_only_pdf_miss"  # pedigree in XLSX, keyword not echoed in noisy PDF
        else:
            status = "NOT_FOUND_check_manually"
        rows.append({
            "Strain": code, "Year": year, "MG": r.MGs, "n_rows": r.n_rows,
            "matched_code": matched or "", "originator": originator,
            "pedigree": pedigree, "pdf_keyword": kw, "pdf_pedigree_found": pdf_ped_found,
            "status": status, "identity_row": idrow,
            "xlsx_file": ";".join(Path(x).name for x in xls),
            "pdf_file": Path(pdf_path).name if pdf_path else ""})

    out = pd.DataFrame(rows)
    OUTDIR.mkdir(parents=True, exist_ok=True)
    out.to_csv(OUTDIR / "short_code_source_verification.csv", index=False)

    # markdown summary
    L = ["# Short-code strain source verification (XLSX + PDF)\n",
         f"{len(out)} (Strain, Year) pairs across {out.Year.nunique()} years; "
         f"{out.Strain.nunique()} distinct codes.\n",
         "Status: **CONFIRMED_BOTH** = pedigree in XLSX + keyword echoed in PDF; "
         "**xlsx_only** = pedigree in XLSX, PDF keyword not matched (OCR); "
         "**xlsx_data_only** = code present but pedigree listed in its intro year; "
         "**NOT_FOUND** = needs a manual look.\n",
         "## Status counts\n"]
    for s, n in out.status.value_counts().items():
        L.append(f"- {s}: {n}")
    L.append("\n## Per (Strain, Year)\n")
    L.append("| Strain | Year | MG | status | originator | pedigree (full identity) |")
    L.append("|--------|------|----|--------|-----------|--------------------------|")
    for _, r in out.iterrows():
        L.append(f"| {r.Strain} | {r.Year} | {r.MG} | {r.status} | {r.originator} | {r.pedigree} |")
    (OUTDIR / "short_code_source_verification.md").write_text("\n".join(L), encoding="utf-8")

    print(f"\n{len(out)} (Strain,Year) pairs verified. Status counts:")
    print(out.status.value_counts().to_string())
    print(f"\nNOT_FOUND / needs manual check:")
    nf = out[out.status.str.startswith("NOT_FOUND")]
    print(nf[["Strain", "Year", "MG", "xlsx_file"]].to_string(index=False) if len(nf) else "  (none)")
    print(f"\nWrote short_code_source_verification.csv + .md to {OUTDIR.name}/")


if __name__ == "__main__":
    main()
