"""
audit_test_map_green.py
=======================
TEST-MAP ACCURACY AUDIT. Verifies, per year and per F4U test, that each test's DATA is under the
correct test LABEL — a dimension the value-accuracy pipeline (08/09) is blind to (it keys on Test as a
coordinate, never verifies it). Uses the GREEN (Sojabone OCR XLSX) as the independent, correctly-
structured reference: the Green stores per-test blocks in PDF page order, so enumerating them and
re-doing the label assignment CORRECTLY (right section count) exposes the positional-`Group_N` scramble
in `combine_nust_outputs.TEST_MAPS`.

Method per year:
  1. Enumerate Green YIELD sections (tp6, fuzzy) in page order -> roster + per-strain yield-vector +
     geometry (nloc/nstrain). Merge wide-UT continuation sub-tables.
  2. Assign each Green section its TRUE identity from TEST_MAPS[year]'s intended SEQUENCE (position i ->
     i-th intended code), cross-checked against geometry (UT wide / PT narrow). Count-mismatch => complex.
  3. Match each F4U test's data -> a Green section by roster-overlap + order-independent yield
     value-fingerprint (decisive, label- & location-dialect-agnostic).
  4. Classify each F4U test: CLEAN / MISLABELED->true / MERGED / DROPPED / UNVERIFIABLE; per-year year_class.

Outputs (analysis/data/analysis_results/Corpus_QC/):
  nust_test_map_1941_1988.csv   (section_order, f4u_code, green_true_code, status, evidence)
  nust_test_map_current.csv     (TEST_MAPS dumped to data)
  audit_test_map_green.csv      (row per (Year,F4U_Test): true_test, status, evidence)
  audit_test_map_summary.md

Usage:  uv run python audit_test_map_green.py [--year YYYY] [--no-write]
Reuses 109_extract_green_section: norm, parse_loc, cellnum.  Read-only wrt the corpus/F4U.
"""
import sys
import re
import glob
import shutil
import argparse
import importlib
from pathlib import Path
from collections import defaultdict
import openpyxl
import pandas as pd

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
G9 = importlib.import_module("109_extract_green_section")
norm, cellnum = G9.norm, G9.cellnum

# 109's parse_loc falls through to "treat as a city" for anything unrecognized, so non-location header
# columns (notably 'Rank' in a mis-tagged YIELD-RANK table) get counted as locations. Reject those here.
# NOTE `\d+ tests?` : the 1970s-era Green writes the MEAN column header as a bare "9 Tests" / "27 Tests"
# (the 'Mean of' sits on the marker row, or OCR dropped it). Without this it is parsed as a CITY and
# inflates nloc. See _MEANCELL_RE for the matching has_mean/mean_n half of the same bug.
_NOTLOC_RE = re.compile(r"^\s*(rank|no\.?|number|mean|strain|strn|tests?|avg|average|index|%|"
                        r"\d+\s*tests?)\s*$", re.I)
# A mean column is 'Mean', 'Mean of N Tests', or (1970s era) a bare 'N Tests'.
_MEANCELL_RE = re.compile(r"mean|^\s*\d+\s*tests?\s*$", re.I)

def parse_loc(c):
    """109.parse_loc + reject non-location header cells (Rank/No./Index/...)."""
    if c is None or _NOTLOC_RE.match(str(c)):
        return None
    return G9.parse_loc(c)

REPO = Path("C:/Users/vramasub/Desktop/UMN_GIT/NUST_Data_Prep")
NUST = Path("C:/Users/vramasub/Desktop/UMN_Projects/NUST_Projects/NUST_Data")
R_GREEN = Path("R:/cfans_agro_lore0149_lorenzlabresearch/NUST_Data/NUST_Data/"
               "NUST_Historical_Data_1941_1988/Green-20260427T181938Z-3-001/Green")
CACHE = Path("C:/Users/vramasub/AppData/Local/Temp/nust_green_cache")
OUT = REPO / "analysis" / "data" / "analysis_results" / "Corpus_QC"

# ---------------------------------------------------------------- Green file location
def _is_year_green(name, year):
    n = name.lower()
    return "sojabone" in n and str(year) in n and n.endswith(".xlsx") and "injury" not in n

def green_files(year):
    """Resolve the year's Sojabone XLSX. Order: (1) repo input_files, (2) local temp cache,
    (3) R: (copied into the cache). Checking the cache BEFORE R: keeps the audit working when the
    R: network drive is offline. Skips marker-less supplemental files (e.g. Soybean Injury)."""
    files = sorted(glob.glob(str(REPO / "input_files" / f"input_{year}" / "*Sojabone*.xlsx")))
    if not files:                                              # (2) cache (R:-independent)
        files = sorted(str(p) for p in CACHE.glob("*.xlsx") if _is_year_green(p.name, year))
    if not files:                                              # (3) R: -> cache
        rdir = R_GREEN / str(year)
        try:
            cand = sorted(glob.glob(str(rdir / "*Sojabone*.xlsx"))) if rdir.exists() else []
        except OSError:
            cand = []
        CACHE.mkdir(parents=True, exist_ok=True)
        out = []
        for f in cand:
            dst = CACHE / Path(f).name
            if not dst.exists():
                try:
                    shutil.copy2(f, dst)
                except OSError:
                    continue
            out.append(str(dst))
        files = out
    return [f for f in files if _is_year_green(Path(f).name, year)]

# ---------------------------------------------------------------- fuzzy tp-marker
def ntp(tag):
    """Canonicalize a column-A marker cell to a tp-token, tolerating OCR damage:
    'tp 7'->tp7, 'tp6 & tp7'->tp6, 'tp7 (20+21_OR.png)'->tp7, 'tp??'->None. Returns token or None."""
    if tag is None:
        return None
    s = str(tag).strip().lower()
    s = re.sub(r"\(.*", "", s).strip()                 # drop PNG-filename tails
    s = s.split("&")[0].strip()                        # combined 'tp6 & tp7' -> first
    s = re.sub(r"\s+", "", s)                           # 'tp 7' -> 'tp7'
    m = re.match(r"^(tp\d+[ab]?)$", s)
    return m.group(1) if m else None

# ---------------------------------------------------------------- Green yield-section enumerator
_FOOTER_RE = re.compile(r"(mean|no\.?\s|c\.?v|l\.?s\.?d|coef|range|row spacing|bu\.|grand|date planted|days to|average)", re.I)
# MULTI-YEAR marker: needs an explicit N-year token ('2-year', 'Three-year', '1983-1984, 2-year mean').
# Bare 'Summary' must NOT match -- it labels the current test's own regional summary.
_NYEAR_RE = re.compile(r"\b(\d{1,2}|one|two|three|four|five|six|seven|eight|nine|ten)[\s-]*year", re.I)
# Some years mis-tag NON-yield trait tables as `tp6` (e.g. 1950 'tp6 | Maturity' should be tp8). A tp6
# whose marker names another trait (and not Yield) is not a yield table -> not a test section.
_NONYIELD_RE = re.compile(r"(maturity|lodging|height|seed\s*qual|seed\s*size|seed\s*weight|protein|"
                          r"\boil\b|chlorosis|shatter|descriptive|iodine|parentage|disease)", re.I)
# A 'Yield Rank' table is a RANK table, not a yield table: same Strain/Mean/location header as the
# yield table it follows, but the cells are ranks (1..n). It contains the word 'yield', so the
# _NONYIELD_RE rule below would clear it, and it has real location columns so the nloc==0 guard
# misses it too -> it would be counted as an extra section (1960: 13 vs 12). Rank always wins.
_RANK_RE = re.compile(r"\brank\b", re.I)

def _read_tp6(grid, i):
    """Read one tp6 yield table starting at marker row i. Returns dict or None.
    {has_mean, is_summary, locs, rows{strainNorm:[cells]}, end}.
    NOTE: `is_summary` means MULTI-YEAR (2-/3-/5-year mean) -> skip. It must require an explicit
    N-YEAR token: a bare 'Summary' is the CURRENT test's own regional-summary label (e.g. 1950's
    'Summary agronomic, chem. data') and marks a REAL single-year section."""
    marker = " ".join(str(c) for c in grid[i] if c)
    is_summary = bool(_NYEAR_RE.search(marker))
    # a tp6 whose label names another trait (and not Yield) is a mis-tagged non-yield table
    is_yield = not (_NONYIELD_RE.search(marker) and not re.search(r"yield", marker, re.I))
    h = i + 1
    # header ('Strain'/'Strn.') can sit 1-6 rows below the marker (a title/label row may intervene)
    while h < len(grid) and h < i + 7 and not re.match(r"str", str(grid[h][0] or "").strip(), re.I):
        if ntp(grid[h][0]) is not None:            # ran into the next marker -> no header here
            return None
        if _NYEAR_RE.search(str(grid[h][0] or "")):
            is_summary = True
        h += 1
    if h >= len(grid) or not re.match(r"str", str(grid[h][0] or "").strip(), re.I):
        return None
    header = grid[h]
    # has_mean marks a SECTION START (continuation loc-groups carry no mean col), so missing a mean
    # spelling silently demotes a real test to a continuation -> UNDERCOUNT. The 1970s Green spells it
    # as a bare 'N Tests' (1972: 'Strain | 9 Tests | Ont. Ridgetown ...').
    has_mean = any(_MEANCELL_RE.search(str(c)) for c in header[1:3] if c is not None)
    # "Mean of N Tests" (or bare "N Tests") -> N. For a current-year test N ~= its location count; a
    # multi-year summary of the SAME test has a strictly larger N (locs x years). Guards the collapse.
    mn = None
    mean_ci = None
    for ci, c in enumerate(header[1:3], start=1):
        if c is not None and _MEANCELL_RE.search(str(c)) and mean_ci is None:
            mean_ci = ci
        m = re.search(r"(?:mean\D{0,6})?(\d+)\s*tests?\b", str(c), re.I)
        if m and mn is None:
            mn = int(m.group(1))
    locs = [parse_loc(c) for c in header[1:]]
    rows = {}
    r = h + 1
    while r < len(grid):
        a = grid[r][0]
        atag = str(a).strip() if a is not None else ""
        if ntp(a) is not None or re.match(r"str", atag, re.I):
            break
        if not atag:
            r += 1
            # Two consecutive blanks mean the table ENDED -- but only once we are inside the data. A
            # blank gap between the header and the first data row is just spacing (1981 row 475 has
            # two), and breaking there yielded a 0-strain section that `yield_sections` then dropped
            # => UNDERCOUNT. The loop is still bounded: it breaks on the next tp marker / Strain header.
            if rows and r < len(grid) and grid[r][0] is None:
                break
            continue
        if _FOOTER_RE.match(atag):
            r += 1; continue
        sk = norm(atag)
        if sk:
            rows[sk] = list(grid[r][1:])
        r += 1
    # RANK-TABLE rejection, decided on the VALUES not the marker text. A rank table repeats the yield
    # table's header but holds ranks (1960 row 60 `tp6 | Yield Rank`, Flambeau 1|1|1). The marker alone
    # is NOT enough: 1959 row 231 `tp6 & tp7 | Yield rank` is a COMBINED yield+rank table whose cells
    # are real yields (Acme 25.8|29.9|31.5) with a Rank column -- rejecting it on the word 'rank' loses
    # a real test. Ranks are small integers bounded by the strain count; yields are not.
    if is_yield and _RANK_RE.search(marker) and rows:
        vals = []
        for _sk, cells in rows.items():
            for ci, L in enumerate(locs):
                if L is None:
                    continue
                v = cellnum(cells[ci]) if ci < len(cells) else None
                if v is not None:
                    vals.append(float(v))
        if vals:
            n_int = sum(1 for v in vals if v.is_integer())
            if n_int / len(vals) >= 0.9 and max(vals) <= len(rows) + 2:
                is_yield = False
    # per-strain Mean-column vector: constant across a test's pages, and distinct between tests (each
    # test is its own trial). Lets _is_same_test() spot a continuation page that REPEATS the Strain +
    # Mean + leading location columns -- see yield_sections.
    means = {}
    if mean_ci is not None:
        for sk, cells in rows.items():
            v = cellnum(cells[mean_ci - 1]) if mean_ci - 1 < len(cells) else None
            if v is not None:
                means[sk] = round(v, 1)
    return {"has_mean": has_mean, "is_summary": is_summary, "is_yield": is_yield, "locs": locs,
            "rows": rows, "mean_n": mn, "means": means, "end": r}


def _is_same_test(a_means, b_means):
    """True when two tp6 tables carry the SAME per-strain Mean vector => same test, so the second is a
    continuation page that re-printed the Mean column (1970 does this) rather than a new section.
    Different tests are different trials, so their per-strain means differ even for shared checks."""
    shared = set(a_means) & set(b_means)
    if len(shared) < 3:
        return False
    same = sum(1 for k in shared if abs(a_means[k] - b_means[k]) < 0.051)
    return same / len(shared) >= 0.8

def yield_sections(year):
    """Enumerate real single-year YIELD tests, page order. A test STARTS at a non-summary tp6 that has
    a 'Mean of N Tests' column; following non-summary tp6's WITHOUT a Mean col are continuation loc-groups
    (merge them). Multi-year 'N-year summary' tp6 tables are skipped. Returns list of
    {order, roster, nstrain, nloc, ylong[(strainNorm,value)], file, row}."""
    secs = []
    order = 0
    for f in green_files(year):
        try:
            wb = openpyxl.load_workbook(f, read_only=True, data_only=True)
        except Exception:
            continue
        if "Sheet1" not in wb.sheetnames:
            wb.close(); continue
        grid = [list(r) for r in wb["Sheet1"].iter_rows(values_only=True)]
        wb.close()
        i = 0
        while i < len(grid):
            tag = ntp(grid[i][0])
            if tag not in ("tp6", "tp5", "tp4"):
                i += 1; continue
            if tag != "tp6":
                # A test's per-location yield table is occasionally mis-tagged tp5/tp4 (1976 row 3588:
                # `tp5 | 1976 YIELD (bu./A)` with a Strain|Mean|locations header) -> scanning tp6 only
                # loses that whole test. tp4/tp5 are normally the REGIONAL SUMMARY / N-year-mean tables,
                # so admit one only when its marker names YIELD for the CURRENT year (no N-year token):
                # that rejects `tp5 1975-1976, 2-YEAR MEAN YIELD`, `tp5 1976, EAST COAST` and bare `tp4`,
                # while the has_mean / is_yield / nloc>0 filters below still apply.
                mk = " ".join(str(c) for c in grid[i] if c)
                if not re.search(r"yield", mk, re.I) or _NYEAR_RE.search(mk):
                    i += 1; continue
            tbl = _read_tp6(grid, i)
            if tbl is None:
                i += 1; continue
            if tbl["is_summary"] or not tbl["has_mean"] or not tbl["is_yield"]:
                i = tbl["end"]; continue        # skip summaries, orphan continuations, mis-tagged non-yield
            # section start -> accumulate this table + following no-Mean, non-summary continuations
            roster, ylong, nloc = set(), [], 0
            def add(t):
                nonlocal nloc
                nloc += sum(1 for L in t["locs"] if L)
                for sk, cells in t["rows"].items():
                    roster.add(sk)
                    for cidx, L in enumerate(t["locs"]):
                        if L is None:
                            continue
                        v = cellnum(cells[cidx]) if cidx < len(cells) else None
                        if v is not None:
                            ylong.append((sk, round(v, 1)))
            add(tbl)
            sec_means = dict(tbl["means"])
            j = tbl["end"]
            while j < len(grid):
                # skip blanks to the next marker
                k = j
                while k < len(grid) and ntp(grid[k][0]) is None and str(grid[k][0] or "").strip() == "":
                    k += 1
                if k >= len(grid) or ntp(grid[k][0]) != "tp6":
                    break
                cont = _read_tp6(grid, k)
                if cont is None or cont["is_summary"] or not cont["is_yield"]:
                    break                                  # next real section / summary -> stop
                # A continuation normally has NO mean col. But some years (1970) re-print Strain+Mean+
                # the leading loc columns on the continuation page; that repeated mean col would
                # otherwise start a phantom section. Same per-strain mean vector => same test.
                if cont["has_mean"] and not _is_same_test(sec_means, cont["means"]):
                    break                                  # a genuinely different test -> stop
                add(cont); j = cont["end"]
            # nloc==0 => no real location columns (e.g. a mis-tagged YIELD-RANK table whose only
            # non-Mean column is 'Rank') -> not a per-location yield test section.
            if roster and nloc > 0:
                secs.append({"roster": roster, "nstrain": len(roster), "mean_n": tbl.get("mean_n"),
                             "means": sec_means,
                             "nloc": nloc, "ylong": ylong, "file": Path(f).name, "row": i})
            i = j
    return _collapse_summaries(secs)


def _collapse_summaries(secs):
    """Drop multi-year-summary yield tables that carry NO parseable 'N-year' label (1950s era: the
    tp?? marker text is OCR-lost). Structure per test = current-year yield FIRST, then its 2-/N-year
    summaries, whose rosters are SUBSETS of the current-year roster (only multi-year-tested strains).
    Adjacent DIFFERENT tests share only a few check varieties, so high containment discriminates.
    Rule: a section whose roster is >=80% contained in the previous KEPT section's roster AND is not
    larger AND has a strictly larger 'Mean of N Tests' (locs x years > locs) => a summary of that same
    test -> drop. The mean_n guard prevents collapsing a genuinely SMALL adjacent test whose roster
    happens to be contained in the previous test's. Keeps the first (current-year) of each group."""
    kept = []
    for s in secs:
        if kept:
            prev = kept[-1]
            # SAME-TEST REPEAT: a page of the previous test that re-printed Strain+Mean+leading locs
            # (1970 does this, with other tables in between so the continuation loop never sees it).
            # Identical per-strain mean vector => same trial => not a new section.
            if _is_same_test(prev.get("means", {}), s.get("means", {})):
                continue
            inter = len(s["roster"] & prev["roster"])
            denom = max(1, min(len(s["roster"]), len(prev["roster"])))
            # mean_n VETO: only when both are known and the candidate's N is NOT larger (=> it's a
            # genuine small adjacent test, not a summary). If either N is unparseable, fall back to
            # containment alone (which is the reliable signal for the OCR-lost 1950s markers).
            sn, pn = s.get("mean_n"), prev.get("mean_n")
            not_summary = (sn is not None and pn is not None and sn <= pn)
            if inter / denom >= 0.8 and len(s["roster"]) <= len(prev["roster"]) and not not_summary:
                continue
        kept.append(s)
    for n, s in enumerate(kept, 1):
        s["order"] = n
    return kept

# ---------------------------------------------------------------- TEST_MAPS intended sequence
def test_map_sequence(year):
    """Ordered list of intended test codes (by Group_N) for the year. Primary source =
    combine_nust_outputs.TEST_MAPS; fallback = the per-year `input_files/input_<year>/<year>_done_test_map.json`
    override (1987/1988/1989 have no TEST_MAPS entry and were combined via --test_map JSON)."""
    tm = _load_test_maps()
    d = tm.get(str(year))
    if d:
        def gnum(k):
            m = re.search(r"(\d+)", k); return int(m.group(1)) if m else 999
        return [d[k] for k in sorted(d, key=gnum)]
    j = REPO / "input_files" / f"input_{year}" / f"{year}_done_test_map.json"
    if j.exists():
        try:
            import json
            g = json.loads(j.read_text(encoding="utf-8")).get("groups", [])
            seq = [x["test_code"] for x in sorted(g, key=lambda x: x.get("group_number", 999))
                   if x.get("test_code")]
            return seq or None
        except Exception:
            return None
    return None

_TM_CACHE = None
def _load_test_maps():
    global _TM_CACHE
    if _TM_CACHE is not None:
        return _TM_CACHE
    src = (REPO / "data_prep" / "stage1_processing" / "combine_nust_outputs.py").read_text(encoding="utf-8")
    m = re.search(r"TEST_MAPS.*?=\s*\{", src)
    # brace-match from the first { after TEST_MAPS
    start = src.index("{", m.start())
    depth, i = 0, start
    while i < len(src):
        if src[i] == "{": depth += 1
        elif src[i] == "}":
            depth -= 1
            if depth == 0:
                break
        i += 1
    body = src[start:i + 1]
    _TM_CACHE = eval(body, {"__builtins__": {}}, {})   # pure literal dict
    return _TM_CACHE

# ---------------------------------------------------------------- F4U side
def f4u_path(year, use_bak=False):
    for base in (NUST / "NUST_Historical_Data_1941_1988", NUST):
        p = base / f"{year}_Processing" / "Files4Upload" / "phenotypesTable1.csv"
        if use_bak:
            for suf in (".csv.bak_pre_1984pt_swap", ".csv.orig_prelabel"):
                b = p.with_suffix(suf)
                if b.exists():
                    return b
        if p.exists():
            return p
    return None

def load_f4u_tests(year, use_bak=False):
    """{test_code: {roster:set(norm), yvals:Counter(rounded yields), nstrain, nyield}} from the F4U."""
    from collections import Counter
    p = f4u_path(year, use_bak)
    if p is None:
        return {}, None
    d = pd.read_csv(p, low_memory=False)
    d.columns = [c.strip() for c in d.columns]
    tcol = next(c for c in d.columns if c.lower() == "test")
    scol = next(c for c in d.columns if c.lower() == "strain")
    pcol = next(c for c in d.columns if c.lower() == "phenotype")
    vcol = next(c for c in d.columns if c.lower() == "value")
    out = {}
    for t, sub in d.groupby(tcol):
        y = sub[sub[pcol] == "YieldBuA"]
        roster = {norm(s) for s in sub[scol].dropna() if str(s).strip().lower() != "mean"}
        roster.discard("")
        yv = Counter(round(float(v), 1) for v in pd.to_numeric(y[vcol], errors="coerce").dropna())
        out[str(t)] = {"roster": roster, "yvals": yv,
                       "nstrain": len({norm(s) for s in y[scol].dropna()} - {""}),
                       "nyield": int(sum(yv.values()))}
    return out, p

def _mset_overlap(a, b):
    """multiset intersection size (Counter)."""
    return sum((a & b).values())

# ---------------------------------------------------------------- audit one year
def audit_year(year, use_bak=False):
    secs = yield_sections(year)
    seq = test_map_sequence(year)
    f4u, fpath = load_f4u_tests(year, use_bak)
    from collections import Counter
    # precompute each Green section's yield multiset + intended identity + geometry proxy
    for s in secs:
        s["yvals"] = Counter(v for _, v in s["ylong"])
        s["intended"] = seq[s["order"] - 1] if seq and s["order"] - 1 <= len(seq) - 1 else None
        s["vps"] = (len(s["ylong"]) / s["nstrain"]) if s["nstrain"] else 0   # values-per-strain (UT~15-25 / PT~7-10)
        s["geom"] = "UT" if s["vps"] >= 13 else "PT"
    rows = []
    for t, info in sorted(f4u.items()):
        if not info["yvals"]:
            rows.append(dict(Year=year, f4u_test=t, true_test="", status="NO_YIELD",
                             roster_ov=0, roster_n=len(info["roster"]), yield_match=0,
                             n_shared=0, nyield=info["nyield"], confidence="", matched_order=""))
            continue
        # best Green section by yield value-multiset overlap
        scored = sorted(((_mset_overlap(info["yvals"], s["yvals"]), s) for s in secs),
                        key=lambda x: x[0], reverse=True)
        best_ov, best = scored[0] if scored else (0, None)
        denom = min(info["nyield"], sum(best["yvals"].values())) if best else 1
        yfrac = best_ov / denom if denom else 0
        rost_ov = len(info["roster"] & best["roster"]) if best else 0
        true = best["intended"] if best else None
        conf = "decisive" if yfrac >= 0.85 else ("moderate" if yfrac >= 0.5 else "weak")
        # roster corroboration + merge detection (a real match needs yield AND roster agreement)
        rost_min = max(4, int(0.30 * len(info["roster"])))
        strong = [s for s in secs if len(info["roster"] & s["roster"]) >= max(8, int(0.4 * s["nstrain"]))]
        if best is None or yfrac < 0.4:
            status = "UNVERIFIABLE"
        elif len(strong) >= 2:
            status = "MERGED"                                  # F4U roster spans >1 Green section
        elif rost_ov < rost_min:
            status = "UNVERIFIABLE"                            # yield matched but roster doesn't corroborate
        elif true is None:
            status = "NO_INTENDED"
        elif _canon(true) == _canon(t):
            status = "CLEAN"
        else:
            status = f"MISLABELED->{true}"
        rows.append(dict(Year=year, f4u_test=t, true_test=true or "", status=status,
                         roster_ov=rost_ov, roster_n=len(info["roster"]),
                         yield_match=round(yfrac, 3), n_shared=best_ov, nyield=info["nyield"],
                         confidence=conf, matched_order=best["order"] if best else ""))
    # section-level table (order, intended, which f4u claimed it)
    claim = defaultdict(list)
    for r in rows:
        if r["matched_order"] != "":
            claim[r["matched_order"]].append(r["f4u_test"])
    sec_rows = []
    for s in secs:
        claimants = claim.get(s["order"], [])
        sec_rows.append(dict(Year=year, section_order=s["order"], green_true_code=s["intended"] or "",
                             geom=s["geom"], nstrain=s["nstrain"], nyield=len(s["ylong"]),
                             f4u_claimants=";".join(claimants),
                             status=("DROPPED" if not claimants else
                                     ("OK" if any(_canon(c) == _canon(s["intended"] or "") for c in claimants) else "MISLABEL")),
                             green_file=s["file"], green_row=s["row"]))
    # per-year class
    st = [r["status"] for r in rows]
    dropped = any(sr["status"] == "DROPPED" for sr in sec_rows)
    multiclaim = any(len(v) > 1 for v in claim.values())
    has_bad = any(s in ("UNVERIFIABLE", "MERGED", "NO_YIELD", "NO_INTENDED") for s in st)
    mislabels = [r for r in rows if r["status"].startswith("MISLABELED")]
    labels = {_canon(r["f4u_test"]) for r in rows}
    trues = [_canon(r["true_test"]) for r in mislabels]
    is_perm = len(trues) == len(set(trues)) and set(trues) <= labels   # bijection into the F4U label set
    if not seq:
        yclass = "no_testmap"                    # 1987/88/89 -> can't name sections
    elif len(secs) != len(seq):
        yclass = "enum_uncertain"                # Green section count != intended map -> parsing not
                                                 # reliable enough to name; do NOT assert a scramble
    elif has_bad or dropped or multiclaim:
        yclass = "complex"                       # merges / drops / unverifiable -> manual
    elif not mislabels:
        yclass = "clean"
    elif is_perm:
        yclass = "clean_shift"                   # pure decisive permutation -> auto-relabel candidate
    else:
        yclass = "complex"
    return rows, sec_rows, yclass, fpath, len(secs), (len(seq) if seq else 0)

_CANON_RE = re.compile(r"[^A-Z0-9]")
def _canon(code):
    c = _CANON_RE.sub("", str(code).upper()).replace("UPT", "PT")
    c = c.replace("UTO", "UT0")            # UT-O (letter) == UT-0 (zero)
    return c

# ---------------------------------------------------------------- CLI / sweep
GREEN_YEARS = list(range(1941, 1989))
SKIP = {1975: "PDF-direct (no Green)", 1990: "PDF-direct", 1989: "Master (no Green)"}

if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--year", type=int)
    ap.add_argument("--bak", action="store_true", help="audit the pre-swap/pre-relabel F4U backup")
    ap.add_argument("--no-write", action="store_true")
    args = ap.parse_args()
    years = [args.year] if args.year else GREEN_YEARS
    all_rows, all_secs, ysum = [], [], []
    for y in years:
        if y in SKIP and not args.year:
            ysum.append(dict(Year=y, year_class="skipped", note=SKIP[y], n_green=0, n_testmap=0)); continue
        try:
            rows, sec_rows, yclass, fpath, nsec, nseq = audit_year(y, use_bak=args.bak)
        except Exception as e:
            ysum.append(dict(Year=y, year_class="ERROR", note=str(e)[:80], n_green=0, n_testmap=0)); continue
        all_rows += rows; all_secs += sec_rows
        reliable = yclass in ("clean", "clean_shift", "complex")
        mis = [f"{r['f4u_test']}->{r['true_test']}" for r in rows if r["status"].startswith("MISLABELED")]
        note = (";".join(mis) if mis else "-") if reliable else "(parse unreliable)"
        ysum.append(dict(Year=y, year_class=yclass, note=note, n_green=nsec, n_testmap=nseq))
        if args.year:
            print(f"\n=== {y}  year_class={yclass}  green_sections={nsec} testmap={nseq}  F4U={fpath}")
            print(pd.DataFrame(rows)[["f4u_test", "true_test", "status", "yield_match", "n_shared",
                                      "roster_ov", "roster_n", "confidence"]].to_string(index=False))
            print("\n-- Green sections --")
            print(pd.DataFrame(sec_rows)[["section_order", "green_true_code", "geom", "nstrain",
                                          "f4u_claimants", "status"]].to_string(index=False))
    print("\n=== YEAR SUMMARY ===")
    print(pd.DataFrame(ysum).to_string(index=False))
    if not args.no_write and not args.year:
        OUT.mkdir(parents=True, exist_ok=True)
        pd.DataFrame(all_rows).to_csv(OUT / "audit_test_map_green.csv", index=False)
        pd.DataFrame(all_secs).to_csv(OUT / "nust_test_map_1941_1988.csv", index=False)
        # dump TEST_MAPS as-coded
        tm = _load_test_maps()
        cur = [dict(Year=y, Group=k, code=v) for y in sorted(tm) for k, v in tm[y].items()]
        pd.DataFrame(cur).to_csv(OUT / "nust_test_map_current.csv", index=False)
        # NB: DataFrame.to_markdown() needs `tabulate`, which is NOT a declared dependency -- it raised
        # ImportError here AFTER the CSVs were written, so every writing run ended in a traceback and
        # left summary.md truncated. Render the table directly instead of adding a dep for one call.
        ysum_df = pd.DataFrame(ysum)
        with open(OUT / "audit_test_map_summary.md", "w", encoding="utf-8") as fh:
            fh.write("# Test-Map Accuracy Audit (Green vs F4U)\n\n")
            fh.write("| " + " | ".join(ysum_df.columns) + " |\n")
            fh.write("|" + "|".join("---" for _ in ysum_df.columns) + "|\n")
            for _, r in ysum_df.iterrows():
                fh.write("| " + " | ".join(str(v) for v in r) + " |\n")
        print(f"\nwrote audit_test_map_green.csv / nust_test_map_1941_1988.csv / _current.csv / summary.md -> {OUT}")
