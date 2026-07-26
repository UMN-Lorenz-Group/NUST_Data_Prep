"""1983 tp2 SCRAMBLE fix (B3 part 2) — the only year needing tp2 REMOVAL as well as inserts.

The 12 yield tables (tp6) enumerate the 12 TRUE tests cleanly, but the tp2 markers are broken: 11 tp2 =
12 real - 3 missing + 2 spurious. Fix, fully determined by aligning tp2 to the intact tp6 sections:

  INSERT tp2 (blank -> tp2), at each missing test's parentage boundary:
    file1 @961  PT-I     (parentage before its mislabeled tp3a@962; merged into UT-I)
    file2 @555  PT-IIIA  (parentage header@556; merged into UT-III)
    file2 @1679 UT-IV    (parentage before its mislabeled tp3a@1680; merged into PT-IIIB)

  BLANK tp2 (tp2 -> ""), the 2 spurious markers (NOT relabel -- relabeling @1341 to tp3a would COLLIDE
  with the genuine tp3a@1378 chlorosis + tp3b@1414 disease tables):
    file1 @1341  UT-II "Descriptive Code" sub-table mislabeled tp2 (real parentage is tp2@1305)
    file2 @2206  duplicate of PT-IV parentage already opened at tp2@2182

Net 11 - 2 + 3 = 12 tp2 = 12 tests. Blanking (not deleting rows) keeps the descriptive-code / duplicate-
parentage rows inside their group's sub-chunk A, read by the API at re-extraction; only the phantom group
boundary is removed. Also registers 1983 in reference/tp2_repair_spec.csv so derive_test_map trusts the
post-fix raw tp2 count (the stale emitted count would otherwise still report a deficit).

`--apply` backs up both Green files to `.bak_pre_1983fix` and writes. Idempotent-ish (re-run safe: inserts
skip non-empty cells, blanks skip already-blank).
"""
import sys
import shutil
import importlib
from pathlib import Path
import openpyxl
import pandas as pd

sys.stdout.reconfigure(encoding="utf-8")
HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE.parents[1] / "data_prep" / "stage0_extraction"))
A = importlib.import_module("audit_test_map_green")
SPEC = HERE.parents[1] / "reference" / "tp2_repair_spec.csv"

F1 = "Sojabone-1983 (1-110 OR).xlsx"
F2 = "Sojabone-1983 (111-215 OR).xlsx"
INSERTS = [  # (green_file, row, before_code, tp6_row, boundary)
    (F1, 961, "PT-I", 1054, "parentage-header(modern)"),
    (F2, 555, "PT-IIIA", 730, "parentage-header(modern)"),
    (F2, 1679, "UT-IV", 1786, "parentage-header(modern)"),
]
BLANKS = [(F1, 1341, "UT-II descriptive-code sub-table"), (F2, 2206, "duplicate PT-IV parentage")]


def path_for(fname):
    return next(p for p in A.green_files(1983) if Path(p).name == fname)


def main():
    apply = "--apply" in sys.argv
    files = {F1: path_for(F1), F2: path_for(F2)}

    for fname, path in files.items():
        wb = openpyxl.load_workbook(path)
        ws = wb["Sheet1"]
        for gf, row, code, *_ in INSERTS:
            if gf != fname:
                continue
            cell = ws.cell(row=row, column=1)
            state = "empty" if cell.value in (None, "") or not str(cell.value).strip() else f"NON-EMPTY {cell.value!r}"
            print(f"  INSERT {fname} @{row} (before {code}): cell is {state}")
        for gf, row, what in BLANKS:
            if gf != fname:
                continue
            cell = ws.cell(row=row, column=1)
            print(f"  BLANK  {fname} @{row} ({what}): cell={cell.value!r}")
        if apply:
            bak = Path(path).with_suffix(".xlsx.bak_pre_1983fix")
            if not bak.exists():
                shutil.copy2(path, bak)
                print(f"    backed up -> {bak.name}")
            for gf, row, *_ in INSERTS:
                if gf == fname:
                    c = ws.cell(row=row, column=1)
                    if c.value in (None, "") or not str(c.value).strip():
                        c.value = "tp2"
            for gf, row, _ in BLANKS:
                if gf == fname:
                    ws.cell(row=row, column=1).value = None
            wb.save(path)
            print(f"    saved {fname}")
        wb.close()

    if apply:
        spec = pd.read_csv(SPEC)
        spec = spec[spec.Year != 1983]
        add = pd.DataFrame([dict(Year=1983, green_file=gf, insert_row=row, before_code=code,
                                 before_tp6_row=tp6, boundary=b, shared_group="scramble")
                            for gf, row, code, tp6, b in INSERTS])
        pd.concat([spec, add], ignore_index=True).to_csv(SPEC, index=False)
        print(f"  registered 1983 in {SPEC.name} (3 inserts)")
    else:
        print("\n(dry run; --apply to write)")


if __name__ == "__main__":
    main()
