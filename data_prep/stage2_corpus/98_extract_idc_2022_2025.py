"""
98_extract_idc_2022_2025.py
===========================
Re-extract the iron-deficiency-chlorosis (IDC) "IDC Score" column from the raw
2022-2025 UTPT Conventional report XLSX files and append it to each year's
phenotypesTable1.csv as Phenotype="Chlorosis" (the historical 1989-2019 label).

WHY
---
The corpus carried "Chlorosis" for 1989-2019 and "IDC" for 2020 only; 2021-2025
had neither. Diagnosis (logged in analysis_notes.md): the IDC Score lives in the
"DESCRIPTIVE AND DISEASE DATA" regional-summary block of the raw report (not in
the per-location yield tables that fed phenotypesTable1.csv), so it was never
extracted. This script recovers it. IDC and the pre-2020 "Chlorosis" share the
same 1-5 visual IDC scale, so they merge cleanly (the wide-builder,
11_build_wide_1941_2025.py, additionally aliases the leftover 2020 "IDC" label to
"Chlorosis").

WHAT IT FINDS (Conventional reports only; verified 2026-06-23)
-------------------------------------------------------------
  2022: UT00/UT0/UTI            -> Danvers MN              (62 vals)
  2023: UT00/UT0/UTI            -> Crookston MN            (59 vals)
        (UTII-IV have the header but no data -> skipped)
  2024: UT00/UT0/UTI            -> Climax + Crookston MN   (106 vals)
  2025: UT00/UT0/UTI/PTI        -> Crookston + Climax MN   (144 vals)
  2021: NO raw report exists (only derived tables, no IDC) -> remains unavailable.

A genuine block is identified by an 'IDC' header cell whose merged span carries a
'Score' sub-header; the 2025 UTIII "IDC" string (a Unique-Traits tag in the
parentage block) is correctly ignored because it has no 'Score' sub-header.

OUTPUT
------
Appends long rows (Phenotype="Chlorosis", Units="score") to:
  2023/2024/2025 phenotypesTable1.csv (already long)
For 2022 (wide phenotypesTable1.csv) it adds a "Chlorosis" column with Danvers
rows; the assemble step melts wide->long so these become Chlorosis long rows.

Idempotent: a one-time *.bak backup is made, prior Chlorosis rows the script
added are dropped before re-appending, so re-running is safe.

Run from repo root with `uv run python data_prep/stage2_corpus/98_extract_idc_2022_2025.py`.
"""
import re
import shutil
import sys
from pathlib import Path

import openpyxl
import pandas as pd

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

NUST_DATA = Path("C:/Users/vramasub/Desktop/UMN_Projects/NUST_Projects/NUST_Data")

# Raw conventional report XLSX + the per-year phenotypesTable1.csv to augment.
YEARS = {
    2022: {
        "xlsx": NUST_DATA / "2022" / "2022_NUST_Processing"
                / "2022 UTPT Report - Conventional - 01252023 Final (no formulas).xlsx",
        "csv":  NUST_DATA / "2022" / "2022_NUST_Data" / "phenotypesTable1.csv",
        "format": "wide",
    },
    2023: {
        "xlsx": NUST_DATA / "2023" / "2023 UTPT Report - Conventional - 20240206 Final no formulas.xlsx",
        "csv":  NUST_DATA / "2023" / "2023_NUST_Processing" / "phenotypesTable1.csv",
        "format": "long",
    },
    2024: {
        "xlsx": NUST_DATA / "2024" / "2024 UTPT Report - Conventional - 20250212 Final no formulas.xlsx",
        "csv":  NUST_DATA / "2024" / "2024_NUST_Processing" / "Files4Upload" / "phenotypesTable1.csv",
        "format": "long",
    },
    2025: {
        "xlsx": NUST_DATA / "2025" / "2025 UTPT Report - Conventional - 20260120 Final no Formulas.xlsx",
        "csv":  NUST_DATA / "2025" / "2025_NUST_Processing" / "Files4Upload" / "phenotypesTable1.csv",
        "format": "long",
    },
}

# IDC nursery location string (fragment-joined header) -> (City, State).
# All IDC nurseries are MN. Guarded: an unknown location aborts so a future
# year's new nursery can't silently get the wrong state.
LOCN = {
    "DanversMN": ("Danvers", "MN"),
    "Crookston": ("Crookston", "MN"),
    "Climax":    ("Climax", "MN"),
    # Present as headers in 2023 southern sheets but carry no data; mapped for safety.
    "Lamberton": ("Lamberton", "MN"),
    "Humboldt":  ("Humboldt", "IA"),
}

PHENOTYPE = "Chlorosis"
UNITS = "score"
SUMMARY_ROWS = {"mean", "c.v. (%)", "l.s.d. (5%)", "average", "c.v.", "lsd"}


def frag_join(fragments):
    """Join stacked header fragments (e.g. ['Crook-','ston'] -> 'Crookston')."""
    s = "".join(f.replace("-", "") for f in fragments)
    return re.sub(r"Crooks?t+on", "Crookston", s)  # normalize OCR-ish 'Crookstton'


def strain_key(s):
    """Space/paren/star-insensitive key for matching to the source file's spelling."""
    s = re.sub(r"\s*\([^)]*\)", "", str(s)).replace("*", "")
    return re.sub(r"\s+", "", s).upper()


def find_idc_blocks(ws):
    """Yield (col_span, strain_col, strain_hdr_row, {col: location}) for each
    genuine IDC Score block on a sheet. A block is genuine iff its merged 'IDC'
    span carries a 'Score' sub-header (excludes parentage Unique-Traits tags)."""
    for row in ws.iter_rows(min_row=1, max_row=60):
        for cell in row:
            if not (isinstance(cell.value, str) and cell.value.strip().upper() == "IDC"):
                continue
            r, c = cell.row, cell.column
            span = (c, c)
            for mr in ws.merged_cells.ranges:
                if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                    span = (mr.min_col, mr.max_col)
            has_score = any(
                isinstance(ws.cell(rr, cc).value, str)
                and ws.cell(rr, cc).value.strip().lower() == "score"
                for rr in range(r + 1, r + 3)
                for cc in range(span[0], span[1] + 1)
            )
            if not has_score:
                continue
            # Strain label (block has its own Strain column to the left of IDC).
            strain_col = strain_hdr_row = None
            for rr in range(r, r + 9):
                for cc in range(1, span[0]):
                    if ws.cell(rr, cc).value == "Strain":
                        strain_col, strain_hdr_row = cc, rr
            if strain_col is None:
                continue
            # Per-column location: concatenate string fragments below 'Score'.
            frags = {cc: [] for cc in range(span[0], span[1] + 1)}
            for rr in range(r + 1, strain_hdr_row + 1):
                for cc in range(span[0], span[1] + 1):
                    v = ws.cell(rr, cc).value
                    if isinstance(v, str) and v.strip() and v.strip().lower() != "score":
                        frags[cc].append(v.strip())
            locs = {cc: frag_join(f) for cc, f in frags.items()}
            yield span, strain_col, strain_hdr_row, locs


def extract_year(year, info):
    """Return a DataFrame: Year, Test, RawStrain, City, State, Value."""
    wb = openpyxl.load_workbook(info["xlsx"], data_only=True)
    out = []
    for sheet in wb.sheetnames:
        ws = wb[sheet]
        for span, strain_col, strain_hdr_row, locs in find_idc_blocks(ws):
            rr = strain_hdr_row + 1
            blanks = 0
            while rr < ws.max_row:
                sv = ws.cell(rr, strain_col).value
                s = str(sv).strip() if sv is not None else ""
                if s == "":
                    blanks += 1
                    if blanks >= 3:
                        break
                    rr += 1
                    continue
                blanks = 0
                if s.lower() in SUMMARY_ROWS:
                    break
                for cc in range(span[0], span[1] + 1):
                    v = ws.cell(rr, cc).value
                    if isinstance(v, (int, float)):
                        loc = locs[cc]
                        if loc not in LOCN:
                            sys.exit(f"ERROR {year} {sheet}: unknown IDC location "
                                     f"{loc!r}; add it to LOCN with the correct state.")
                        city, state = LOCN[loc]
                        out.append((year, sheet, s, city, state, round(float(v), 4)))
                rr += 1
    return pd.DataFrame(out, columns=["Year", "Test", "RawStrain", "City", "State", "Value"])


def backup_once(path):
    bak = path.with_suffix(path.suffix + ".bak")
    if not bak.exists():
        shutil.copy2(path, bak)
        print(f"    backup -> {bak.name}")


def append_long(year, info, ext):
    """Append Chlorosis long rows to a long-format phenotypesTable1.csv,
    matching its exact column schema and mapping strains to its spelling."""
    path = info["csv"]
    src = pd.read_csv(path, low_memory=False)
    cols = list(src.columns)
    key2spelling = {strain_key(s): str(s).strip() for s in src["Strain"].astype(str).unique()}

    # Idempotency: drop any Chlorosis/IDC rows a prior run added.
    if "Phenotype" in cols:
        src = src[~src["Phenotype"].isin([PHENOTYPE, "IDC"])].copy()

    new = pd.DataFrame({
        "Year": year,
        "Test": ext["Test"],
        "City": ext["City"],
        "State": ext["State"],
        "Strain": ext["RawStrain"].map(lambda s: key2spelling[strain_key(s)]),
        "OriginalStrain": ext["RawStrain"].map(lambda s: key2spelling[strain_key(s)]),
        "Phenotype": PHENOTYPE,
        "Value": ext["Value"],
        "Units": UNITS,
    })
    # 2025 carries a combined Location (City_State) column alongside City.
    if "Location" in cols:
        new["Location"] = new["City"] + "_" + new["State"]
    new = new[[c for c in cols if c in new.columns]]
    # Guard: every source column must be produced (else schema drift).
    missing = [c for c in cols if c not in new.columns]
    if missing:
        sys.exit(f"ERROR {year}: cannot fill source columns {missing}")

    backup_once(path)
    pd.concat([src, new], ignore_index=True)[cols].to_csv(path, index=False)
    print(f"    appended {len(new)} Chlorosis rows -> {path.name} "
          f"(cities: {sorted(new['City'].unique())})")


def append_wide(year, info, ext):
    """Add a 'Chlorosis' column with IDC-nursery rows to a wide-format
    phenotypesTable1.csv. The assemble step melts wide->long, so each populated
    Chlorosis cell becomes one long Chlorosis row."""
    path = info["csv"]
    src = pd.read_csv(path, low_memory=False)
    cols = list(src.columns)
    key2spelling = {strain_key(s): str(s).strip() for s in src["Strain"].astype(str).unique()}

    # Idempotency: drop previously-added IDC rows (uniquely flagged by a non-null
    # Chlorosis cell), then drop the column so we rebuild cleanly.
    if "Chlorosis" in src.columns:
        src = src[src["Chlorosis"].isna()].drop(columns=["Chlorosis"]).copy()
        cols = [c for c in cols if c != "Chlorosis"]

    new = pd.DataFrame({
        "Year": year,
        "Test": ext["Test"],
        "Location": ext["City"],   # wide file's Location col holds the city
        "State": ext["State"],
        "Strain": ext["RawStrain"].map(lambda s: key2spelling[strain_key(s)]),
        "Chlorosis": ext["Value"],
    })
    backup_once(path)
    out = pd.concat([src, new], ignore_index=True)
    out_cols = cols + ["Chlorosis"]
    out[out_cols].to_csv(path, index=False)
    print(f"    appended {len(new)} Chlorosis rows (+Chlorosis col) -> {path.name} "
          f"(cities: {sorted(new['Location'].unique())})")


def main():
    print("=" * 70)
    print("IDC Score -> Chlorosis re-extraction (2022-2025 Conventional reports)")
    print("=" * 70)
    grand = 0
    for year, info in YEARS.items():
        print(f"\n[{year}] {info['xlsx'].name}")
        if not info["xlsx"].exists():
            sys.exit(f"  ERROR: missing XLSX {info['xlsx']}")
        if not info["csv"].exists():
            sys.exit(f"  ERROR: missing CSV {info['csv']}")
        ext = extract_year(year, info)
        if ext.empty:
            print("    no IDC values found")
            continue
        per_city = ext.groupby("City").size().to_dict()
        print(f"    extracted {len(ext)} IDC values  {per_city}")
        if info["format"] == "wide":
            append_wide(year, info, ext)
        else:
            append_long(year, info, ext)
        grand += len(ext)
    print(f"\nDone. {grand} Chlorosis rows written across {len(YEARS)} years.")
    print("Next: rebuild corpus (10_assemble_corpus.py, 11_build_wide_1941_2025.py),"
          " then 13_qc_strain_germplasm_audit.py + P2 traits heatmap.")


if __name__ == "__main__":
    main()
