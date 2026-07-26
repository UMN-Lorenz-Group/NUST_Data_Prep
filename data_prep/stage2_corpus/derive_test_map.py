"""Derive the TRUE per-year test map from EVIDENCE, not from `combine_nust_outputs.TEST_MAPS`.

WHY. `TEST_MAPS` is a hardcoded per-year POSITIONAL dict (`Group_N` -> code) where `Group_N` is merely
the ordinal of the Nth tp2 marker in an OCR'd Green sheet, and `apply_test_map()` is a bare
`dict.get(v, v)` with no validation: one missing tp2 permutes the whole year, one extra yields an
orphan `Group_11` that `10_assemble_corpus.parse_test_code` silently DROPS. Every post-hoc repair in
this tree (108_relabel_1959, 110_relabel_year, the 1984 rebuild, 115 recovery) exists to undo that.
And TEST_MAPS is itself wrong for several years, so auditing the Green against it is circular.

EVIDENCE MODEL (three independent sources; the PDF arbitrates):
  * Red PDF caption  -> the LABEL. The only source that states a section's code outright (pdf_captions).
  * Green XLSX       -> the STRUCTURE: section order, roster, per-strain yield values.
  * F4U              -> the artefact under audit.

Per year: align PDF sections <-> Green sections by roster overlap, giving each Green section its true
code; match each F4U test to a section by yield value-multiset fingerprint (decisive, label-agnostic);
finally map each section's tp6 row into its owning tp2 `Group_N` so the result can be emitted as a
`--test_map` JSON that `combine_nust_outputs` already accepts (no edit to TEST_MAPS needed).

Deliberately NOT a count gate. `n_green == n_testmap` is proven unsafe: 1970 and 1953 both matched by
a coincidence of cancelling errors. Confidence rests on caption + roster + yield agreement.
"""
import argparse
import importlib
import json
import re
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl
import pandas as pd

HERE = Path(__file__).resolve().parent
REPO = HERE.parents[1]
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(REPO / "data_prep" / "stage0_extraction"))

A = importlib.import_module("audit_test_map_green")
from pdf_captions import caption_sections, merge_noise_runs, sort_key, match_key, _FOOT  # noqa: E402
from extract_nust_xlsx import _normalize_tp  # noqa: E402

OUT_CSV = REPO / "reference" / "nust_test_map_verified.csv"
OUT_JSON_DIR = REPO / "reference" / "test_maps"


def tp2_groups(path):
    """1-based tp2 marker rows for one Green file -> the `Group_N` ordinals combine_nust_outputs uses.
    Mirrors extract_nust_xlsx.find_group_boundaries' core (enumerate(..., start=1))."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    ws = wb["Sheet1"] if "Sheet1" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = []
    for i, row in enumerate(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True), start=1):
        v = row[0]
        if not v:
            continue
        n = _normalize_tp(str(v).strip())
        if n == "tp2":
            rows.append(i)
    wb.close()
    return rows


def emitted_group_count(year):
    """The number of Group_N the EXTRACTOR actually produced -- the authoritative denominator and what
    `combine`'s Group_N lookup keys on -- NOT a re-count of today's input Green. Differs for pre-1960
    PREPROCESSED-Green years (1945: raw Green has 18 tp2, extractor emitted 5). Each per-prefix
    `*_phenotypes.csv` numbers its groups from 1 (per FILE), and `combine` renumbers file-2 by +max(file
    1); so the global count = SUM of per-file distinct Group_N counts (1984: file1 5 + file2 6 = 11),
    NOT the union. Returns int, or 0 if no outputs."""
    import glob
    total = 0
    for f in sorted(glob.glob(str(REPO / "output_files" / f"output_{year}" / "*_phenotypes.csv"))):
        gs = set()
        try:
            for line in open(f, encoding="utf-8", errors="replace"):
                for m in re.findall(r"Group_(\d+)", line):
                    gs.add(int(m))
        except Exception:
            pass
        total += len(gs)
    return total


def _mk(roster):
    """Comparison view of a roster: rank-prefix folded, footers dropped. The Green's roster keeps the
    printed rank ('1. Elgin' -> '1E1G1N' under 109.norm) while the PDF's does not, so raw sets can
    miss entirely -- 1984 UT-I overlapped on nothing but {'REPS'} before this."""
    return {match_key(x) for x in roster if str(x) not in _FOOT} - {""}


def overlap(a, b):
    a, b = _mk(a), _mk(b)
    if not a or not b:
        return 0.0
    return len(a & b) / max(1, min(len(a), len(b)))


def align(pdf_secs, green_secs):
    """Assign each Green section its PDF code by roster overlap, preferring the order-preserving
    (diagonal) assignment. Returns (mapping green_idx->pdf_idx, unaligned_green, unaligned_pdf)."""
    n, m = len(green_secs), len(pdf_secs)
    score = [[overlap(g["roster"], p["roster"]) for p in pdf_secs] for g in green_secs]
    pairs = sorted(((score[i][j], i, j) for i in range(n) for j in range(m)),
                   key=lambda x: -x[0])
    gmap, used_g, used_p = {}, set(), set()
    for s, i, j in pairs:
        if s <= 0.15 or i in used_g or j in used_p:
            continue
        gmap[i] = j
        used_g.add(i)
        used_p.add(j)
    return gmap, [i for i in range(n) if i not in used_g], [j for j in range(m) if j not in used_p]


def derive_no_green(year, pdf_secs, testmaps):
    """PDF-ONLY path (1987/1988): their Green lives on R: in a different format and is absent locally,
    so there is no tp2 count and no yield fingerprint. The PDF caption sequence is still the label
    truth, and the F4U roster still identifies which test each label holds -> we can VERIFY the
    existing `input_files/input_<year>/<year>_done_test_map.json` (which is what actually built the
    F4U) without the Green at all. Matching is by ROSTER overlap, not yields.
    """
    try:
        f4u, _ = A.load_f4u_tests(year)
    except Exception:
        f4u = {}
    rows = []
    for i, p in enumerate(pdf_secs, start=1):
        best, bo = None, 0.0
        for t, info in f4u.items():
            o = overlap(p["roster"], info["roster"])
            if o > bo:
                bo, best = o, t
        rows.append(dict(
            Year=year, section_order=i, group_n=i, true_code=p["code"],
            pdf_pages=f"{p['pages'][0]}-{p['pages'][-1]}", caption_pages=p["n_caption_pages"],
            roster_overlap=round(bo, 3), green_file="(none: PDF-only year)", green_row=-1,
            nstrain=len(p["roster"]), nloc=-1, geom="",
            f4u_claimant=best or "", yield_match=0.0,
            testmaps_code=testmaps[i - 1] if i - 1 < len(testmaps) else "",
        ))
    df = pd.DataFrame(rows)
    if not df.empty:
        df["agrees_with_testmaps"] = [
            bool(r.true_code) and bool(r.testmaps_code)
            and A._canon(r.true_code) == A._canon(r.testmaps_code) for r in df.itertuples()]
        df["status"] = [
            "CLEAN" if (r.f4u_claimant and A._canon(r.f4u_claimant) == A._canon(r.true_code))
            else ("DROPPED" if not r.f4u_claimant or r.roster_overlap < 0.3
                  else f"MISLABEL->{r.true_code}")
            for r in df.itertuples()]
        # Two sections claiming ONE F4U label: only the stronger claim is real; the weaker section is
        # a test the F4U does not hold (1988's true UT-00 was dropped, so its PDF section falls back
        # to UT-0 at 0.833 while the real UT-0 section claims it at 0.886).
        for lab, grp in df[df.f4u_claimant != ""].groupby("f4u_claimant"):
            if len(grp) > 1:
                keep = grp.roster_overlap.idxmax()
                for i in grp.index:
                    df.loc[i, "status"] = "CLEAN" if i == keep else "DROPPED(no F4U cell)"
    return df


def tp2_repaired_years():
    """Years whose Green tp2 markers were REPAIRED (inserts applied; see reference/tp2_repair_spec.csv).

    For these the RAW tp2 count is the post-repair AUTHORITY. The `emitted` output_files count is stale
    (it predates the not-yet-run re-extraction) and is smaller than raw, so the ordinary
    `n_groups = min(offset, emitted)` + `preprocessed = emitted < offset` logic would treat a fixed year
    as a preprocessed Green, report a phantom deficit, and refuse the JSON. Trusting raw for these years
    lets the map emit now (the JSON is the post-re-extraction target); it is NOT applied to genuinely
    preprocessed pre-1960 years, whose emitted count really is the authority.
    """
    f = REPO / "reference" / "tp2_repair_spec.csv"
    if not f.exists():
        return set()
    try:
        return set(pd.read_csv(f)["Year"].astype(int).unique())
    except Exception:
        return set()


def derive(year):
    pdf_secs = merge_noise_runs(caption_sections(year))
    green_secs = A.yield_sections(year)
    testmaps = A.test_map_sequence(year) or []
    if not green_secs:
        # PDF-only year (1987/88). n_groups is taken from the year's own _done_test_map.json length,
        # which IS the group count the extractor used to build the F4U.
        df = derive_no_green(year, pdf_secs, testmaps)
        return df, pdf_secs, [], [], [], testmaps, len(testmaps)
    for s in green_secs:
        from collections import Counter
        s["yvals"] = Counter(v for _, v in s["ylong"])
    try:
        f4u, _ = A.load_f4u_tests(year)
    except Exception:
        f4u = {}

    gmap, un_g, un_p = align(pdf_secs, green_secs)

    # --- Group_N per Green section (tp2 ordinal, continuing across the year's files) ---
    files = A.green_files(year)
    g_of_row, offset = {}, 0
    for f in files:
        t2 = tp2_groups(f)
        for k, r1 in enumerate(t2):
            end = t2[k + 1] - 1 if k + 1 < len(t2) else 10 ** 9
            g_of_row[(Path(f).name, r1, end)] = offset + k + 1
        offset += len(t2)

    def group_of(fname, row0):
        row1 = row0 + 1                      # green section row is a 0-based list index
        for (fn, a, b), gn in g_of_row.items():
            if fn == fname and a <= row1 <= b:
                return gn
        return None

    # Authoritative group count = what the EXTRACTOR emitted (output_files), falling back to the raw
    # tp2 count only if outputs are absent. Pre-1960 preprocessed-Green years diverge (raw tp2 is
    # per-table); the emitted count is what `combine` keys on, so it is the correct denominator.
    # Denominator = min(raw tp2, emitted). Preprocessed pre-1960 years consolidated tp2 per-test, so
    # emitted < raw (1945: emitted 5 < raw 18) -> take emitted. But the emitted reader OVER-counts some
    # years (1950/1972) because output_files carry sub-chunk/per-file Group_N labels that inflate the
    # sum above the real group count -> there emitted > raw and raw is right. min() is correct in both.
    emitted = emitted_group_count(year)
    repaired = year in tp2_repaired_years()   # raw tp2 is post-repair authoritative; ignore stale emitted
    n_groups = offset if repaired else (min(offset, emitted) if emitted else offset)

    rows = []
    for i, g in enumerate(green_secs):
        j = gmap.get(i)
        code = pdf_secs[j]["code"] if j is not None else None
        gn = group_of(g["file"], g["row"])
        # best F4U claimant by yield fingerprint
        best, bf = None, 0.0
        for t, info in f4u.items():
            ov = A._mset_overlap(info["yvals"], g["yvals"])
            den = min(info["nyield"], sum(g["yvals"].values())) or 1
            fr = ov / den
            if fr > bf:
                bf, best = fr, t
        rows.append(dict(
            Year=year, section_order=g["order"], group_n=gn, true_code=code,
            pdf_pages=f"{pdf_secs[j]['pages'][0]}-{pdf_secs[j]['pages'][-1]}" if j is not None else "",
            caption_pages=pdf_secs[j]["n_caption_pages"] if j is not None else 0,
            roster_overlap=round(overlap(g["roster"], pdf_secs[j]["roster"]), 3) if j is not None else 0,
            green_file=g["file"], green_row=g["row"], nstrain=g["nstrain"], nloc=g["nloc"],
            geom="UT" if g["nloc"] >= 12 else "PT",
            f4u_claimant=best or "", yield_match=round(bf, 3),
            # TEST_MAPS is keyed on Group_N, so index it by group_n -- NOT by section_order. Indexing
            # by section_order shifts the whole comparison whenever the Green enumerator misses a
            # section (1970 loses PT-00's yield table), reporting phantom disagreement for a year
            # whose map is in fact correct.
            testmaps_code=(testmaps[int(gn) - 1]
                           if gn and int(gn) - 1 < len(testmaps) else ""),
        ))
    df = pd.DataFrame(rows)
    # PREPROCESSED-GREEN years (raw tp2 count != emitted count, e.g. pre-1960 where tp2 is per-table):
    # the today-Green section -> Group_N mapping via raw tp2 rows is MEANINGLESS (the extractor saw a
    # consolidated Green we don't have). Null group_n so the JSON authors positionally from the PDF
    # (Group_i -> PDF code_i) and the Green-disagreement veto is skipped -- the emitted count is the
    # authority, and PDF order is the label truth.
    # consolidated Green -> fewer groups than raw tp2. NOT the tp2-repaired years: there emitted<offset
    # only because the re-extraction that would refresh emitted hasn't run yet, and raw is authoritative.
    preprocessed = (not repaired) and bool(emitted) and emitted < offset
    if preprocessed and not df.empty:
        df["group_n"] = pd.NA
        df["testmaps_code"] = ""
    if not df.empty:
        # compare CANONICALLY: TEST_MAPS writes the letter-O form 'UT-O' for some years while the PDF
        # prints the digit 'UT-0'; that is a dialect, not a disagreement.
        df["agrees_with_testmaps"] = [
            bool(r.true_code) and bool(r.testmaps_code)
            and A._canon(r.true_code) == A._canon(r.testmaps_code)
            for r in df.itertuples()]
        # status primarily reflects the DERIVED map's confidence, which is the ROSTER overlap to the
        # PDF (the label oracle) -- NOT the F4U-claimant yield fingerprint, which is noisy when adjacent
        # tests share check varieties (A4: 1954 sec2 truly UT-I at roster 0.857 but the weak fingerprint
        # (0.514) matched F4U 'UT-0', a false MISLABEL). A green section that aligned to NO PDF section
        # (roster 0) is a parse artefact, not a merge.
        def _status(r):
            if not r.true_code:
                return "NO_CODE"                    # a PDF section with no aligned green section
            if r.roster_overlap == 0:
                return "NO_SECTION"                 # green section matched no PDF -> enumerator junk
            if r.roster_overlap >= 0.5:
                return "CLEAN"                       # derived code well-supported by the roster
            if r.f4u_claimant and A._canon(r.f4u_claimant) == A._canon(r.true_code):
                return "CLEAN"
            return "LOW_CONF"                        # weak roster + fingerprint -> needs an eye
        df["status"] = [_status(r) for r in df.itertuples()]
    return df, pdf_secs, green_secs, un_g, un_p, testmaps, n_groups


def group_map(pdf_secs, n_groups, df, testmaps=None):
    """The `--test_map` payload: Group_N -> true code.

    Keyed on **Group_N <-> PDF section**, both of which are complete and ordered, NOT on the Green's
    yield sections (the Green enumerator can miss a section -- 1970 loses PT-00's yield table while its
    tp2 group still exists, so keying on the Green would emit a map with a HOLE at Group_2, which
    `combine_nust_outputs` turns into an orphan that `parse_test_code` silently drops).
    The Green is used to VERIFY: every green section that aligned must agree with the positional map.

    Returns (mapping|None, reason). n_groups != n_pdf means the tp2 markers are wrong -> a JSON cannot
    express it (one Group cannot carry two codes) -> the year needs a tp2 repair first.
    """
    codes = [p["code"] for p in pdf_secs]
    if n_groups != len(codes):
        d = len(codes) - n_groups
        return None, (f"tp2 DEFICIT: {n_groups} Group_N vs {len(codes)} PDF tests ({d:+d}) -> "
                      f"{'a Group carries >1 test; insert tp2' if d > 0 else 'spurious tp2'} "
                      f"BEFORE a map can be authored")
    gmap = {i + 1: c for i, c in enumerate(codes)}
    # If the PDF sequence canonically EQUALS TEST_MAPS at matching length, the positional map is
    # already confirmed correct -- the Green-section->group veto is then pure enumerator noise and must
    # not block (1941/1948/1958/1964 all have PDF==emitted==TEST_MAPS but the Green enumerator's
    # section->group_n mapping is ragged). A real scramble shows up as PDF != TEST_MAPS, where the veto
    # still matters.
    if testmaps and len(testmaps) == len(codes) and all(
            A._canon(a) == A._canon(b) for a, b in zip(codes, testmaps)):
        return gmap, "ok (PDF==TEST_MAPS; green veto skipped)"
    bad = []
    for r in df.itertuples():
        # Only a section with real evidence may veto. A green section that matched NO pdf section
        # (true_code NaN) is a parse artefact -- 1966's sec11 matches nothing at all (roster overlap
        # 0.000, best F4U yield 0.048) yet vetoed an otherwise sound map. NB `bool(float('nan'))` is
        # True, so the emptiness test must be pd.notna(), not a bare truthiness check.
        if not pd.notna(r.group_n) or not pd.notna(r.true_code) or not str(r.true_code).strip():
            continue
        if gmap.get(int(r.group_n)) != r.true_code:
            bad.append(f"Group_{int(r.group_n)}={gmap.get(int(r.group_n))} but green sec"
                       f"{r.section_order} derives {r.true_code}")
    if bad:
        return None, "Green disagrees with the positional Group<->PDF map: " + "; ".join(bad[:4])
    return gmap, "ok"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("years", nargs="+", type=int)
    ap.add_argument("--write", action="store_true", help="write the verified CSV + staged JSONs")
    args = ap.parse_args()
    sys.stdout.reconfigure(encoding="utf-8")

    allrows, ok_years = [], []
    for y in args.years:
        df, pdf_secs, green_secs, un_g, un_p, tm, n_groups = derive(y)
        codes = [p["code"] for p in pdf_secs]
        print(f"\n{'='*100}\n{y}:  PDF {len(pdf_secs)} sections | Green {len(green_secs)} | "
              f"tp2 Groups {n_groups} | TEST_MAPS {len(tm)}")
        print(f"   PDF sequence : {codes}")
        print(f"   TEST_MAPS    : {tm}")
        if codes != sorted(codes, key=sort_key):
            print("   ! PDF sequence is NOT in canonical order")
        if un_p:
            print(f"   ! PDF sections with NO Green match: {[pdf_secs[j]['code'] for j in un_p]}")
        if un_g:
            print(f"   ! Green sections with NO PDF match: {[green_secs[i]['order'] for i in un_g]}")
        if not df.empty:
            print(df[["section_order", "group_n", "true_code", "roster_overlap", "geom", "nstrain",
                      "nloc", "f4u_claimant", "yield_match", "status", "testmaps_code",
                      "agrees_with_testmaps"]].to_string(index=False))
        gmap, why = group_map(pdf_secs, n_groups, df, testmaps=tm)
        if gmap is None:
            print(f"   ! NO JSON — {why}")
            gcount = defaultdict(list)
            for r in df.itertuples():
                if pd.notna(r.group_n):        # NB bool(float('nan')) is True -- must use notna()
                    gcount[int(r.group_n)].append(r.true_code)
            for g, c in sorted(gcount.items()):
                if len(c) > 1:
                    print(f"       Group_{g} carries {c}  -> insert a tp2 before '{c[1]}'")
        else:
            print(f"   JSON ready: {len(gmap)} groups")
            ok_years.append((y, gmap))
        allrows.append(df)

    if args.write:
        out = pd.concat(allrows, ignore_index=True)
        OUT_CSV.parent.mkdir(parents=True, exist_ok=True)
        if OUT_CSV.exists():
            prev = pd.read_csv(OUT_CSV)
            prev = prev[~prev.Year.isin(out.Year.unique())]
            out = pd.concat([prev, out], ignore_index=True).sort_values(["Year", "section_order"])
        out.to_csv(OUT_CSV, index=False)
        print(f"\nwrote {OUT_CSV.relative_to(REPO)}  ({len(out)} rows)")
        OUT_JSON_DIR.mkdir(parents=True, exist_ok=True)
        for y, gmap in ok_years:
            p = OUT_JSON_DIR / f"{y}_test_map.json"
            p.write_text(json.dumps(
                {"year": y, "source": "derive_test_map.py (PDF captions + Green rosters)",
                 "groups": [{"group_number": g, "test_code": c} for g, c in sorted(gmap.items())]},
                indent=2))
            print(f"  staged {p.relative_to(REPO)}  ({len(gmap)} groups)")
        skipped = [y for y in args.years if y not in [a for a, _ in ok_years]]
        if skipped:
            print(f"  NO JSON for: {skipped} (see flags above)")


if __name__ == "__main__":
    main()
