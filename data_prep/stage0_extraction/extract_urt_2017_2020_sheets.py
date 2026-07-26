"""
extract_urt_2017_2020_sheets.py
===============================
The 2017-2020 URT phenotype data is delivered as combined "UTPT Report" workbooks
(one sheet per trial) rather than the per-trial CSV files used for 1989-2016. This
splits each trial sheet out to a CSV named like the rest, into
  R:\\...\\NUST_project_1989_2020\\URT Phenotype Data\\Years\\<year>\\<Trial>.csv
so the Format-A maturity parser (script 92) can process 2017-2020 uniformly.

- Conventional / combined workbooks -> <Sheet>.csv
- Dedicated "Traited" workbooks      -> <Sheet>TM.csv (unless already ...TM)
- Excel date cells (the anchor row)  -> "M/D" text (matches the rest; parseable by to_doy)
- Only trial sheets are exported (name starts UT/PT OR sheet has a 'MATURITY (date)' cell)
- ADDITIVE: never overwrites an existing CSV (skips + warns).

Usage:
    PYTHONUTF8=1 uv run python data_prep/stage0_extraction/extract_urt_2017_2020_sheets.py
"""
import csv
import datetime
import re
import sys
from pathlib import Path

import openpyxl

sys.stdout.reconfigure(encoding="utf-8")

URT = Path(r"R:\cfans_agro_lore0149_lorenzlabresearch\NUST_project_1989_2020\URT Phenotype Data")

# (year, workbook relative path, is_traited)
WORKBOOKS = [
    (2017, "URT_2017_Phenotype_Data/2017 UTPT Report - Conventional 2.13.18.xlsx", False),
    (2018, "URT_2018_Phenotype_Data/UTPT Report - Conventional 2018 working.xlsx", False),
    (2018, "URT_2018_Phenotype_Data/UTPT 2018 Report - Traited Material.xlsx", True),
    (2019, "URT_2019_Phenotype_Data/UTPT_Report_Conventional_Traited2019_working.xlsx", False),
    (2020, "URT_2020_Phenotype_Data/2020 UTPT Report - Conventional - 02012021 Final.xlsx", False),
    (2020, "URT_2020_Phenotype_Data/2020 UTPT Report - Traited Material - 02022021 Final.xlsx", True),
]

TRIAL_NAME = re.compile(r"^(UT|PT)[0-9IVABR]*$", re.IGNORECASE)


def fmt(v):
    if v is None:
        return ""
    if isinstance(v, (datetime.datetime, datetime.date)):
        return f"{v.month}/{v.day}"
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v)


def sheet_to_rows(ws):
    return [[fmt(c.value) for c in row] for row in ws.iter_rows()]


def is_trial_sheet(name, rows):
    if TRIAL_NAME.match(name.strip().replace(" ", "")):
        return True
    return any(str(c).strip().upper() == "MATURITY (DATE)" for r in rows for c in r)


def main():
    written = skipped = 0
    for year, rel, traited in WORKBOOKS:
        wb_path = URT / rel
        if not wb_path.exists():
            print(f"{year}: workbook missing — {rel}")
            continue
        out_dir = URT / "Years" / str(year)
        out_dir.mkdir(parents=True, exist_ok=True)
        print(f"\n=== {year} {'[Traited]' if traited else '[Conventional]'} {wb_path.name} ===")
        wb = openpyxl.load_workbook(wb_path, read_only=True, data_only=True)
        for sn in wb.sheetnames:
            rows = sheet_to_rows(wb[sn])
            if not is_trial_sheet(sn, rows):
                print(f"  skip non-trial sheet {sn!r}")
                continue
            trial = sn.strip().replace(" ", "").upper()
            if traited and not trial.endswith("TM"):
                trial += "TM"
            out = out_dir / f"{trial}.csv"
            if out.exists():
                print(f"  EXISTS (skip) {out.name}")
                skipped += 1
                continue
            with open(out, "w", newline="", encoding="utf-8") as f:
                csv.writer(f).writerows(rows)
            mi = next((i for i, r in enumerate(rows)
                       if any(c.strip().upper() == "MATURITY (DATE)" for c in r)), None)
            print(f"  wrote {out.name:14s} rows={len(rows):4d} MATURITY@{mi}")
            written += 1
        wb.close()
    print(f"\nDone. wrote {written} CSVs, skipped {skipped} existing.")


if __name__ == "__main__":
    main()
